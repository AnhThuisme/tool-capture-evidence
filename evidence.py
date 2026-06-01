from __future__ import annotations

import os
import time
import threading
import shutil
import json
import base64
import re
import signal
import calendar
import io
import sys
import socket
import csv
import html as html_lib
import zipfile
from xml.sax.saxutils import escape as xml_escape
from concurrent.futures import ThreadPoolExecutor, as_completed
try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except Exception:
    tk = None
    ttk = None
    messagebox = None
    filedialog = None
from datetime import datetime
import unicodedata
import subprocess
import difflib
from urllib.parse import quote, urlparse, parse_qs, urljoin
try:
    import requests
except Exception:
    requests = None

import gspread
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.driver_cache import DriverCacheManager
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

# ================= CONFIG =================
BASE_DIR = os.environ.get("EVIDENCE_BASE_DIR", os.path.dirname(os.path.abspath(__file__)))
APP_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR
TEMP_ROOT = os.path.join(os.environ.get("LOCALAPPDATA", BASE_DIR), "ToolEvidence")
TEMP_DIR = os.path.join(os.environ.get("EVIDENCE_TEMP_DIR", TEMP_ROOT), "temp_screenshots")
FB_PROFILE_PATH = os.path.join(BASE_DIR, "FB_Session")
FB_PROFILE_PATH_ALT = os.path.join(BASE_DIR, "FB_Session_Selenium")
LOCAL_PROFILE_PATH = os.path.join(os.environ.get("LOCALAPPDATA", os.path.join(BASE_DIR, ".local_profile")), "EvidenceTool_Profile")
WDM_ROOT = os.path.join(os.environ.get("LOCALAPPDATA", BASE_DIR), "EvidenceTool_WDM")
LOG_PATH = os.path.join(BASE_DIR, "log.txt")
SETTINGS_PATH = os.path.join(BASE_DIR, "app_settings.json")
ERROR_HISTORY_PATH = os.path.join(BASE_DIR, "error_history.json")

DEFAULT_DRIVE_FOLDER_ID = "1JJuG1ja80ThO_V14XnkOlwBz9Ey-3kmn"
DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1wKLirm10BTEhkfVHZJxeo5iR4fjCpesaylUPYlF2UV0"
DEFAULT_SHEET_NAME_TARGET = "Nghiệm thu"
CAPTURE_WINDOW_SIZE = "1920,1400"
CAPTURE_ZOOM_PERCENT = 90
PAGE_READY_TIMEOUT = 3
PAGE_READY_FALLBACK_SLEEP = 0.45
PER_LINK_BASE_WAIT = 0.22
TIKTOK_SCROLL_WAIT_1 = 0.35
TIKTOK_SCROLL_WAIT_2 = 0.5
ZOOM_SETTLE_SLEEP = 0.08
SCREENSHOT_CAPTURE_DELAY = 1.0
# Extra buffer for TikTok before first screenshot to let video/player UI settle.
TIKTOK_FIRST_CAPTURE_EXTRA_SEC = 1.0
TIKTOK_CAPTCHA_MAX_WAIT_SEC = 15.0
TIKTOK_CAPTCHA_POLL_SEC = 1.0
TIKTOK_CAPTCHA_POST_CLEAR_WAIT_SEC = 0.7
TIKTOK_BRING_TO_FRONT_INTERVAL_SEC = 8.0
TIKTOK_CAPTCHA_FORCE_FOCUS = True
TIKTOK_ACCESS_DENIED_RETRY_MAX = 4
TIKTOK_ACCESS_DENIED_RETRY_SLEEP_SEC = 1.6
TIKTOK_REDIRECT_WAIT_SEC = 4.0
PLEASE_WAIT_EXTRA_CAPTURE_DELAY_SEC = 2.0
PLEASE_WAIT_MAX_WAIT_SEC = 8.0
PLEASE_WAIT_POLL_SEC = 1.0
BLANK_SCREEN_RETRY_DELAY_SEC = 2.0
BLANK_SCREEN_MAX_RETRIES = 1
MULTI_CAPTURE_INTERVAL_SEC = 5.0
FB_COMMENT_READY_WAIT = 4.0
UI_CLICK_SETTLE_SLEEP = 0.15
UI_SCROLL_SETTLE_SLEEP = 0.1
TIKTOK_OEMBED_TIMEOUT_SEC = 10.0


def _normalize_profile_dir(path_value: str | None) -> str:
    raw = str(path_value or "").strip()
    if not raw:
        raw = LOCAL_PROFILE_PATH
    if not os.path.isabs(raw):
        raw = os.path.join(BASE_DIR, raw)
    return os.path.abspath(raw)


def _fallback_profile_dir(browser_port: int = 9223) -> str:
    if os.name == "nt":
        root = os.path.join(os.environ.get("LOCALAPPDATA", BASE_DIR), "ToolEvidence")
    elif sys.platform == "darwin":
        root = os.path.join(os.path.expanduser("~"), "Library", "Application Support", "ToolEvidence")
    else:
        root = os.path.join(os.path.expanduser("~"), ".local", "share", "ToolEvidence")
    return os.path.abspath(os.path.join(root, "profiles", f"chrome_port_{int(browser_port)}"))


def _ensure_profile_dir_writable(profile_dir: str) -> tuple[bool, str]:
    try:
        os.makedirs(profile_dir, exist_ok=True)
        probe = os.path.join(profile_dir, ".evidence_profile_write_test")
        with open(probe, "w", encoding="utf-8") as f:
            f.write("ok")
        try:
            os.remove(probe)
        except Exception:
            pass
        return True, ""
    except Exception as exc:
        return False, str(exc)


def _resolve_writable_profile_dir(
    profile_path: str | None,
    *,
    browser_port: int = 9223,
    log_prefix: str = "",
) -> str:
    profile = _normalize_profile_dir(profile_path)
    ok, reason = _ensure_profile_dir_writable(profile)
    if ok:
        return profile
    fallback = _fallback_profile_dir(browser_port)
    ok_fb, reason_fb = _ensure_profile_dir_writable(fallback)
    if ok_fb:
        write_log(
            f"[WARN] {log_prefix}Profile dir not writable: '{profile}' ({reason}). "
            f"Use fallback profile: '{fallback}'."
        )
        return fallback
    raise RuntimeError(
        f"{log_prefix}Không ghi được thư mục profile Chrome.\n"
        f"- Primary: {profile} ({reason})\n"
        f"- Fallback: {fallback} ({reason_fb})"
    )


def get_post_port(post_index: int, base_port: int = 9223) -> int:
    """
    Post 1 -> 9223, Post 2 -> 9324, Post 3 -> 9325, ...
    Keep compatible with existing worker/profile mapping.
    """
    if post_index <= 0:
        return base_port
    return base_port + 100 + post_index


def get_block_profile(block_index: int, mode: str = "seeding", browser_port: int | None = None) -> str:
    mode_name = str(mode or "seeding").strip().lower() or "seeding"
    idx = max(0, int(block_index or 0))
    if browser_port:
        try:
            port = int(browser_port)
            if port > 0:
                if mode_name == "seeding" and port == 9223:
                    return _normalize_profile_dir(LOCAL_PROFILE_PATH)
                return _normalize_profile_dir(os.path.join(TEMP_DIR, f"chrome_profile_{mode_name}_port_{port}"))
        except Exception:
            pass
    if mode_name == "seeding":
        if idx <= 0:
            return _normalize_profile_dir(LOCAL_PROFILE_PATH)
        return _normalize_profile_dir(os.path.join(TEMP_DIR, f"chrome_profile_worker_{idx}"))
    suffix = f"{mode_name}_{idx}" if idx > 0 else f"{mode_name}_main"
    return _normalize_profile_dir(os.path.join(TEMP_DIR, f"chrome_profile_{suffix}"))


_BROWSER_BINARY_CHECK_CACHE: dict[str, tuple[bool, str]] = {}
_BROWSER_BINARY_CHECK_LOCK = threading.Lock()


def _iter_browser_binary_candidates() -> list[str]:
    candidates: list[str] = []
    env_overrides = [
        os.environ.get("EVIDENCE_CHROME_BINARY", "").strip(),
        os.environ.get("EVIDENCE_BROWSER_BINARY", "").strip(),
    ]
    for item in env_overrides:
        if item:
            candidates.append(item)

    if os.name == "nt":
        candidates.extend(
            [
                os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "Google", "Chrome", "Application", "chrome.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "Google", "Chrome", "Application", "chrome.exe"),
                os.path.join(os.environ.get("LOCALAPPDATA", ""), "Google", "Chrome", "Application", "chrome.exe"),
                os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "Chromium", "Application", "chrome.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "Chromium", "Application", "chrome.exe"),
                os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "Microsoft", "Edge", "Application", "msedge.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "Microsoft", "Edge", "Application", "msedge.exe"),
                os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
            ]
        )
    elif sys.platform == "darwin":
        candidates.extend(
            [
                "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                "/Applications/Chromium.app/Contents/MacOS/Chromium",
                "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
                "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser",
                "/Applications/Arc.app/Contents/MacOS/Arc",
                "/Applications/Vivaldi.app/Contents/MacOS/Vivaldi",
                "/Applications/Opera.app/Contents/MacOS/Opera",
            ]
        )
    else:
        for cmd in [
            "google-chrome",
            "google-chrome-stable",
            "chromium",
            "chromium-browser",
            "microsoft-edge",
            "microsoft-edge-stable",
            "brave-browser",
            "vivaldi",
            "opera",
        ]:
            p = shutil.which(cmd)
            if p:
                candidates.append(p)

    for cmd in [
        "chrome",
        "chrome.exe",
        "google-chrome",
        "chromium",
        "msedge",
        "msedge.exe",
        "brave",
        "brave-browser",
        "vivaldi",
        "opera",
    ]:
        p = shutil.which(cmd)
        if p:
            candidates.append(p)

    seen: set[str] = set()
    ordered: list[str] = []
    for p in candidates:
        norm = os.path.normpath(str(p or "").strip())
        if not norm or norm in seen:
            continue
        seen.add(norm)
        ordered.append(norm)
    return ordered


def _check_browser_binary_runtime(path: str) -> tuple[bool, str]:
    normalized = os.path.normpath(str(path or "").strip())
    if not normalized:
        return False, "empty path"
    with _BROWSER_BINARY_CHECK_LOCK:
        cached = _BROWSER_BINARY_CHECK_CACHE.get(normalized)
    if cached is not None:
        return cached
    if not os.path.exists(normalized):
        result = (False, "not found")
        with _BROWSER_BINARY_CHECK_LOCK:
            _BROWSER_BINARY_CHECK_CACHE[normalized] = result
        return result
    if sys.platform != "darwin":
        result = (True, "")
        with _BROWSER_BINARY_CHECK_LOCK:
            _BROWSER_BINARY_CHECK_CACHE[normalized] = result
        return result
    try:
        probe = subprocess.run(
            [normalized, "--version"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=4,
            check=False,
        )
        combined = "\n".join([str(probe.stdout or "").strip(), str(probe.stderr or "").strip()]).strip()
        lowered = combined.lower()
        if ("or later required" in lowered) and ("have instead" in lowered):
            reason = combined.splitlines()[0].strip() if combined else "binary requires newer macOS"
            result = (False, reason)
        else:
            result = (True, "")
    except Exception as exc:
        result = (True, f"probe skipped: {exc}")
    with _BROWSER_BINARY_CHECK_LOCK:
        _BROWSER_BINARY_CHECK_CACHE[normalized] = result
    return result


def find_compatible_browser_binary() -> tuple[str | None, list[str]]:
    skipped: list[str] = []
    for path in _iter_browser_binary_candidates():
        ok, reason = _check_browser_binary_runtime(path)
        if ok:
            return path, skipped
        skipped.append(f"{path} ({reason})")
    return None, skipped


def _bootstrap_env_credentials_path() -> str:
    raw = os.environ.get("GOOGLE_CREDENTIALS_JSON_B64", "").strip()
    if not raw:
        return ""
    # Vercel filesystem is read-only except /tmp.
    if str(os.environ.get("VERCEL", "")).strip():
        target = os.path.join("/tmp", "tool-evidence", "credentials.env.json")
    else:
        target = os.path.join(BASE_DIR, "credentials.env.json")
    try:
        if raw.startswith("{"):
            data = json.loads(raw)
        else:
            padded = raw + ("=" * (-len(raw) % 4))
            decoded = base64.b64decode(padded.encode("utf-8")).decode("utf-8")
            data = json.loads(decoded)
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with open(target, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return target
    except Exception as exc:
        print(f"[startup-config] failed to materialize GOOGLE_CREDENTIALS_JSON_B64: {exc}")
        return ""


def is_fixed_credentials_path(path: str | None) -> bool:
    raw = str(path or "").strip()
    if not raw:
        return False
    return os.path.basename(raw).lower() in {"credentials.inline.json", "credentials.env.json"}


def resolve_credentials_path() -> str:
    env_path = os.environ.get("GOOGLE_CREDENTIALS_PATH", "").strip()
    if env_path:
        norm_env_path = os.path.normpath(env_path)
        if os.path.exists(norm_env_path):
            return norm_env_path

    env_b64_path = _bootstrap_env_credentials_path()
    if env_b64_path:
        return env_b64_path

    candidates = [
        os.path.join(APP_DIR, "credentials.inline.json"),  # saved once from web UI / committed fixed file
        os.path.join(APP_DIR, "credentials.json"),     # next to .exe / script
        os.path.join(os.getcwd(), "credentials.inline.json"),
        os.path.join(os.getcwd(), "credentials.json"), # current working directory
        os.path.join(BASE_DIR, "credentials.inline.json"),
        os.path.join(BASE_DIR, "credentials.json"),    # source directory
    ]
    if hasattr(sys, "_MEIPASS"):
        candidates.append(os.path.join(getattr(sys, "_MEIPASS"), "credentials.inline.json"))
        candidates.append(os.path.join(getattr(sys, "_MEIPASS"), "credentials.json"))

    for p in candidates:
        if os.path.exists(p):
            return p

    # default location for error messages when file is missing
    return candidates[0]


JSON_PATH = resolve_credentials_path()


def normalize_sheet_input(sheet_text: str) -> str:
    s = (sheet_text or "").strip()
    if not s:
        return ""
    if "docs.google.com/spreadsheets/" in s:
        return s
    # Accept raw spreadsheet id and normalize to URL.
    if len(s) >= 20 and "/" not in s and " " not in s:
        return f"https://docs.google.com/spreadsheets/d/{s}"
    return s


def extract_sheet_gid(sheet_text: str) -> int | None:
    s = normalize_sheet_input(sheet_text or "")
    if not s:
        return None
    try:
        parsed = urlparse(s)
        query = parse_qs(parsed.query or "")
        raw_gid = (query.get("gid") or [""])[0].strip()
        if not raw_gid and parsed.fragment:
            fragment_query = parse_qs(parsed.fragment or "")
            raw_gid = (fragment_query.get("gid") or [""])[0].strip()
            if not raw_gid and parsed.fragment.isdigit():
                raw_gid = parsed.fragment.strip()
        if not raw_gid:
            return None
        gid = int(raw_gid)
        return gid if gid >= 0 else None
    except Exception:
        return None


def resolve_worksheet(spreadsheet, sheet_name: str = "", sheet_url: str = ""):
    raw_name = str(sheet_name or "").strip()
    if raw_name:
        return spreadsheet.worksheet(raw_name)
    gid = extract_sheet_gid(sheet_url)
    if gid is not None:
        try:
            for ws in spreadsheet.worksheets():
                if int(getattr(ws, "id", -1) or -1) == gid:
                    return ws
        except Exception:
            pass
    raise Exception("Thiếu Sheet Name")


def normalize_drive_folder_input(folder_text: str) -> str:
    s = (folder_text or "").strip()
    if not s:
        return ""
    if "drive.google.com" not in s:
        return s
    try:
        parsed = urlparse(s)
        parts = [p for p in (parsed.path or "").split("/") if p]
        if "folders" in parts:
            i = parts.index("folders")
            if i + 1 < len(parts):
                cand = parts[i + 1].strip()
                if cand:
                    return cand
        q = parse_qs(parsed.query or "")
        cand = (q.get("id") or [""])[0].strip()
        if cand:
            return cand
    except Exception:
        pass
    return s


def get_default_credentials_input() -> str:
    try:
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False)
    except Exception:
        return JSON_PATH


def resolve_chromedriver_service() -> Service:
    """
    Resolve chromedriver with a writable cache dir to avoid WinError 5 on locked home dirs.
    """
    local_driver = shutil.which("chromedriver")
    if local_driver:
        write_log(f"[INFO] Use local chromedriver: {local_driver}")
        return Service(local_driver)

    wdm_root = os.environ.get("EVIDENCE_WDM_DIR", "").strip() or WDM_ROOT
    os.makedirs(wdm_root, exist_ok=True)
    write_log(f"[INFO] WebDriver cache dir: {wdm_root}")

    cache_manager = DriverCacheManager(root_dir=wdm_root)
    driver_path = ChromeDriverManager(cache_manager=cache_manager).install()
    write_log(f"[INFO] WebDriver installed: {driver_path}")
    return Service(driver_path)


def create_chrome_driver(options: Options, service: Service | None = None):
    """
    Try an explicit chromedriver service first, then fall back to Selenium Manager.
    This makes local runs more resilient when webdriver-manager cached binaries drift
    from the installed Chrome version.
    """
    errors = []
    if service is not None:
        try:
            return webdriver.Chrome(service=service, options=options)
        except Exception as exc:
            errors.append(f"service={exc}")
            write_log(f"[WARN] Chrome via explicit service failed, fallback to Selenium Manager: {exc}")
    try:
        return webdriver.Chrome(options=options)
    except Exception as exc:
        errors.append(f"selenium_manager={exc}")
        raise Exception(" | ".join(errors) if errors else str(exc)) from exc

# ================= HELPERS =================
def get_service_account_email(path: str | None = None):
    """Đọc email service account từ credentials.json để hướng dẫn user chia sẻ Sheet/Drive."""
    try:
        import json
        cred_path = path or JSON_PATH
        with open(cred_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return (data.get("client_email") or "").strip()
    except Exception:
        return ""


def col_letter_to_index(letter):
    """Convert Excel/Sheets column ref to 1-based index (A, AC, 29, ...)."""
    s = str(letter or "").strip()
    if not s:
        return None

    # Allow numeric column input directly (e.g. "29").
    if s.isdigit():
        idx = int(s)
        return idx if idx > 0 else None

    s = s.upper()
    idx = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            return None
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def col_index_to_letter(index: int) -> str:
    """Convert 1-based column index to letter (1->A, 27->AA)."""
    try:
        n = int(index)
    except Exception:
        return ""
    if n <= 0:
        return ""
    out = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out.append(chr(ord("A") + rem))
    return "".join(reversed(out))


def extract_url_from_hyperlink_formula(formula_text: str) -> str:
    """
    Extract URL from Sheets formula:
    =HYPERLINK("https://...","label")
    """
    s = str(formula_text or "").strip()
    if not s:
        return ""
    m = re.search(r'^\s*=\s*HYPERLINK\s*\(\s*"((?:[^"]|"")*)"', s, flags=re.IGNORECASE)
    if not m:
        return ""
    return m.group(1).replace('""', '"').strip()


def resolve_links_for_scan(worksheet, col_idx: int, start_row: int = 4, total_rows: int | None = None) -> list[str]:
    """
    Build effective link list for scan mode.
    If displayed cell value is not an URL, try to read URL from HYPERLINK formula.
    """
    if not col_idx:
        return []

    col_letter = col_index_to_letter(col_idx)
    display_slice: list[str] = []
    if total_rows is not None and total_rows > 0 and col_letter:
        end_row = start_row + total_rows - 1
        try:
            display_rows = worksheet.get(
                f"{col_letter}{start_row}:{col_letter}{end_row}",
                value_render_option="UNFORMATTED_VALUE",
            ) or []
        except Exception:
            display_rows = []
        for r in display_rows:
            if r and len(r) > 0:
                display_slice.append(str(r[0]).strip())
            else:
                display_slice.append("")
        if len(display_slice) < total_rows:
            display_slice.extend([""] * (total_rows - len(display_slice)))
    else:
        display_vals = worksheet.col_values(col_idx)
        display_slice = display_vals[start_row - 1 :] if len(display_vals) >= start_row else []

    formula_rows = []
    try:
        if col_letter:
            if total_rows is not None and total_rows > 0:
                end_row = start_row + total_rows - 1
                formula_rows = worksheet.get(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    value_render_option="FORMULA",
                ) or []
            else:
                formula_rows = worksheet.get(
                    f"{col_letter}{start_row}:{col_letter}",
                    value_render_option="FORMULA",
                ) or []
    except Exception as e:
        write_log(f"[WARN] resolve_links_for_scan formulas read failed: {e}")

    size = max(int(total_rows or 0), len(display_slice), len(formula_rows))
    out: list[str] = []
    for i in range(size):
        shown = str(display_slice[i]).strip() if i < len(display_slice) else ""
        shown_norm = normalize_scan_source_url(shown)
        if shown_norm:
            out.append(shown_norm)
            continue
        formula_cell = ""
        if i < len(formula_rows) and formula_rows[i]:
            formula_cell = str(formula_rows[i][0]).strip()
        parsed = extract_url_from_hyperlink_formula(formula_cell)
        out.append(normalize_scan_source_url(parsed or shown))
    return out


def resolve_links_for_scan_values(display_slice: list[str] | tuple[str, ...], formula_rows: list | tuple | None = None) -> list[str]:
    """
    Build effective link list for scan mode from preloaded display/formula values.
    """
    formula_rows = list(formula_rows or [])
    size = max(len(display_slice or []), len(formula_rows))
    out: list[str] = []
    for i in range(size):
        shown = str(display_slice[i]).strip() if i < len(display_slice or []) else ""
        shown_norm = normalize_scan_source_url(shown)
        if shown_norm:
            out.append(shown_norm)
            continue
        formula_cell = ""
        if i < len(formula_rows) and formula_rows[i]:
            formula_cell = str(formula_rows[i][0]).strip()
        parsed = extract_url_from_hyperlink_formula(formula_cell)
        out.append(normalize_scan_source_url(parsed or shown))
    return out


def filldown_scan_links_for_merged_rows(
    links: list[str],
    expected_texts: list[str] | None = None,
    result_values: list[str] | None = None,
) -> list[str]:
    """
    When a scan link cell is merged across multiple rows, Sheets only keeps the URL
    on the first visible row. Carry that URL downward for subsequent rows that still
    have row-level payload (expected text/result), so each row is checked separately.
    """
    expected_texts = list(expected_texts or [])
    result_values = list(result_values or [])
    size = max(len(links), len(expected_texts), len(result_values))
    out: list[str] = []
    last_link = ""
    for i in range(size):
        current_link = normalize_scan_source_url(links[i] if i < len(links) else "")
        if current_link:
            last_link = current_link
            out.append(current_link)
            continue
        expected = str(expected_texts[i]).strip() if i < len(expected_texts) else ""
        result = str(result_values[i]).strip() if i < len(result_values) else ""
        has_row_payload = bool(expected or result)
        out.append(last_link if (last_link and has_row_payload) else "")
    return out


def resolve_column_values_aligned(worksheet, col_idx: int, start_row: int = 4, total_rows: int | None = None) -> list[str]:
    """
    Read a column while preserving row alignment and blanks.
    """
    if not col_idx:
        return []
    col_letter = col_index_to_letter(col_idx)
    if not col_letter:
        return []
    if total_rows is not None and total_rows > 0:
        end_row = start_row + total_rows - 1
        rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    else:
        rng = f"{col_letter}{start_row}:{col_letter}"
    try:
        rows = worksheet.get(rng, value_render_option="UNFORMATTED_VALUE") or []
    except Exception as e:
        write_log(f"[WARN] resolve_column_values_aligned failed: {e}")
        rows = []
    out: list[str] = []
    for r in rows:
        if r and len(r) > 0:
            out.append(str(r[0]).strip())
        else:
            out.append("")
    if total_rows is not None and total_rows > 0 and len(out) < total_rows:
        out.extend([""] * (total_rows - len(out)))
    return out


def normalize_match_text(text: str) -> str:
    s = unicodedata.normalize("NFD", str(text or ""))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_scan_terms_text(value: str) -> str:
    lines: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[\r\n,;]+", str(value or "")):
        term = normalize_match_text(part)
        if not term or term in seen:
            continue
        seen.add(term)
        lines.append(term)
    return "\n".join(lines)


def _parse_scan_terms(value: str | list[str] | tuple[str, ...] | set[str] | None) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        joined = "\n".join(str(item or "") for item in value)
    else:
        joined = str(value or "")
    normalized = _normalize_scan_terms_text(joined)
    return [line for line in normalized.splitlines() if line.strip()]


def detect_scan_negative_term(*texts: str, keywords: list[str] | tuple[str, ...] | set[str] | None = None) -> str:
    normalized_terms = _parse_scan_terms(list(keywords or []))
    if not normalized_terms:
        return ""
    source_parts: list[str] = []
    token_set: set[str] = set()
    for raw_text in texts:
        normalized_text = normalize_match_text(raw_text)
        if not normalized_text:
            continue
        source_parts.append(normalized_text)
        token_set.update(token for token in normalized_text.split() if token)
    if not source_parts:
        return ""
    haystack = f" {' '.join(source_parts)} "
    for term in normalized_terms:
        if " " in term:
            if f" {term} " in haystack:
                return term
        elif term in token_set:
            return term
    return ""


def detect_scan_keyword_term(*texts: str, keywords: list[str] | tuple[str, ...] | set[str] | None = None) -> str:
    normalized_terms = _parse_scan_terms(list(keywords or []))
    if not normalized_terms:
        return ""
    source_parts: list[str] = []
    token_set: set[str] = set()
    for raw_text in texts:
        normalized_text = normalize_match_text(raw_text)
        if not normalized_text:
            continue
        source_parts.append(normalized_text)
        token_set.update(token for token in normalized_text.split() if token)
    if not source_parts:
        return ""
    haystack = f" {' '.join(source_parts)} "
    for term in normalized_terms:
        if " " in term:
            if f" {term} " in haystack:
                return term
        elif term in token_set:
            return term
    return ""


def extract_drive_file_id(url: str) -> str:
    s = str(url or "").strip()
    if not s:
        return ""
    try:
        if "/file/d/" in s:
            return s.split("/file/d/", 1)[1].split("/", 1)[0].strip()
        parsed = urlparse(s)
        q = parse_qs(parsed.query or "")
        cand = (q.get("id") or [""])[0].strip()
        if cand:
            return cand
    except Exception:
        pass
    return ""


def normalize_scan_source_url(raw_url: str) -> str:
    s = str(raw_url or "").strip()
    if not s:
        return ""
    if s.lower().startswith("http://") or s.lower().startswith("https://"):
        return s
    if "drive.google.com" in s:
        return "https://" + s.lstrip("/")
    fid = extract_drive_file_id(s)
    if fid:
        return f"https://drive.google.com/file/d/{fid}/view"
    return ""


def normalize_web_source_url(raw_url: str) -> str:
    s = str(raw_url or "").strip()
    if not s:
        return ""
    low = s.lower()
    if low.startswith("http://") or low.startswith("https://"):
        return s
    if "://" in s:
        return s
    if "drive.google.com" in low:
        return "https://" + s.lstrip("/")
    if re.match(r"^[a-z0-9][a-z0-9.-]*\.[a-z]{2,}(/.*)?$", low):
        return "https://" + s
    return ""


def build_candidate_image_urls(src_url: str) -> list[str]:
    base = str(src_url or "").strip()
    if not base:
        return []
    out = [base]
    fid = extract_drive_file_id(base)
    if fid:
        out.extend(
            [
                f"https://drive.google.com/uc?export=download&id={fid}",
                f"https://drive.google.com/uc?export=view&id={fid}",
                f"https://lh3.googleusercontent.com/d/{fid}",
            ]
        )
    # preserve order, remove duplicates
    seen = set()
    uniq = []
    for u in out:
        if u not in seen:
            seen.add(u)
            uniq.append(u)
    return uniq


def download_image_bytes_for_scan(url: str, timeout: int = 20, drive_service=None) -> bytes:
    fid = extract_drive_file_id(url)
    if drive_service is not None and fid:
        try:
            req = drive_service.files().get_media(fileId=fid, supportsAllDrives=True)
            buf = io.BytesIO()
            downloader = MediaIoBaseDownload(buf, req)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            content = buf.getvalue() or b""
            if content and (
                content.startswith(b"\x89PNG")
                or content.startswith(b"\xff\xd8")
                or content[:4] == b"RIFF"
            ):
                return content
        except Exception as e:
            write_log(f"[WARN] Drive API download failed for {fid}: {e}")

    headers = {"User-Agent": "Mozilla/5.0"}
    for cand in build_candidate_image_urls(url):
        try:
            if requests is not None:
                r = requests.get(cand, timeout=timeout, headers=headers, allow_redirects=True)
                if r.status_code >= 400:
                    continue
                ctype = (r.headers.get("content-type") or "").lower()
                content = r.content or b""
            else:
                from urllib.request import Request, urlopen
                req = Request(cand, headers=headers)
                with urlopen(req, timeout=timeout) as resp:
                    ctype = str(resp.headers.get("Content-Type", "")).lower()
                    content = resp.read() or b""
            if not content:
                continue
            # Accept typical image content-type or PNG/JPG/WebP bytes signature.
            if (
                "image/" in ctype
                or content.startswith(b"\x89PNG")
                or content.startswith(b"\xff\xd8")
                or content[:4] == b"RIFF"
            ):
                return content
        except Exception:
            continue
    return b""


def fetch_tiktok_oembed_data(url: str, timeout_sec: float = TIKTOK_OEMBED_TIMEOUT_SEC) -> dict:
    """
    Fetch TikTok public oEmbed payload for a video URL.
    Useful when normal browser navigation is blocked by Access Denied.
    """
    src = normalize_web_source_url(url)
    if not src:
        src = str(url or "").strip()
    if not src:
        return {}
    headers = {"User-Agent": "Mozilla/5.0"}

    candidates = [src]
    if "vt.tiktok.com" in src.lower():
        try:
            if requests is not None:
                r = requests.get(src, timeout=max(3, int(timeout_sec)), headers=headers, allow_redirects=True)
                final_url = str(getattr(r, "url", "") or "").strip()
            else:
                from urllib.request import Request, urlopen
                req = Request(src, headers=headers)
                with urlopen(req, timeout=max(3, float(timeout_sec or 10.0))) as resp:
                    final_url = str(getattr(resp, "url", "") or "").strip()
            if final_url and final_url not in candidates:
                candidates.insert(0, final_url)
        except Exception:
            pass

    seen = set()
    for cand in candidates:
        c = str(cand or "").strip()
        if not c or c in seen:
            continue
        seen.add(c)
        endpoint = f"https://www.tiktok.com/oembed?url={quote(c, safe='')}"
        try:
            if requests is not None:
                resp = requests.get(
                    endpoint,
                    timeout=max(3, int(timeout_sec)),
                    headers=headers,
                    allow_redirects=True,
                )
                if int(getattr(resp, "status_code", 0) or 0) >= 400:
                    continue
                payload = resp.json() if hasattr(resp, "json") else {}
            else:
                from urllib.request import Request, urlopen
                req = Request(endpoint, headers=headers)
                with urlopen(req, timeout=max(3, float(timeout_sec or 10.0))) as r2:
                    payload = json.loads((r2.read() or b"").decode("utf-8", errors="ignore") or "{}")
            if isinstance(payload, dict) and payload:
                payload["_source_url"] = c
                return payload
        except Exception:
            continue
    return {}


def ocr_text_from_image_bytes(image_bytes: bytes, expected_text: str = "") -> str:
    if not image_bytes:
        return ""
    try:
        from PIL import Image
        import pytesseract
        try:
            tcmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or "tesseract"
            if not shutil.which(str(tcmd)):
                candidates = [
                    os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs", "Tesseract-OCR", "tesseract.exe"),
                    os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "Tesseract-OCR", "tesseract.exe"),
                    os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "Tesseract-OCR", "tesseract.exe"),
                ]
                for p in candidates:
                    if p and os.path.exists(p):
                        pytesseract.pytesseract.tesseract_cmd = p
                        break
        except Exception:
            pass
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        w, h = img.size
        gray = img.convert("L")
        # Fast-first OCR plan; only fallback to slower variants if needed.
        variants = [
            gray,
            img.resize((max(1, w * 2), max(1, h * 2))).convert("L"),
            gray.point(lambda p: 255 if p > 165 else 0),
        ]
        configs = ["--oem 1 --psm 6", "--oem 1 --psm 11"]
        langs = ["vie+eng", "eng"]
        texts = []
        seen = set()
        for v in variants:
            for lang in langs:
                for cfg in configs:
                    try:
                        t = (pytesseract.image_to_string(v, lang=lang, config=cfg) or "").strip()
                        if t and t not in seen:
                            seen.add(t)
                            texts.append(t)
                            # Early stop when already matched -> much faster for positive rows.
                            if expected_text and is_scan_match(expected_text, t):
                                return t
                    except Exception:
                        continue
        if texts:
            return "\n".join(texts)
    except Exception as e:
        write_log(f"[WARN] OCR engine unavailable/failed: {e}")
    return ""


def build_collage_png(image_bytes_list: list[bytes]) -> bytes:
    """
    Merge multiple screenshots into a single collage image so Sheets can render
    all shots in one IMAGE() cell.
    """
    if not image_bytes_list:
        return b""
    try:
        from PIL import Image
    except Exception:
        return b""

    images = []
    for b in image_bytes_list:
        try:
            img = Image.open(io.BytesIO(b)).convert("RGB")
            images.append(img)
        except Exception:
            continue
    if not images:
        return b""
    if len(images) == 1:
        out_buf = io.BytesIO()
        images[0].save(out_buf, format="PNG")
        return out_buf.getvalue()

    cols = min(3, len(images))
    rows = (len(images) + cols - 1) // cols
    tile_w = 360
    tile_h = 260
    pad = 8

    thumbs = []
    for img in images:
        im = img.copy()
        im.thumbnail((tile_w, tile_h))
        thumbs.append(im)

    canvas_w = cols * tile_w + (cols + 1) * pad
    canvas_h = rows * tile_h + (rows + 1) * pad
    canvas = Image.new("RGB", (canvas_w, canvas_h), (245, 246, 250))

    for i, im in enumerate(thumbs):
        r = i // cols
        c = i % cols
        x0 = pad + c * (tile_w + pad)
        y0 = pad + r * (tile_h + pad)
        x = x0 + max(0, (tile_w - im.width) // 2)
        y = y0 + max(0, (tile_h - im.height) // 2)
        canvas.paste(im, (x, y))

    out_buf = io.BytesIO()
    canvas.save(out_buf, format="PNG")
    return out_buf.getvalue()


def is_blank_like_screenshot_png(image_bytes: bytes) -> bool:
    """
    Heuristic for blank/placeholder screenshots (white/black/near-solid canvas).
    Returns True when frame has very low visual variance or one dominant color.
    """
    if not image_bytes:
        return True
    try:
        from PIL import Image, ImageStat
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception:
        return False

    try:
        sample = img.resize((160, 90))
        stat = ImageStat.Stat(sample)
        std_vals = list(stat.stddev or [0.0, 0.0, 0.0])
        mean_std = sum(float(v) for v in std_vals) / max(1, len(std_vals))

        pixels = sample.getdata()
        total = max(1, len(pixels))
        bright = 0
        dark = 0
        for r, g, b in pixels:
            if r >= 245 and g >= 245 and b >= 245:
                bright += 1
            if r <= 14 and g <= 14 and b <= 14:
                dark += 1
        bright_ratio = bright / total
        dark_ratio = dark / total

        counts = sample.quantize(colors=8, method=2).getcolors() or []
        dominant_ratio = 0.0
        if counts:
            dominant_ratio = max(c for c, _ in counts) / total

        if bright_ratio >= 0.93 or dark_ratio >= 0.93:
            return True
        if dominant_ratio >= 0.90 and mean_std <= 12.0:
            return True
        if mean_std <= 5.0:
            return True
        return False
    except Exception:
        return False


def check_ocr_dependencies() -> tuple[bool, str]:
    try:
        from PIL import Image  # noqa: F401
    except Exception:
        return False, "Thiếu Pillow. Cài: pip install pillow"
    try:
        import pytesseract
    except Exception:
        return False, "Thiếu pytesseract. Cài: pip install pytesseract"
    try:
        tcmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or "tesseract"
        if not shutil.which(str(tcmd)):
            candidates = [
                os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs", "Tesseract-OCR", "tesseract.exe"),
                os.path.join(os.environ.get("PROGRAMFILES", r"C:\Program Files"), "Tesseract-OCR", "tesseract.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"), "Tesseract-OCR", "tesseract.exe"),
            ]
            for p in candidates:
                if p and os.path.exists(p):
                    pytesseract.pytesseract.tesseract_cmd = p
                    break
    except Exception:
        pass
    try:
        _ = pytesseract.get_tesseract_version()
    except Exception:
        return (
            False,
            "Thiếu Tesseract OCR (binary). Cài Tesseract và thêm vào PATH, rồi mở lại app.",
        )
    return True, ""


def is_scan_match(expected_text: str, ocr_text: str) -> bool:
    expected = normalize_match_text(expected_text)
    got = normalize_match_text(ocr_text)
    if not expected or not got:
        return False
    if expected in got:
        return True
    if got in expected and len(got) >= 18:
        return True
    e_tokens = [t for t in expected.split() if len(t) >= 2]
    g_tokens = set(t for t in got.split() if len(t) >= 2)
    got_pad = f" {got} "
    # Strong signal: a consecutive phrase appears in OCR text.
    if len(e_tokens) >= 3:
        for win in (5, 4, 3):
            if len(e_tokens) >= win:
                for i in range(len(e_tokens) - win + 1):
                    phrase = " " + " ".join(e_tokens[i : i + win]) + " "
                    if phrase in got_pad:
                        return True
    if e_tokens and g_tokens:
        overlap = sum(1 for t in e_tokens if t in g_tokens) / max(1, len(e_tokens))
        if overlap >= 0.42:
            return True
    ratio = difflib.SequenceMatcher(None, expected, got).ratio()
    return ratio >= 0.52


def is_scan_text_strict_match(expected_text: str, source_text: str) -> bool:
    """
    Stricter matching for Scan Only Text mode to reduce false positives.
    Compare against per-line candidate comments, not only whole-page text.
    """
    expected = normalize_match_text(expected_text)
    source = normalize_match_text(source_text)
    if not expected or not source:
        return False

    # Very short expected text is too ambiguous.
    e_tokens = [t for t in expected.split() if len(t) >= 2]
    if len(e_tokens) < 4 or len(expected) < 12:
        return False

    exp_pad = f" {expected} "
    exp_compact = expected.replace(" ", "")
    exp_len = len(expected)

    # Build candidate comment blocks from lines + sliding windows.
    noise_phrases = {
        "like", "reply", "share", "follow", "see more", "view more",
        "xem them", "xem them binh luan", "xem them phan hoi",
        "tat ca binh luan", "binh luan", "phan hoi",
    }
    raw_lines = [str(x).strip() for x in str(source_text or "").splitlines()]
    normalized_lines: list[str] = []
    for ln in raw_lines:
        n = normalize_match_text(ln)
        if len(n) < 10:
            continue
        if n in noise_phrases:
            continue
        normalized_lines.append(n)
    if not normalized_lines:
        normalized_lines = [source]

    candidates: list[str] = []
    seen_cands: set[str] = set()

    def _push_candidate(txt: str):
        t = (txt or "").strip()
        if len(t) < 10:
            return
        if t in seen_cands:
            return
        seen_cands.add(t)
        candidates.append(t)

    for n in normalized_lines:
        _push_candidate(n)

    max_window = 6
    for i in range(len(normalized_lines)):
        merged = normalized_lines[i]
        for w in range(2, max_window + 1):
            j = i + w - 1
            if j >= len(normalized_lines):
                break
            merged = f"{merged} {normalized_lines[j]}".strip()
            if len(merged) > 450:
                break
            _push_candidate(merged)

    if len(candidates) > 5000:
        candidates = candidates[:5000]

    for cand in candidates:
        if not cand:
            continue
        cand_pad = f" {cand} "

        # Primary rule: expected sentence must appear as a contiguous phrase.
        if exp_pad in cand_pad:
            return True

        # Also allow exact-compact containment (for cases where page collapses spaces).
        cand_compact = cand.replace(" ", "")
        if exp_compact and exp_compact in cand_compact:
            return True

        c_tokens = [t for t in cand.split() if len(t) >= 2]
        if not c_tokens:
            continue
        c_set = set(c_tokens)
        overlap = sum(1 for t in e_tokens if t in c_set) / max(1, len(e_tokens))
        ratio = difflib.SequenceMatcher(None, expected, cand).ratio()

        # Ordered token coverage: robust for line breaks / emoji / minor OCR-like distortions.
        j = 0
        ordered_hits = 0
        for tok in e_tokens:
            while j < len(c_tokens) and c_tokens[j] != tok:
                j += 1
            if j < len(c_tokens):
                ordered_hits += 1
                j += 1
        order_cov = ordered_hits / max(1, len(e_tokens))

        # Fuzzy fallback on one candidate only (not whole-page), to reduce false matches.
        len_gap = abs(len(cand) - exp_len) / max(1, exp_len)
        if exp_len >= 90:
            need_order, need_overlap, need_ratio, need_gap = 0.80, 0.82, 0.78, 0.68
        elif exp_len >= 55:
            need_order, need_overlap, need_ratio, need_gap = 0.84, 0.84, 0.82, 0.55
        else:
            need_order, need_overlap, need_ratio, need_gap = 0.90, 0.90, 0.88, 0.38

        if order_cov >= need_order and overlap >= need_overlap and len_gap <= need_gap:
            return True
        if overlap >= need_overlap and ratio >= need_ratio and len_gap <= need_gap:
            return True

    return False


def to_mbasic_facebook_url(raw_url: str) -> str:
    u = str(raw_url or "").strip()
    if not u:
        return ""
    low = u.lower()
    if "facebook.com" not in low and "fb.watch" not in low:
        return ""
    try:
        if "fb.watch/" in low:
            tail = u.split("fb.watch/", 1)[1].strip("/")
            if tail:
                return f"https://mbasic.facebook.com/watch/?v={tail}"
        parsed = urlparse(u)
        path = parsed.path or "/"
        query = parsed.query or ""
        frag = parsed.fragment or ""
        base = f"https://mbasic.facebook.com{path}"
        if query:
            base += f"?{query}"
        if frag:
            base += f"#{frag}"
        return base
    except Exception:
        return ""


def _collect_mbasic_visible_text(driver) -> str:
    try:
        txt = driver.execute_script(
            """
            const sels = [
              'div[data-ft]',
              'article',
              'div[role="article"]',
              'div[id*="ufi"]',
              'h3 + div',
              'p'
            ];
            const out = [];
            const seen = new Set();
            for (const s of sels) {
              const nodes = document.querySelectorAll(s);
              for (const n of nodes) {
                const t = (n && n.innerText) ? n.innerText.trim() : '';
                if (!t || t.length < 6) continue;
                if (seen.has(t)) continue;
                seen.add(t);
                out.push(t);
                if (out.length >= 1200) break;
              }
              if (out.length >= 1200) break;
            }
            if (!out.length && document.body && document.body.innerText) {
              out.push(document.body.innerText);
            }
            return out.join('\\n');
            """
        ) or ""
        return str(txt).strip()
    except Exception:
        return ""


def _html_to_plain_text(html: str) -> str:
    s = str(html or "")
    if not s:
        return ""
    # Remove non-content blocks.
    s = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", s)
    s = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", s)
    s = re.sub(r"(?is)<!--.*?-->", " ", s)
    # Keep line breaks around common block tags.
    s = re.sub(r"(?is)<\s*br\s*/?\s*>", "\n", s)
    s = re.sub(r"(?is)</\s*(p|div|li|tr|h[1-6]|article|section)\s*>", "\n", s)
    # Drop remaining tags.
    s = re.sub(r"(?is)<[^>]+>", " ", s)
    s = html_lib.unescape(s)
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{2,}", "\n", s)
    return s.strip()


def _extract_follow_links_mbasic_html(html: str) -> list[str]:
    out: list[str] = []
    if not html:
        return out
    words = [
        "xem them",
        "xem them binh luan",
        "xem them phan hoi",
        "tat ca binh luan",
        "see more",
        "more comments",
        "view more",
        "replies",
        "more replies",
        "all comments",
        "load more",
        "view previous comments",
    ]
    seen: set[str] = set()
    for m in re.finditer(r'(?is)<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html):
        href = html_lib.unescape(str(m.group(1) or "").strip())
        label_raw = _html_to_plain_text(m.group(2) or "")
        label = normalize_match_text(label_raw)
        if not href or not label:
            continue
        if not any(w in label for w in words):
            continue
        full = urljoin("https://mbasic.facebook.com", href)
        if full in seen:
            continue
        seen.add(full)
        out.append(full)
        if len(out) >= 160:
            break
    return out


def extract_fb_comments_via_mbasic(driver, src_url: str, max_hops: int = 28) -> str:
    """
    Crawl mbasic Facebook pages to expand and collect comments text without relying on
    heavy dynamic UI interactions on the normal Facebook surface.
    """
    mbasic_url = to_mbasic_facebook_url(src_url)
    if not mbasic_url:
        return ""

    # 1) Prefer requests + browser cookies (more stable than dynamic clicking).
    if requests is not None:
        try:
            session = requests.Session()
            session.headers.update(
                {
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0.0.0 Safari/537.36"
                    ),
                    "Accept-Language": "vi,en-US;q=0.9,en;q=0.8",
                }
            )
            try:
                for ck in (driver.get_cookies() or []):
                    n = str((ck or {}).get("name", "")).strip()
                    v = str((ck or {}).get("value", "")).strip()
                    d = str((ck or {}).get("domain", "")).strip() or ".facebook.com"
                    p = str((ck or {}).get("path", "")).strip() or "/"
                    if n:
                        session.cookies.set(n, v, domain=d, path=p)
            except Exception:
                pass

            rq_chunks: list[str] = []
            rq_visited: set[str] = set()
            rq_queue: list[str] = [mbasic_url]
            rq_hop = 0
            while rq_queue and rq_hop < max_hops:
                rq_hop += 1
                cur = rq_queue.pop(0)
                if not cur or cur in rq_visited:
                    continue
                rq_visited.add(cur)
                try:
                    resp = session.get(cur, timeout=20, allow_redirects=True)
                except Exception:
                    continue
                if not resp or resp.status_code >= 400:
                    continue
                html = str(resp.text or "")
                txt = _html_to_plain_text(html)
                if txt:
                    rq_chunks.append(txt[:120000])
                for nxt in _extract_follow_links_mbasic_html(html):
                    if nxt not in rq_visited and nxt not in rq_queue:
                        rq_queue.append(nxt)

            if rq_chunks:
                merged_rq = "\n".join(rq_chunks)
                write_log(f"[SCAN_TEXT] mbasic(requests) collected: pages={len(rq_visited)} chars={len(merged_rq)}")
                if len(merged_rq) >= 500:
                    return merged_rq
                write_log("[SCAN_TEXT][WARN] mbasic(requests) text short, fallback to selenium mbasic.")
        except Exception as e:
            write_log(f"[SCAN_TEXT][WARN] mbasic(requests) failed: {e}")

    # 2) Fallback: selenium-based mbasic traversal.
    chunks: list[str] = []
    visited: set[str] = set()
    queue: list[str] = [mbasic_url]
    hop = 0

    while queue and hop < max_hops:
        hop += 1
        cur = queue.pop(0)
        if not cur or cur in visited:
            continue
        visited.add(cur)
        try:
            driver.get(cur)
            time.sleep(0.7)
        except Exception:
            continue

        txt = _collect_mbasic_visible_text(driver)
        if txt:
            chunks.append(txt[:120000])

        try:
            found = driver.execute_script(
                """
                const words = [
                  'xem them binh luan', 'xem thêm bình luận',
                  'xem them phan hoi', 'xem thêm phản hồi',
                  'xem them', 'xem thêm',
                  'see more comments', 'view more comments',
                  'more comments', 'more replies', 'replies', 'view previous comments',
                  'all comments', 'load more'
                ];
                const norm = (s) => (s || '')
                  .toLowerCase()
                  .normalize('NFD')
                  .replace(/[\\u0300-\\u036f]/g, '')
                  .replace(/\\s+/g, ' ')
                  .trim();
                const out = [];
                const seen = new Set();
                const as = document.querySelectorAll('a[href]');
                for (const a of as) {
                  const t = norm(a.innerText || a.textContent || '');
                  if (!t) continue;
                  let ok = false;
                  for (const w of words) {
                    if (t.includes(norm(w))) { ok = true; break; }
                  }
                  if (!ok) continue;
                  const href = a.getAttribute('href') || '';
                  if (!href) continue;
                  if (seen.has(href)) continue;
                  seen.add(href);
                  out.push({href: href, text: t});
                  if (out.length >= 120) break;
                }
                return out;
                """
            ) or []
        except Exception:
            found = []

        for item in found:
            try:
                href = str((item or {}).get("href", "")).strip()
            except Exception:
                href = ""
            if not href:
                continue
            nxt = urljoin("https://mbasic.facebook.com", href)
            if nxt not in visited and nxt not in queue:
                queue.append(nxt)

    if chunks:
        merged = "\n".join(chunks)
        write_log(f"[SCAN_TEXT] mbasic collected: pages={len(visited)} chars={len(merged)}")
        return merged
    write_log(f"[SCAN_TEXT] mbasic collected empty from: {mbasic_url}")
    return ""


def extract_text_from_link_for_scan(driver, url: str) -> str:
    """
    Extract main textual content from a post/article page for Scan Only Text mode.
    """
    if "facebook.com" in (url or "").lower() or "fb.watch" in (url or "").lower():
        try:
            mbasic_text = extract_fb_comments_via_mbasic(driver, url)
            if len(mbasic_text) >= 500:
                return mbasic_text
            if mbasic_text:
                write_log(
                    f"[SCAN_TEXT][WARN] mbasic text short (len={len(mbasic_text)}), fallback dynamic extraction."
                )
        except Exception as e:
            write_log(f"[SCAN_TEXT][WARN] mbasic extraction failed: {e}")

    try:
        # Expand all "xem them/see more/comments/replies" first, then scan text.
        expand_script = """
            const words = [
              'xem them', 'xem thêm', 'xem them binh luan', 'xem thêm bình luận',
              'xem them phan hoi', 'xem thêm phản hồi', 'tat ca binh luan', 'tất cả bình luận',
              'see more', 'more comments', 'view more', 'view more comments',
              'replies', 'more replies', 'all comments', 'load more'
            ];
            const deny = ['thich', 'like', 'share', 'chia se', 'follow', 'theo doi'];
            const norm = (s) => (s || '')
              .toLowerCase()
              .normalize('NFD')
              .replace(/[\\u0300-\\u036f]/g, '')
              .replace(/\\s+/g, ' ')
              .trim();
            const mayClick = (el) => {
              if (!el) return false;
              const raw = el.innerText || el.textContent || '';
              const t = norm(raw);
              if (!t || t.length < 3 || t.length > 220) return false;
              for (const d of deny) {
                if (t === d || t.startsWith(d + ' ')) return false;
              }
              let hit = false;
              for (const w of words) {
                const wn = norm(w);
                if (t.includes(wn)) { hit = true; break; }
              }
              if (!hit) return false;
              try { el.scrollIntoView({block: 'center'}); } catch (_) {}
              try { el.click(); return true; } catch (_) {}
              try {
                const evt = new MouseEvent('click', {bubbles: true, cancelable: true, view: window});
                el.dispatchEvent(evt);
                return true;
              } catch (_) {}
              return false;
            };
            const nodes = Array.from(document.querySelectorAll('a,button,[role="button"],div,span'));
            let clicked = 0;
            for (const n of nodes) {
              if (clicked >= 240) break;
              if (mayClick(n)) clicked++;
            }
            return clicked;
        """
        total_clicked = 0
        no_click_rounds = 0
        max_rounds = 36
        for _ in range(max_rounds):
            # Click at current viewport first.
            try:
                clicked_1 = int(driver.execute_script(expand_script) or 0)
            except Exception:
                clicked_1 = 0
            total_clicked += max(0, clicked_1)
            if clicked_1 > 0:
                no_click_rounds = 0
            else:
                no_click_rounds += 1

            time.sleep(0.28)
            try:
                m = driver.execute_script(
                    "return {y:(window.pageYOffset||document.documentElement.scrollTop||0),"
                    "vh:(window.innerHeight||0),h:(document.body&&document.body.scrollHeight)||0};"
                ) or {}
                y = float(m.get("y", 0) or 0)
                vh = float(m.get("vh", 0) or 0)
                h = float(m.get("h", 0) or 0)
                at_bottom = (y + vh) >= (h - 8)
            except Exception:
                at_bottom = False

            if at_bottom:
                # At bottom: one more expand pass; if still no click for a while -> done.
                try:
                    clicked_bottom = int(driver.execute_script(expand_script) or 0)
                except Exception:
                    clicked_bottom = 0
                total_clicked += max(0, clicked_bottom)
                if clicked_bottom > 0:
                    no_click_rounds = 0
                    time.sleep(0.25)
                    continue
                if no_click_rounds >= 3:
                    break
                # Rewind then scan downward again to catch delayed loaded controls.
                try:
                    driver.execute_script("window.scrollTo(0, 0);")
                except Exception:
                    pass
                time.sleep(0.35)
                continue

            # Not at bottom: keep scrolling to force comment lazy-load.
            try:
                driver.execute_script("window.scrollBy(0, Math.max(760, Math.floor(window.innerHeight * 0.95)));")
            except Exception:
                pass
            # Also scroll comment containers (FB often uses inner scroll regions).
            try:
                driver.execute_script(
                    """
                    const els = Array.from(document.querySelectorAll('div,section,main,article'));
                    let moved = 0;
                    for (const el of els) {
                      try {
                        const canScroll = el.scrollHeight > (el.clientHeight + 60);
                        if (!canScroll) continue;
                        const oldTop = el.scrollTop;
                        el.scrollTop = Math.min(el.scrollHeight, oldTop + Math.max(700, Math.floor(el.clientHeight * 0.9)));
                        if (el.scrollTop !== oldTop) moved++;
                      } catch (_) {}
                    }
                    return moved;
                    """
                )
            except Exception:
                pass
            time.sleep(0.35)

        write_log(f"[SCAN_TEXT] expand-all before scan: clicks={total_clicked}, rounds={max_rounds}")
    except Exception:
        pass

    chunks = []
    try:
        if "facebook.com" in (url or "").lower() or "fb.watch" in (url or "").lower():
            _p, cap = get_fb_profile_and_caption(driver, url)
            if cap and cap.strip():
                chunks.append(cap.strip())
    except Exception:
        pass
    try:
        if "youtube.com" in (url or "").lower() or "youtu.be" in (url or "").lower():
            title = (get_youtube_title(driver) or "").strip()
            if title:
                chunks.append(title)
    except Exception:
        pass
    try:
        comment_like_texts = driver.execute_script(
            """
            const selectors = [
              '[aria-label*="comment" i]',
              '[aria-label*="bình luận" i]',
              '[data-testid*="comment" i]',
              '[class*="comment" i]',
              'ytd-comment-thread-renderer',
              '[data-e2e*="comment"]',
              '[class*="Comment" i]',
              'div[role="article"] div[dir="auto"]'
            ];
            const out = [];
            const seen = new Set();
            for (const sel of selectors) {
              const nodes = document.querySelectorAll(sel);
              for (const n of nodes) {
                const t = (n && n.innerText) ? n.innerText.trim() : '';
                if (!t) continue;
                if (t.length < 6) continue;
                if (seen.has(t)) continue;
                seen.add(t);
                out.push(t);
                if (out.length >= 800) break;
              }
              if (out.length >= 800) break;
            }
            return out.join('\\n');
            """
        ) or ""
        comment_like_texts = str(comment_like_texts).strip()
        if comment_like_texts:
            chunks.append(comment_like_texts[:180000])
    except Exception:
        pass
    try:
        body_text = (
            driver.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : '';"
            )
            or ""
        )
        body_text = str(body_text).strip()
        if body_text:
            chunks.append(body_text[:180000])
    except Exception:
        pass
    if not chunks:
        return ""
    # De-dup while preserving order.
    out = []
    seen = set()
    for c in chunks:
        key = normalize_match_text(c)[:300]
        if key and key not in seen:
            seen.add(key)
            out.append(c)
    return "\n".join(out)


# ================= LOG =================
def write_log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except:
        pass


def _normalize_sheet_key(sheet_url: str) -> str:
    u = (sheet_url or "").strip()
    if not u:
        return ""
    # Prefer stable spreadsheet id from URL: /spreadsheets/d/<id>/
    try:
        parts = u.split("/spreadsheets/d/")
        if len(parts) > 1:
            tail = parts[1]
            sid = tail.split("/", 1)[0].strip()
            if sid:
                return f"sheet_id:{sid}"
    except Exception:
        pass
    return f"sheet_url:{u}"


def _sheet_history_key(sheet_url: str, sheet_name: str = "") -> str:
    # History is tracked by Sheet link (not worksheet name).
    return _normalize_sheet_key(sheet_url)


def load_error_history() -> dict:
    try:
        if not os.path.exists(ERROR_HISTORY_PATH):
            return {}
        with open(ERROR_HISTORY_PATH, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        if isinstance(data, dict):
            return data
    except Exception as e:
        write_log(f"[WARN] Load error history failed: {e}")
    return {}


def save_error_history(data: dict):
    try:
        with open(ERROR_HISTORY_PATH, "w", encoding="utf-8") as f:
            json.dump(data or {}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        write_log(f"[WARN] Save error history failed: {e}")


def get_error_rows_for_sheet(sheet_url: str, sheet_name: str = "") -> set[int]:
    key = _sheet_history_key(sheet_url, sheet_name)
    db = load_error_history()
    item = db.get(key) or {}
    rows = item.get("rows") or []

    # Backward compatibility: merge old keys that included sheet_name suffix.
    if not rows:
        legacy_prefix = f"{(sheet_url or '').strip()}|"
        merged = []
        for k, v in db.items():
            if isinstance(k, str) and k.startswith(legacy_prefix):
                merged.extend((v or {}).get("rows") or [])
        rows = merged
    out = set()
    for r in rows:
        try:
            rv = int(r)
            if rv >= 1:
                out.add(rv)
        except Exception:
            continue
    return out


def get_error_details_for_sheet(sheet_url: str, sheet_name: str = "") -> dict[int, str]:
    key = _sheet_history_key(sheet_url, sheet_name)
    db = load_error_history()
    item = db.get(key) or {}
    raw = item.get("details") or {}
    out: dict[int, str] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            try:
                rk = int(k)
                if rk >= 1:
                    out[rk] = str(v or "").strip()
            except Exception:
                continue
    return out


def set_error_rows_for_sheet(
    sheet_url: str,
    sheet_name: str = "",
    rows: set[int] = None,
    details: dict[int, str] | None = None,
):
    rows = rows or set()
    details = details or {}
    key = _sheet_history_key(sheet_url, sheet_name)
    db = load_error_history()
    sorted_rows = sorted({int(r) for r in rows if int(r) >= 1})
    details_clean = {}
    for r in sorted_rows:
        msg = str(details.get(r, "")).strip()
        if msg:
            details_clean[str(r)] = msg[:220]
    if sorted_rows:
        db[key] = {
            "sheet_url": (sheet_url or "").strip(),
            "sheet_name": (sheet_name or "").strip(),
            "rows": sorted_rows,
            "details": details_clean,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    else:
        db.pop(key, None)
    save_error_history(db)


def parse_target_rows_text(raw_text: str) -> tuple[list[int], str]:
    """
    Parse manual row selection text.
    Supported formats: "7", "7,8,9", "7-12", "7, 10-15; 20".
    Returns (sorted_unique_rows, error_message).
    """
    raw = str(raw_text or "").strip()
    if not raw:
        return [], ""
    tokens = [t.strip() for t in re.split(r"[,\n;]+", raw) if str(t).strip()]
    if not tokens:
        return [], ""
    selected: set[int] = set()
    for token in tokens:
        normalized = re.sub(r"\s*-\s*", "-", token.strip())
        if re.fullmatch(r"\d+", normalized):
            row_num = int(normalized)
            if row_num < 1:
                return [], f"Hàng không hợp lệ: '{token}' (phải >= 1)."
            selected.add(row_num)
            continue
        range_match = re.fullmatch(r"(\d+)-(\d+)", normalized)
        if range_match:
            left = int(range_match.group(1))
            right = int(range_match.group(2))
            if left < 1 or right < 1:
                return [], f"Khoảng hàng không hợp lệ: '{token}' (phải >= 1)."
            if left > right:
                return [], f"Khoảng hàng không hợp lệ: '{token}' (đầu <= cuối)."
            if (right - left) > 50000:
                return [], f"Khoảng hàng quá lớn: '{token}'."
            for row_num in range(left, right + 1):
                selected.add(row_num)
            continue
        return [], f"Định dạng không hợp lệ: '{token}'. Dùng ví dụ: 7,8,10-15"
    rows = sorted(selected)
    if len(rows) > 100000:
        return [], "Danh sách hàng quá dài (tối đa 100000 hàng)."
    return rows, ""


def list_saved_error_sheets() -> list[dict]:
    db = load_error_history()
    items = []
    for _, v in db.items():
        if not isinstance(v, dict):
            continue
        url = str(v.get("sheet_url", "")).strip()
        if not url:
            continue
        rows = v.get("rows") or []
        updated_at = str(v.get("updated_at", "")).strip()
        sheet_name = str(v.get("sheet_name", "")).strip()
        items.append(
            {
                "sheet_url": url,
                "sheet_name": sheet_name,
                "rows_count": len(rows),
                "updated_at": updated_at,
            }
        )
    # Newest first (timestamp format YYYY-MM-DD HH:MM:SS is lexicographically sortable)
    items.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    return items

# ================= FB COMMENT PARSE =================
def extract_comment_id(url):
    comment_id = None
    if "comment_id=" in url:
        start = url.find("comment_id=") + len("comment_id=")
        end = url.find("&", start)
        if end == -1:
            comment_id = url[start:]
        else:
            comment_id = url[start:end]
    elif "reply_comment_id=" in url:
        start = url.find("reply_comment_id=") + len("reply_comment_id=")
        end = url.find("&", start)
        if end == -1:
            comment_id = url[start:]
        else:
            comment_id = url[start:end]
    return comment_id

def get_highlighted_fb_comment(driver, url):
    time.sleep(FB_COMMENT_READY_WAIT)
    
    comment_id = extract_comment_id(url)
    if comment_id:
        try:
            # Find the exact comment element by comment_id
            comment_element = driver.find_element(By.XPATH, f"//div[contains(@data-ft, '\"comment_id\":\"{comment_id}\"')]")
            # Find the text content within the comment
            text_elements = comment_element.find_elements(By.XPATH, ".//div[@dir='auto']")
            text = ""
            for elem in text_elements:
                t = elem.text.strip()
                if len(t) > 5:
                    text = t
                    break
            if text:
                return text
        except:
            pass
    
    # Fallback: Look for comment by checking highlighted/focused elements
    try:
        # Try to find the most prominent comment text
        all_comment_divs = driver.find_elements(By.XPATH, "//div[@data-testid='comment']")
        if all_comment_divs:
            # Get the first visible comment's text
            for comment_div in all_comment_divs[:3]:  # Check first 3 comments
                try:
                    text_elem = comment_div.find_element(By.XPATH, ".//div[@dir='auto']")
                    text = text_elem.text.strip()
                    if len(text) > 5:
                        return text
                except:
                    continue
    except:
        pass

    # Position-based detection as last resort
    candidates = driver.find_elements(By.XPATH, "//div[@dir='auto']")
    best_text = ""

    for c in candidates:
        try:
            text = c.text.strip()

            if len(text) < 8:
                continue

            # loÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡i text rÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡c
            if text.lower() in ["thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­ch", "trÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â£ lÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âi", "xem thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âªm", "tiÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¿p tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¥c", "tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â£i thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âªm"]:
                continue

            rect = driver.execute_script("""
                const r = arguments[0].getBoundingClientRect();
                return {top: r.top, bottom: r.bottom};
            """, c)

            # comment ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â£c highlight thÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âng nÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â±m gÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â§n ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â§u viewport
            if rect["top"] > 120 and rect["top"] < 450:
                best_text = text
                break

        except:
            continue

    return best_text

# ================= TIKTOK CAPTION =================
def get_tiktok_caption(driver):
    caption = ""
    
    # Try meta tags first (og:description) - TikTok should have this
    try:
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            prop = m.get_attribute("property")
            if prop == "og:description":
                caption = m.get_attribute("content") or ""
                if caption:
                    return caption.strip()
    except:
        pass
    
    # If not found from meta, try to get from DOM
    if not caption:
        try:
            # Look for description text in TikTok DOM - multiple selectors
            selectors = [
                "//span[@data-e2e='video-desc']",
                "//div[@data-testid='video-desc']//span",
                "//h2//span",
                "//h1//span",
            ]
            for selector in selectors:
                desc_elements = driver.find_elements(By.XPATH, selector)
                for elem in desc_elements:
                    text = elem.text.strip()
                    if len(text) > 5 and len(text) < 10000:  # TikTok captions can be long
                        caption = text
                        return caption.strip()
        except:
            pass
    
    return caption


def get_tiktok_profile_name(driver, source_url: str = "") -> str:
    # Resolve handle from final URL/source URL first.
    expected_handle = ""
    try:
        current_url = (driver.current_url or "").strip()
    except Exception:
        current_url = ""
    for candidate_url in [current_url, source_url]:
        h = extract_account_name_from_url(candidate_url)
        if h:
            expected_handle = h.strip()
            break

    # 1) Try JSON-LD metadata (usually ties directly to current video author).
    try:
        scripts = driver.find_elements(By.XPATH, "//script[@type='application/ld+json']")
        for s in scripts:
            raw = (s.get_attribute("textContent") or "").strip()
            if not raw:
                continue
            try:
                data = json.loads(raw)
            except Exception:
                continue

            objs = data if isinstance(data, list) else [data]
            for obj in objs:
                if not isinstance(obj, dict):
                    continue
                author = obj.get("author")
                if isinstance(author, list):
                    author = author[0] if author else {}
                if not isinstance(author, dict):
                    continue

                display = clean_account_name_candidate(str(author.get("name", "")).strip())
                alt = str(author.get("alternateName", "")).strip()
                alt_handle = alt if alt.startswith("@") else ""
                if not alt_handle and alt and alt.startswith("@"):
                    alt_handle = alt

                # If we know expected handle, only trust matching author blocks.
                if expected_handle and alt_handle and alt_handle != expected_handle:
                    continue

                if display and not display.startswith("@") and is_likely_account_name(display):
                    return display
    except Exception:
        pass

    # 2) Prefer display name (nickname) but only when paired with expected handle.
    try:
        selectors = [
            "//*[@data-e2e='video-author-nickname']",
            "//*[@data-e2e='browse-user-nickname']",
            "//h3[contains(@data-e2e,'nickname')]",
            "//h2[contains(@data-e2e,'nickname')]",
        ]
        for sel in selectors:
            elems = driver.find_elements(By.XPATH, sel)
            for e in elems:
                txt = clean_account_name_candidate((e.text or "").strip())
                if not (txt and not txt.startswith("@") and is_likely_account_name(txt)):
                    continue

                if expected_handle:
                    try:
                        nearby = driver.execute_script(
                            """
                            const el = arguments[0];
                            const root = el.closest('article, section, div') || el.parentElement;
                            return (root && root.innerText) ? root.innerText : '';
                            """,
                            e,
                        ) or ""
                        if expected_handle not in nearby:
                            continue
                    except Exception:
                        continue

                    return txt
    except Exception:
        pass

    # 3) Meta fallback: "<display name> on TikTok"
    try:
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            if m.get_attribute("property") == "og:title":
                title = (m.get_attribute("content") or "").strip()
                if title:
                    lowered = title.lower()
                    marker = " on tiktok"
                    if marker in lowered:
                        title = title[:lowered.find(marker)].strip()
                    txt = clean_account_name_candidate(title)
                    if txt and not txt.startswith("@") and is_likely_account_name(txt):
                        return txt
    except Exception:
        pass

    # 4) Fallback: resolved URL after redirects (@handle)
    for candidate_url in [current_url, source_url]:
        uname = extract_account_name_from_url(candidate_url)
        if uname:
            return uname.strip()

    # 5) DOM fallback: profile links that point to /@username
    try:
        links = driver.find_elements(By.XPATH, "//a[contains(@href, '/@')]")
        for a in links:
            href = (a.get_attribute("href") or "").strip()
            uname = extract_account_name_from_url(href)
            if uname:
                return uname.strip()
            txt = (a.text or "").strip()
            if txt.startswith("@") and len(txt) > 1:
                return txt
    except Exception:
        pass

    return ""


# ================= YOUTUBE TITLE =================
def get_youtube_title(driver):
    title = ""
    try:
        # Try og:title meta
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            if m.get_attribute("property") == "og:title":
                title = m.get_attribute("content") or ""
                if title:
                    return title.strip()
    except:
        pass

    try:
        # Fallback to document title
        t = driver.title or driver.execute_script("return document.title")
        if t:
            return t.strip()
    except:
        pass

    return title


def get_youtube_channel(driver):
    channel = ""
    try:
        # Try meta article:author first (often has channel name)
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            prop = m.get_attribute("property") or m.get_attribute("name") or ""
            if "author" in prop.lower():
                channel = m.get_attribute("content") or ""
                if channel and len(channel) > 2:
                    return channel.strip()
    except:
        pass

    try:
        # Try og:site_name
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            if m.get_attribute("property") == "og:site_name" or m.get_attribute("name") == "og:site_name":
                channel = m.get_attribute("content") or ""
                if channel and len(channel) > 2:
                    return channel.strip()
    except:
        pass

    try:
        # Try to extract from page title (usually has channel: | Uploaded by channel)
        title = driver.title or ""
        if " - " in title:
            parts = title.split(" - ")
            if len(parts) > 1:
                potential_channel = parts[-1].strip()
                if len(potential_channel) > 2 and len(potential_channel) < 100:
                    return potential_channel
    except:
        pass

    try:
        # Fallback: look for channel link/button in header
        channel_link = driver.find_element(By.XPATH, "//a[contains(@href, 'youtube.com/@') or contains(@href, '/channel/') or contains(@href, '/user/')][1]")
        t = channel_link.text.strip()
        if t and len(t) > 2:
            return t
    except:
        pass

    return channel


# ================= NAME CLEAN =================
def clean_fb_profile_name(name: str) -> str:
    if not name:
        return name
    n = name.strip()

    def _strip_phrase(original: str, phrase: str) -> str:
        # Normalize to ASCII for matching while keeping original for slicing
        norm = ""
        mapping = []
        for i, ch in enumerate(original):
            decomp = unicodedata.normalize("NFD", ch)
            for dc in decomp:
                if unicodedata.category(dc) == "Mn":
                    continue
                norm += dc
                mapping.append(i)
        norm_lower = norm.lower()
        phrase_lower = phrase.lower()
        idx = norm_lower.find(phrase_lower)
        if idx != -1:
            end_norm = idx + len(phrase_lower)
            end_orig = mapping[end_norm - 1] + 1
            return original[end_orig:].strip(" :-")
        return original

    n2 = _strip_phrase(n, "bai viet cua")
    if n2 != n:
        return n2.strip()

    return n


def is_likely_account_name(text: str) -> bool:
    if not text:
        return False
    t = text.strip()
    if not t:
        return False
    if "\n" in t or "\r" in t:
        return False
    if len(t) > 80:
        return False

    lower = t.lower()
    noise_markers = [
        " views", " view", " reactions", " reaction", " comments", " comment",
        " shares", " share", " like", " thich", " binh luan", " xem them",
        "http://", "https://", "www.", "#"
    ]
    if any(m in lower for m in noise_markers):
        return False
    return True


def clean_account_name_candidate(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    t = t.splitlines()[0].strip()
    for sep in [" · ", "Â·", " Â· ", "â€¢", "•", "|", " - "]:
        if sep in t:
            t = t.split(sep, 1)[0].strip()
    t = t.strip(":- ")
    t = clean_fb_profile_name(t)
    return t.strip()


def is_numeric_like_account_name(name: str) -> bool:
    t = (name or "").strip().lstrip("@")
    if not t:
        return False
    if t.startswith("profile_") and t[8:].isdigit():
        return True
    return t.isdigit() and len(t) >= 5


def extract_account_name_from_title(title: str) -> str:
    t = clean_account_name_candidate(title)
    if not t:
        return ""
    if t.lower() in {"facebook", "instagram", "tiktok", "youtube"}:
        return ""
    return t


def extract_account_name_from_url(url: str) -> str:
    if not url:
        return ""
    try:
        parsed = urlparse(url)
    except Exception:
        return ""

    host = (parsed.netloc or "").lower()
    path_parts = [p for p in (parsed.path or "").split("/") if p]
    query = parse_qs(parsed.query or "")

    # TikTok: /@username/video/...
    if "tiktok.com" in host:
        for p in path_parts:
            if p.startswith("@") and len(p) > 1:
                return p

    # Instagram: /username/p/... or /username/reel/...
    if "instagram.com" in host or "instagr.am" in host:
        if path_parts:
            first = path_parts[0]
            if first not in {"p", "reel", "tv", "stories", "explore"}:
                return first

    # YouTube: /@handle, /channel/<id>, /user/<name>, /c/<name>
    if "youtube.com" in host or "youtu.be" in host:
        if path_parts:
            first = path_parts[0]
            if first.startswith("@") and len(first) > 1:
                return first
            if first in {"channel", "user", "c"} and len(path_parts) > 1:
                return path_parts[1]

    # Facebook common paths
    if "facebook.com" in host or "fb.watch" in host:
        target = (query.get("u") or [""])[0].strip()
        if target:
            nested = extract_account_name_from_url(target)
            if nested:
                return nested

        if path_parts:
            first = path_parts[0]
            reserved = {
                "watch", "reel", "reels", "story.php", "permalink.php",
                "photo", "photos", "photo.php", "groups", "events", "share", "plugins",
                "login", "hashtag"
            }
            if first == "profile.php":
                pid = (query.get("id") or [""])[0].strip()
                if pid:
                    return f"profile_{pid}"
            if first == "people" and len(path_parts) > 1:
                return path_parts[1]
            if first not in reserved:
                return first

    return ""


def normalize_account_name(name: str, url: str) -> str:
    n = clean_account_name_candidate(name)
    if n and not is_likely_account_name(n):
        n = ""
    if n and is_numeric_like_account_name(n):
        n = ""
    if not n:
        n = clean_account_name_candidate(extract_account_name_from_url(url))
    if n and is_numeric_like_account_name(n):
        n = ""
    return n.strip()


def get_post_caption(driver):

    caption = ""

    # Try to expand "See more" / "Xem thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âªm" buttons to reveal full caption
    try:
        buttons = driver.find_elements(By.XPATH, "//div[contains(text(), 'Xem') or contains(text(), 'See more')]")
        for b in buttons:
            try:
                driver.execute_script("arguments[0].click();", b)
                time.sleep(UI_CLICK_SETTLE_SLEEP)
            except:
                continue
    except:
        pass

    try:
        driver.execute_script("window.scrollTo(0, 600);")
        time.sleep(UI_SCROLL_SETTLE_SLEEP)
    except:
        pass


    # Strategy 0: Facebook main caption container (data-ad-preview="message")
    try:
        elems = driver.find_elements(By.XPATH, "//div[@data-ad-preview='message']//div[@dir='auto'] | //div[@data-ad-preview='message']")
        for elem in elems:
            text = elem.text.strip()
            if len(text) > 5 and len(text) < 5000:
                return text
    except:
        pass

    # Strategy 0.5: Reel / main caption areas
    try:
        selectors = [
            "//div[@data-testid='post_message']//div[@dir='auto']",
            "//div[@data-pagelet='Reel']//div[@dir='auto']",
            "//div[@role='main']//div[@dir='auto']",
        ]
        noise_words = ["thich", "tra loi", "xem", "tiep tuc", "chia se", "binh luan", "tai xuong",
                       "like", "comment", "see more", "show more", "share", "download"]
        for selector in selectors:
            elems = driver.find_elements(By.XPATH, selector)
            for elem in elems:
                text = elem.text.strip()
                text_lower = text.lower()
                if (len(text) > 10 and len(text) < 5000 and text_lower not in noise_words and
                    "facebook.com" not in text_lower and not text.startswith("http") and not text.startswith("www.")):
                    return text
    except:
        pass

    # Strategy 1: Look for shared article/link content first (for shared posts)
    try:
        shared_selectors = [
            "//div[@data-testid='share_content']//div[@dir='auto']",
            "//div[contains(@class, 'xwib8y')]//div[@dir='auto']",
            "//div[contains(@class, 'x1iyjqo2')]//h2//span | //div[contains(@class, 'x1iyjqo2')]//div[@dir='auto']",
            "//div[@role='article']//a[contains(@href, 'facebook.com') or contains(@href, 'l.facebook.com')]//parent::*//*[@dir='auto']",
            "//div[@data-testid='message']//div[@dir='auto']",
        ]

        for selector in shared_selectors:
            try:
                elems = driver.find_elements(By.XPATH, selector)
                for elem in elems:
                    text = elem.text.strip()
                    if len(text) > 10 and len(text) < 5000:
                        caption = text
                        break
                if caption:
                    return caption
            except:
                continue
    except:
        pass

    # Strategy 2: Extract from the main article container text
    try:
        article = None
        # try several article selectors
        try:
            article = driver.find_element(By.XPATH, "//article")
        except:
            try:
                article = driver.find_element(By.XPATH, "//div[@role='article']")
            except:
                article = None

        if article:
            full_text = article.text or ""
            lines = [l.strip() for l in full_text.split('\n') if l.strip()]
            # filter noise and pick the longest reasonable line
            candidates = [l for l in lines if len(l) > 10 and 'facebook.com' not in l.lower() and 'see more' not in l.lower()]
            if candidates:
                caption = max(candidates, key=len)
                return caption
    except:
        pass

    # Strategy 3: Look for all div[@dir='auto'] and filter intelligently
    try:
        all_divs = driver.find_elements(By.XPATH, "//div[@dir='auto']")
        candidates = []
        noise_words = ["thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­ch", "trÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â£ lÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âi", "xem thÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âªm", "tiÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¿p tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¥c", "chia sÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»", "bÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬nh luÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­n", "tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂºÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â£i xuÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â»ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œng",
                       "like", "comment", "see more", "show more", "share", "download"]
        for div in all_divs:
            text = div.text.strip()
            text_lower = text.lower()
            if (len(text) > 10 and len(text) < 5000 and text_lower not in noise_words and
                "facebook.com" not in text_lower and not text.startswith("http") and not text.startswith("www.")):
                candidates.append(text)

        if candidates:
            candidates.sort(key=len, reverse=True)
            for c in candidates:
                if 12 < len(c) < 2000:
                    caption = c
                    return caption
    except:
        pass

    # Strategy 4: Fallback to meta tags
    try:
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            prop = m.get_attribute("property")
            if prop == "og:description":
                caption = m.get_attribute("content") or ""
                if caption:
                    return caption
    except:
        pass

    return caption


# ================= INSTAGRAM PARSE =================
def get_instagram_profile_and_caption(driver, url):
    name = ""
    caption = ""
    og_title = ""
    og_desc = ""

    try:
        metas = driver.find_elements(By.TAG_NAME, "meta")
        for m in metas:
            prop = m.get_attribute("property") or m.get_attribute("name") or ""
            if prop == "og:title":
                og_title = m.get_attribute("content") or ""
            if prop in ("og:description", "description"):
                og_desc = m.get_attribute("content") or ""
    except:
        pass

    # Try to parse profile from URL path
    try:
        if "instagram.com/" in url:
            path_part = url.split("instagram.com/", 1)[1]
            path_part = path_part.split("?", 1)[0]
            first = path_part.strip("/").split("/")[0]
            if first and first not in ("p", "reel", "tv", "stories", "explore"):
                name = first
    except:
        pass

    # Try to parse display name from og:title
    if not name and og_title:
        t = og_title
        if " on Instagram" in t:
            name = t.split(" on Instagram", 1)[0].strip()
        elif "Instagram:" in t:
            name = t.split("Instagram:", 1)[0].strip()
        elif "Instagram" in t:
            name_part = t.split("Instagram", 1)[0].strip(" -|")
            name = name_part.split("(", 1)[0].strip() or name_part

    # Caption from og:description
    if og_desc:
        c = og_desc
        if "on Instagram:" in c:
            c = c.split("on Instagram:", 1)[1].strip()
        elif "Instagram:" in c:
            c = c.split("Instagram:", 1)[1].strip()
        caption = c.strip("\"' " )

    name = normalize_account_name(name, url)
    return name, caption.strip()

# ================= FB PARSE =================
def get_facebook_actor_name(driver) -> str:
    candidates = []
    selectors = [
        # Common actor/title link areas on Facebook posts/reels
        "//h2//a | //h3//a | //strong//a",
        "//div[@role='article']//a[contains(@href,'facebook.com') or contains(@href,'/profile.php') or contains(@href,'/people/') or contains(@href,'/reel/')]/span",
        "//a[contains(@href,'/profile.php') or contains(@href,'/people/')]/span",
    ]
    for xp in selectors:
        try:
            elems = driver.find_elements(By.XPATH, xp)
            for e in elems:
                t = (e.text or "").strip()
                if t:
                    candidates.append(t)
        except:
            pass

    # aria-label fallback
    try:
        elems = driver.find_elements(By.XPATH, "//a[@aria-label]")
        for e in elems[:80]:
            t = (e.get_attribute("aria-label") or "").strip()
            if t:
                candidates.append(t)
    except:
        pass

    noise = {"Thích", "Bình luận", "Chia sẻ", "Like", "Comment", "Share", "Follow", "Theo dõi"}
    for c in candidates:
        cc = clean_account_name_candidate(c)
        if 2 < len(cc) < 80 and cc not in noise and is_likely_account_name(cc):
            return cc
    return ""


def get_fb_profile_and_caption(driver, url):
    name = ""
    caption = ""

    url_l = (url or "").lower()
    # Only treat as comment when link contains the word "comment"
    is_comment = "comment" in url_l
    is_tiktok = "tiktok.com" in url_l or "vt.tiktok.com" in url_l
    is_instagram = "instagram.com" in url_l or "instagr.am" in url_l
    is_facebook = ("facebook.com" in url_l) or ("fb.watch" in url_l) or ("m.facebook.com" in url_l)

    # ===== TIKTOK MODE =====
    if is_tiktok:
        caption = get_tiktok_caption(driver)
        # Prefer resolved URL/profile link over broad "@..." DOM text to avoid wrong mentions.
        name = get_tiktok_profile_name(driver, url)

    # ===== INSTAGRAM MODE =====
    elif is_instagram:
        name, caption = get_instagram_profile_and_caption(driver, url)

    # ===== FACEBOOK MODE =====
    elif is_facebook:
        name = get_facebook_actor_name(driver)
        if is_comment:
            caption = get_highlighted_fb_comment(driver, url)
        else:
            caption = get_post_caption(driver)

    # ===== FALLBACK MODE =====
    if not caption:
        try:
            metas = driver.find_elements(By.TAG_NAME, "meta")
            for m in metas:
                prop = m.get_attribute("property")
                if prop == "og:description" and not caption:
                    caption = m.get_attribute("content") or ""
                if prop == "og:title" and not name:
                    name = m.get_attribute("content") or ""
        except:
            pass

    # ===== PROFILE NAME FALLBACK =====
    if not name:
        try:
            elems = driver.find_elements(By.XPATH, "//h2//span | //strong//span")
            for e in elems:
                t = e.text.strip()
                if 2 < len(t) < 60:
                    name = t
                    break
        except:
            pass

    if not name and is_facebook:
        name = get_facebook_actor_name(driver)

    if not name:
        try:
            name = extract_account_name_from_title(driver.title or "")
        except:
            pass

    name = normalize_account_name(name, url)
    return name.strip(), caption.strip()


def get_fb_post_datetime(driver):
    post_time = ""

    try:
        elems = driver.find_elements(By.XPATH, "//abbr | //time")
        for e in elems:
            dt_attr = (e.get_attribute("datetime") or "").strip()
            if dt_attr:
                post_time = dt_attr
                break
            t = e.get_attribute("title")
            if t:
                post_time = t
                break
    except:
        pass

    if not post_time:
        try:
            metas = driver.find_elements(By.TAG_NAME, "meta")
            for m in metas:
                prop = (m.get_attribute("property") or "").strip().lower()
                name = (m.get_attribute("name") or "").strip().lower()
                content = (m.get_attribute("content") or "").strip()
                if not content:
                    continue
                if prop in {"article:published_time", "og:updated_time"}:
                    post_time = content
                    break
                if name in {"pubdate", "publishdate", "date", "datepublished"}:
                    post_time = content
                    break
        except:
            pass

    return post_time.strip()


def get_air_date_token(post_time: str) -> str:
    raw = (post_time or "").strip()
    if not raw:
        return ""

    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.strftime("%Y%m%d")
    except Exception:
        pass

    m = re.search(r"(20\d{2})[/-](\d{1,2})[/-](\d{1,2})", raw)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).strftime("%Y%m%d")
        except Exception:
            pass

    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](20\d{2})", raw)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).strftime("%Y%m%d")
        except Exception:
            pass

    return ""


def detect_platform_label(url: str) -> str:
    u = (url or "").lower()
    if "tiktok.com" in u or "vt.tiktok.com" in u:
        return "TikTok"
    if "instagram.com" in u or "instagr.am" in u:
        return "Instagram"
    if "youtube.com" in u or "youtu.be" in u:
        return "YouTube"
    if "facebook.com" in u or "fb.watch" in u or "m.facebook.com" in u:
        return "Facebook"
    return "Other"


def _extract_tiktok_video_id(url: str) -> str:
    text = str(url or "").strip().lower()
    if not text:
        return ""
    m = re.search(r"/video/(\d+)", text)
    return str(m.group(1)) if m else ""


def _extract_tiktok_handle(url: str) -> str:
    text = str(url or "").strip().lower()
    if not text:
        return ""
    m = re.search(r"/@([^/?#]+)/video/\d+", text)
    return str(m.group(1)).strip() if m else ""


def _is_expected_tiktok_page(requested_url: str, current_url: str) -> bool:
    req = str(requested_url or "").strip().lower()
    cur = str(current_url or "").strip().lower()
    if not req or not cur:
        return False
    req_video_id = _extract_tiktok_video_id(req)
    cur_video_id = _extract_tiktok_video_id(cur)
    # If request includes a concrete video id, current page must match exactly.
    if req_video_id:
        if not cur_video_id:
            return False
        if req_video_id != cur_video_id:
            return False
    # Intentionally ignore handle (@username) because TikTok can rewrite it
    # while still pointing to the same canonical video.
    return True


def sanitize_filename_token(text: str, fallback: str = "Unknown", max_len: int = 64) -> str:
    t = (text or "").strip()
    if not t:
        return fallback
    t = re.sub(r'[\\/:*?"<>|]+', "_", t)
    t = re.sub(r"\s+", "_", t)
    t = t.strip("._- ")
    if not t:
        return fallback
    return t[:max_len]


def is_unavailable_content_page(driver, source_url: str = "") -> bool:
    """
    Detect pages that opened but content is unavailable/private/deleted.
    This prevents saving blank/blocked screenshots as successful rows.
    """
    try:
        txt_raw = (driver.execute_script("return (document.body && document.body.innerText) ? document.body.innerText : ''") or "")
    except Exception:
        txt_raw = ""
    txt = str(txt_raw or "").lower()
    txt_norm = normalize_match_text(txt_raw or "")
    try:
        cur = (driver.current_url or "").lower()
    except Exception:
        cur = ""
    try:
        title = str(driver.title or "").lower()
    except Exception:
        title = ""
    url = (source_url or "").lower()

    # Facebook login gate popup (e.g. "See more on Facebook") should be treated
    # as unavailable for this workflow to avoid false "success" screenshots.
    fb_scope = "facebook.com" in (url + cur)
    if fb_scope:
        fb_login_markers_raw = [
            "see more on facebook",
            "xem them tren facebook",
            "email address or phone number",
            "forgotten password",
            "create new account",
            "scan the qr code and confirm",
            "confirm that the codes match to log in",
            "dang nhap",
            "log in",
            "login",
        ]
        fb_login_markers_norm = [normalize_match_text(m) for m in fb_login_markers_raw]
        marker_hits = 0
        for marker_raw, marker_norm in zip(fb_login_markers_raw, fb_login_markers_norm):
            if marker_raw in txt or (marker_norm and marker_norm in txt_norm):
                marker_hits += 1
        has_password_field = False
        try:
            has_password_field = bool(
                driver.execute_script(
                    "return !!document.querySelector(\"input[type='password'], input[name='pass'], input#pass\");"
                )
            )
        except Exception:
            has_password_field = False
        if marker_hits >= 2 and has_password_field:
            return True

    # TikTok onboarding/login interest modal should be treated as unavailable
    # because it blocks post content extraction/screenshot quality.
    tiktok_scope = "tiktok.com" in (url + cur)
    if tiktok_scope:
        if is_tiktok_shop_app_only_notice(driver, source_url):
            return True
        tiktok_gate_markers_raw = [
            "what would you like to watch on tiktok",
            "what would you like to watch",
            "continue (0/3)",
            "continue 0/3",
            "continue ( 0/3 )",
            "by continuing with an account located in",
        ]
        tiktok_gate_markers_norm = [normalize_match_text(m) for m in tiktok_gate_markers_raw]
        marker_hits = 0
        for marker_raw, marker_norm in zip(tiktok_gate_markers_raw, tiktok_gate_markers_norm):
            if marker_raw in txt or (marker_norm and marker_norm in txt_norm):
                marker_hits += 1
        has_gate_overlay = False
        try:
            has_gate_overlay = bool(
                driver.execute_script(
                    "return !!document.querySelector(\"[role='dialog'], div[class*='Modal'], div[class*='modal'], div[data-e2e*='modal']\");"
                )
            )
        except Exception:
            has_gate_overlay = False
        if marker_hits >= 2 or (marker_hits >= 1 and has_gate_overlay):
            return True

    markers_raw = [
        "bạn hiện không xem được nội dung này",
        "không xem được nội dung này",
        "nội dung này hiện không khả dụng",
        "nội dung không khả dụng",
        "không có nội dung",
        "bài viết này hiện không còn",
        "bài viết này không còn khả dụng",
        "trang này hiện không khả dụng",
        "liên kết này có thể đã bị hỏng",
        "nội dung này đã bị gỡ",
        "video này hiện không khả dụng",
        "khong xem duoc noi dung nay",
        "noi dung khong kha dung",
        "khong co noi dung",
        "this content isn't available",
        "this page isn't available",
        "content isn't available right now",
        "this post is no longer available",
        "the page isn't available",
        "this video is unavailable",
        "video not available",
        "this video is not available",
        "video is not available",
        "video currently unavailable",
        "this video is currently unavailable",
        "looking for videos? try browsing our trending creators",
        "no content available",
        "you cannot view this content",
        "access to www.tiktok.com was denied",
        "access to tiktok.com was denied",
        "you don't have authorization to view this page",
        "you do not have authorization to view this page",
        "http error 403",
        "err_blocked_by_client",
        "error 403",
        "access denied",
    ]
    markers_norm = [normalize_match_text(m) for m in markers_raw]
    if any((m in txt) or (normalize_match_text(m) in txt_norm) for m in markers_raw):
        return True
    if any(mn and mn in txt_norm for mn in markers_norm):
        return True
    title_markers = [
        "access denied",
        "http error 403",
        "error 403",
        "forbidden",
    ]
    if any(tm in title for tm in title_markers):
        if "tiktok.com" in (url + cur + txt):
            return True

    # Common Facebook dead-end routes.
    if fb_scope:
        dead_routes = ["/checkpoint/", "/login/", "/login.php", "/recover/"]
        if any(r in cur for r in dead_routes):
            return True

    return False


def is_tiktok_shop_app_only_notice(driver, source_url: str = "") -> bool:
    """
    Detect TikTok Shop pages that can only be viewed in TikTok app.
    These pages should be treated as unavailable for web screenshot workflow.
    """
    try:
        txt_raw = (
            driver.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:
        txt_raw = ""
    txt = str(txt_raw or "").lower()
    txt_norm = normalize_match_text(txt_raw or "")
    try:
        cur = str(driver.current_url or "").lower()
    except Exception:
        cur = ""
    src = str(source_url or "").lower()
    scope = f"{src} {cur} {txt}"
    if "tiktok.com" not in scope:
        return False
    markers_raw = [
        "view tiktok shop videos in the tiktok app",
        "view tiktok shop videos in the app",
        "xem video tiktok shop tren ung dung tiktok",
        "xem video tiktok shop trong ung dung tiktok",
    ]
    markers_norm = [normalize_match_text(m) for m in markers_raw]
    if any(m in txt for m in markers_raw):
        return True
    if any(mn and mn in txt_norm for mn in markers_norm):
        return True
    return False


def is_tiktok_access_denied_page(driver, source_url: str = "") -> bool:
    try:
        txt_raw = (
            driver.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:
        txt_raw = ""
    txt = str(txt_raw or "").lower()
    txt_norm = normalize_match_text(txt_raw or "")
    try:
        cur = str(driver.current_url or "").lower()
    except Exception:
        cur = ""
    try:
        title = str(driver.title or "").lower()
    except Exception:
        title = ""
    src = str(source_url or "").lower()
    scope = f"{src} {cur} {title} {txt}"
    if "tiktok.com" not in scope:
        return False
    markers_raw = [
        "access to www.tiktok.com was denied",
        "access to tiktok.com was denied",
        "you don't have authorization to view this page",
        "you do not have authorization to view this page",
        "http error 403",
        "error 403",
        "access denied",
        "reference #18.",
    ]
    markers_norm = [normalize_match_text(m) for m in markers_raw]
    if any(m in txt for m in markers_raw):
        return True
    if any(m in title for m in ("access denied", "http error 403", "error 403", "forbidden")):
        return True
    if any(mn and mn in txt_norm for mn in markers_norm):
        return True
    return False


def wait_tiktok_redirect_ready(
    driver,
    requested_url: str,
    timeout_sec: float = TIKTOK_REDIRECT_WAIT_SEC,
) -> str:
    deadline = time.time() + max(0.5, float(timeout_sec or TIKTOK_REDIRECT_WAIT_SEC))
    last_url = ""
    stable_hits = 0
    while time.time() < deadline:
        try:
            cur = str(driver.current_url or "").strip()
        except Exception:
            cur = ""
        if cur:
            if cur == last_url:
                stable_hits += 1
            else:
                stable_hits = 1
                last_url = cur
            # url stops changing for a short window => redirect likely completed
            if stable_hits >= 2:
                return cur
            # if final page already matches expected video id, no need to wait more
            if _is_expected_tiktok_page(requested_url, cur):
                return cur
        time.sleep(0.35)
    return last_url


def has_please_wait_overlay(driver) -> bool:
    try:
        txt_raw = (
            driver.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:
        txt_raw = ""
    txt_norm = normalize_match_text(txt_raw or "")
    markers = [
        "please wait",
        "please wait...",
        "vui long cho",
        "vui long cho...",
        "vui lòng chờ",
        "vui lòng chờ...",
    ]
    return any(normalize_match_text(marker) in txt_norm for marker in markers)


def wait_for_please_wait_clear(
    driver,
    timeout_sec: float = PLEASE_WAIT_MAX_WAIT_SEC,
    poll_sec: float = PLEASE_WAIT_POLL_SEC,
) -> tuple[bool, float]:
    if not has_please_wait_overlay(driver):
        return True, 0.0
    start_ts = time.time()
    deadline = start_ts + max(0.5, float(timeout_sec or PLEASE_WAIT_MAX_WAIT_SEC))
    interval = max(0.2, float(poll_sec or PLEASE_WAIT_POLL_SEC))
    while time.time() < deadline:
        time.sleep(interval)
        if not has_please_wait_overlay(driver):
            return True, max(0.0, time.time() - start_ts)
    return False, max(0.0, time.time() - start_ts)


def bring_current_tab_to_front(driver):
    try:
        driver.execute_cdp_cmd("Page.bringToFront", {})
    except Exception:
        pass


def focus_browser_window_os(title_hint: str = "") -> bool:
    if os.name != "nt":
        return False
    title_hint = str(title_hint or "").strip()
    title_expr = title_hint.replace("'", "''")
    script = """
$ws = New-Object -ComObject WScript.Shell
$targets = @()
if ('__TITLE__'.Length -gt 0) {
  $targets += '__TITLE__'
}
$targets += @('TikTok', 'Google Chrome', 'Chrome', 'Microsoft Edge', 'Edge')
foreach ($t in $targets) {
  try {
    if ($ws.AppActivate($t)) { exit 0 }
  } catch {}
}
exit 1
"""
    script = script.replace("__TITLE__", title_expr)
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=4,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def focus_chrome_for_tiktok_challenge(driver) -> bool:
    if not bool(TIKTOK_CAPTCHA_FORCE_FOCUS):
        return False
    bring_current_tab_to_front(driver)
    page_title = ""
    try:
        page_title = str(driver.title or "").strip()
    except Exception:
        page_title = ""
    if focus_browser_window_os(page_title):
        return True
    return focus_browser_window_os("TikTok")


def is_tiktok_slider_challenge_present(driver) -> bool:
    text_raw = ""
    try:
        text_raw = (
            driver.execute_script(
                "return (document.body && document.body.innerText) ? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:
        text_raw = ""
    text_l = str(text_raw or "").lower()
    text_norm = normalize_match_text(text_raw or "")
    markers_raw = [
        "drag the slider",
        "fit the puzzle",
        "security verification",
        "verify to continue",
        "kéo thanh trượt",
        "xác minh",
        "vui lòng kéo",
    ]
    markers_norm = [normalize_match_text(m) for m in markers_raw]
    if any(m in text_l for m in markers_raw):
        return True
    if any(mn and mn in text_norm for mn in markers_norm):
        return True

    selectors = [
        "[data-e2e*='captcha']",
        "[data-testid*='captcha']",
        "iframe[src*='captcha']",
        "div[class*='captcha']",
        "div[id*='captcha']",
    ]
    try:
        for sel in selectors:
            if driver.find_elements(By.CSS_SELECTOR, sel):
                return True
    except Exception:
        pass
    return False


def wait_for_tiktok_slider_clear(
    driver,
    max_wait_sec: float = TIKTOK_CAPTCHA_MAX_WAIT_SEC,
    poll_sec: float = TIKTOK_CAPTCHA_POLL_SEC,
) -> tuple[bool, float]:
    start = time.time()
    if not is_tiktok_slider_challenge_present(driver):
        return True, 0.0
    if bool(TIKTOK_CAPTCHA_FORCE_FOCUS):
        focus_chrome_for_tiktok_challenge(driver)
    last_bring = start
    while (time.time() - start) < max(1.0, float(max_wait_sec or 0)):
        time.sleep(max(0.2, float(poll_sec or 0.5)))
        if not is_tiktok_slider_challenge_present(driver):
            return True, time.time() - start
        now = time.time()
        if bool(TIKTOK_CAPTCHA_FORCE_FOCUS) and (now - last_bring) >= TIKTOK_BRING_TO_FRONT_INTERVAL_SEC:
            focus_chrome_for_tiktok_challenge(driver)
            last_bring = now
    return False, time.time() - start


def write_colored_xlsx_builtin(path: str, headers: list[str], rows_with_tags: list[tuple[list, list]]):
    """
    Create a minimal .xlsx with row background colors using only stdlib.
    Styles:
    - 0: default
    - 1: ok
    - 2: fail
    - 3: unavailable
    """
    def col_name(idx: int) -> str:
        n = idx + 1
        out = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            out = chr(65 + r) + out
        return out

    def style_id_for_row(vals: list, tags: list) -> int:
        tag_set = set(tags or [])
        state = str(vals[3]).strip().upper() if len(vals) > 3 else ""
        msg = str(vals[4]).lower() if len(vals) > 4 else ""
        if "fail" in tag_set or state == "FAIL":
            return 2
        if "unavailable" in tag_set or "nội dung không khả dụng" in msg:
            return 3
        if "ok" in tag_set or state == "OK":
            return 1
        return 0

    rows_xml = []
    # Header row
    header_cells = []
    for c, h in enumerate(headers):
        ref = f"{col_name(c)}1"
        header_cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{xml_escape(str(h))}</t></is></c>')
    rows_xml.append(f'<row r="1">{"".join(header_cells)}</row>')

    # Data rows
    for i, (vals, tags) in enumerate(rows_with_tags, start=2):
        s_id = style_id_for_row(vals, tags)
        cells = []
        for c, v in enumerate(vals):
            ref = f"{col_name(c)}{i}"
            txt = xml_escape(str(v))
            if s_id > 0:
                cells.append(f'<c r="{ref}" s="{s_id}" t="inlineStr"><is><t>{txt}</t></is></c>')
            else:
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{txt}</t></is></c>')
        rows_xml.append(f'<row r="{i}">{"".join(cells)}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(rows_xml)}</sheetData>'
        '</worksheet>'
    )

    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="5">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFD8F3DC"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFD9D9"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFE6C7"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="4">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="0" xfId="0" applyFill="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>
"""

    workbook_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Log" sheetId="1" r:id="rId1"/></sheets>
</workbook>
"""

    rels_root = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>
"""

    rels_workbook = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>
"""

    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>
"""

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels_root)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", rels_workbook)
        zf.writestr("xl/styles.xml", styles_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

# ================= UI =================
class ProgressApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Evidence Tool")
        self.root.geometry(self._get_initial_geometry())
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        min_w = min(1220, max(980, screen_w - 140))
        min_h = min(860, max(720, screen_h - 180))
        self.root.minsize(min_w, min_h)
        self.root.configure(bg="#f3f4f7")
        self.is_running = True
        self.is_paused = False
        self.driver = None

        self.main_canvas = tk.Canvas(self.root, bg="#f3f4f7", highlightthickness=0)
        self.v_scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        self.main_canvas.configure(yscrollcommand=self.v_scrollbar.set)
        self.v_scrollbar.pack(side="right", fill="y")
        self.main_canvas.pack(side="left", fill="both", expand=True)

        self.main_frame = tk.Frame(self.main_canvas, bg="#ffffff")
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        self.main_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )
        self.main_canvas.bind("<Configure>", self._on_canvas_configure)
        self._bind_scroll_events()

        self.force_run_all = tk.BooleanVar(value=False)
        self.only_run_error_rows = tk.BooleanVar(value=False)
        self.target_rows_var = tk.StringVar(value="")
        self.auto_launch_chrome = tk.BooleanVar(value=True)
        self.capture_five_per_link = tk.BooleanVar(value=False)
        self.mapping_mode_var = tk.StringVar(value="Seeding")
        self.sheet_url_var = tk.StringVar(value=DEFAULT_SHEET_URL)
        self.sheet_name_var = tk.StringVar(value=DEFAULT_SHEET_NAME_TARGET)
        self.drive_id_var = tk.StringVar(value=DEFAULT_DRIVE_FOLDER_ID)
        self.credentials_path_var = tk.StringVar(value=get_default_credentials_input())
        self.mapping_blocks = []
        self.mapping_blocks_by_mode: dict[str, list[dict]] = {}
        self._active_mapping_mode = "Seeding"
        self._is_loading_settings = False
        self.mapping_entries = []
        self.mapping_remove_buttons = []
        self.mapping_launch_buttons = []
        self.chk_capture5 = None
        self.btn_add_block = None
        self.load_settings()
        if not self.mapping_blocks:
            self._ensure_default_mapping_blocks()
        self._build_menu()
        self.main_frame.configure(bg="#f3f4f7")

        # Header
        header = tk.Frame(self.main_frame, bg="#f7f7fa", relief="ridge", bd=1, padx=10, pady=6)
        header.pack(fill="x", padx=12, pady=(8, 8))

        self.label_status = tk.Label(
            header, text="● STATUS: READY",
            font=("Arial", 10, "bold"),
            bg="#f7f7fa", fg="#2e7d32", anchor="w"
        )
        self.label_status.pack(side="left")

        self.reload_btn = tk.Button(
            header, text="⟳", command=self.reload_app,
            width=4, bg="#eeeeee", fg="#444444"
        )
        self.reload_btn.pack(side="right", padx=(4, 0))

        self.pause_btn = tk.Button(
            header, text="⏸", command=self.toggle_pause,
            width=4, bg="#fff3cd", fg="#ff6b6b", state="disabled"
        )
        self.pause_btn.pack(side="right", padx=(4, 0))

        self.save_btn = tk.Button(
            header, text="Save Config", command=self.save_settings,
            width=10, bg="#e6f4ea", fg="#137333"
        )
        self.save_btn.pack(side="right", padx=(4, 0))

        # Content area
        content = tk.Frame(self.main_frame, bg="#f3f4f7")
        content.pack(fill="x", padx=12, pady=4)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        left_card = tk.LabelFrame(content, text="DATA SOURCE", bg="#f7f7fa", fg="#4a4a4a", padx=8, pady=8)
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        right_card = tk.LabelFrame(content, text="COLUMN MAPPING", bg="#f7f7fa", fg="#4a4a4a", padx=8, pady=8)
        right_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        def add_source_row(parent, r, label, text_var, btn_var_name=None):
            tk.Label(parent, text=label, bg="#f7f7fa", anchor="w", width=12).grid(row=r, column=0, sticky="w", pady=2)
            ent = tk.Entry(parent, textvariable=text_var, width=34)
            ent.grid(row=r, column=1, sticky="ew", padx=4, pady=2)
            if btn_var_name:
                btn = tk.Button(parent, text="DÁN", width=6, command=lambda v=text_var: self.paste_to(v))
                btn.grid(row=r, column=2, pady=2)
                setattr(self, btn_var_name, btn)
            return ent

        left_card.grid_columnconfigure(1, weight=1)
        self.entry_sheet_url = add_source_row(left_card, 0, "Sheet URL", self.sheet_url_var, "btn_paste_sheet_url")
        tk.Label(left_card, text="Rows to rerun", bg="#f7f7fa", anchor="w", width=12).grid(row=1, column=0, sticky="w", pady=2)
        rows_filter_source_row = tk.Frame(left_card, bg="#f7f7fa")
        rows_filter_source_row.grid(row=1, column=1, columnspan=2, sticky="ew", padx=4, pady=2)
        self.entry_target_rows = tk.Entry(rows_filter_source_row, textvariable=self.target_rows_var, width=34)
        self.entry_target_rows.pack(side="left", fill="x", expand=True)
        self.btn_clear_target_rows = tk.Button(
            rows_filter_source_row,
            text="Clear",
            width=6,
            command=lambda: self.target_rows_var.set(""),
        )
        self.btn_clear_target_rows.pack(side="left", padx=(4, 0))
        tk.Label(
            left_card,
            text="Ví dụ: 7,8,10-15 (để trống = chạy toàn bộ)",
            bg="#f7f7fa",
            fg="#6b7280",
            anchor="w",
            font=("Arial", 9),
        ).grid(row=2, column=1, columnspan=2, sticky="w", padx=4, pady=(0, 2))
        self.entry_sheet_name = add_source_row(left_card, 3, "Sheet Name", self.sheet_name_var, "btn_paste_sheet_name")
        self.entry_drive_id = add_source_row(left_card, 4, "Drive Folder", self.drive_id_var, "btn_paste_drive_id")
        self.entry_credentials_path = add_source_row(left_card, 5, "Credentials", self.credentials_path_var, None)

        mode_row = tk.Frame(right_card, bg="#f7f7fa")
        mode_row.pack(fill="x", pady=(0, 4))
        tk.Label(mode_row, text="Mode:", bg="#f7f7fa", anchor="w", width=12).pack(side="left")
        self.mapping_mode_combo = ttk.Combobox(
            mode_row,
            textvariable=self.mapping_mode_var,
            values=("Seeding", "Booking", "Scan"),
            state="readonly",
            width=14,
        )
        self.mapping_mode_combo.pack(side="left")
        self.mapping_mode_combo.bind("<<ComboboxSelected>>", self._on_mode_changed)
        self.mapping_header = tk.Frame(right_card, bg="#f7f7fa")
        self.mapping_header.pack(fill="x", pady=(0, 6))
        self.mapping_grid = tk.Frame(right_card, bg="#f7f7fa")
        self.mapping_grid.pack(fill="x")
        self._render_mapping_blocks()

        run_mode = tk.LabelFrame(content, text="RUN MODE", bg="#f7f7fa", fg="#4a4a4a", padx=8, pady=8)
        run_mode.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 2), padx=0)
        self.checkbox = tk.Checkbutton(
            run_mode, text="Run All (overwrite)", variable=self.force_run_all,
            bg="#f7f7fa", anchor="w"
        )
        self.checkbox.pack(anchor="w")
        self.checkbox_errors_only = tk.Checkbutton(
            run_mode, text="Retry Failed Only", variable=self.only_run_error_rows,
            bg="#f7f7fa", anchor="w"
        )
        self.checkbox_errors_only.pack(anchor="w")
        self.action_row = tk.Frame(run_mode, bg="#f7f7fa")
        self.action_row.pack(fill="x", pady=(8, 0))

        self.start_btn = tk.Button(
            self.action_row, text="▶ RUN", command=self.start_processing,
            width=11,
            bg="#2f80ed",
            fg="#ffffff",
            activebackground="#2f80ed",
            activeforeground="#ffffff",
            disabledforeground="#ffffff",
            relief="raised",
            overrelief="raised",
            bd=1,
            highlightthickness=0,
        )
        self.start_btn.pack(side="left", padx=2)

        self.export_log_btn = tk.Button(
            self.action_row, text="Export Log", command=self.export_live_log_excel,
            width=10, bg="#e8f0fe", fg="#1a3d8f"
        )
        self.export_log_btn.pack(side="left", padx=2)

        # Share + Progress
        share_frame = tk.Frame(self.main_frame, bg="#eef3ff", relief="ridge", bd=1, padx=8, pady=6)
        share_frame.pack(fill="x", pady=8, padx=12)
        tk.Label(
            share_frame, text="Chia sẻ Sheet & Drive folder cho (quyền Editor):",
            bg="#eef3ff", fg="#1a73e8", font=("Arial", 9, "bold"), anchor="w"
        ).pack(anchor="w")
        email = get_service_account_email(self.credentials_path_var.get().strip()) or "link-verification@hazel-tea-485816-u3.iam.gserviceaccount.com"
        self.share_email_var = tk.StringVar(value=email)
        row_share = tk.Frame(share_frame, bg="#eef3ff")
        row_share.pack(fill="x", pady=4)
        self.entry_share_email = tk.Entry(row_share, textvariable=self.share_email_var, state="readonly")
        self.entry_share_email.pack(side="left", fill="x", expand=True, padx=(0, 4))
        tk.Button(row_share, text="Copy", width=6, command=lambda: self.copy_share_email()).pack(side="left")

        self.error_card = tk.Frame(self.main_frame, bg="#f7f7fa", relief="ridge", bd=1)
        self.error_card.pack(fill="x", padx=12, pady=(0, 8))

        self.error_card_header = tk.Frame(self.error_card, bg="#f7f7fa")
        self.error_card_header.pack(fill="x", padx=8, pady=(6, 4))

        self.error_header_var = tk.StringVar(value="⚠ Lỗi theo link Sheet: chưa có")
        self.error_header_label = tk.Label(
            self.error_card_header,
            textvariable=self.error_header_var,
            font=("Arial", 11, "bold"),
            bg="#f7f7fa",
            fg="#2f3b52",
            anchor="w",
        )
        self.error_header_label.pack(side="left", fill="x", expand=True)

        self.error_card_save_btn = tk.Button(
            self.error_card_header,
            text="💾 Save",
            width=9,
            command=self._save_error_history_current_sheet,
            bg="#eef2ff",
            fg="#2f3b52",
        )
        self.error_card_save_btn.pack(side="right")

        self.error_card_clear_btn = tk.Button(
            self.error_card_header,
            text="🗑 Xóa",
            width=9,
            command=self._clear_current_sheet_error_history,
            bg="#ffe9e9",
            fg="#8a1c1c",
        )
        self.error_card_clear_btn.pack(side="right", padx=(0, 6))

        self.error_rows_var = tk.StringVar(value="• Chưa có dòng lỗi đã lưu.")
        self.error_rows_frame = tk.Frame(self.error_card, bg="#f7f7fa")
        self.error_rows_frame.pack(fill="x", padx=14, pady=(0, 6))
        self.error_rows_text = tk.Text(
            self.error_rows_frame,
            height=5,
            wrap="word",
            font=("Arial", 10),
            bg="#f7f7fa",
            fg="#5f6673",
            relief="flat",
            borderwidth=0,
            highlightthickness=0,
        )
        self.error_rows_text.pack(side="left", fill="both", expand=True)
        self.error_rows_scroll = ttk.Scrollbar(
            self.error_rows_frame,
            orient="vertical",
            command=self.error_rows_text.yview,
        )
        self.error_rows_scroll.pack(side="right", fill="y")
        self.error_rows_text.configure(yscrollcommand=self.error_rows_scroll.set)
        self.error_rows_text.insert("1.0", self.error_rows_var.get())
        self.error_rows_text.config(state="disabled")

        self.error_sep = tk.Frame(self.error_card, bg="#d9dde7", height=1)
        self.error_sep.pack(fill="x", padx=10, pady=(0, 6))

        self.progress_summary_var = tk.StringVar(value="✔ Progress: 0/0 | Success: 0 | Failed: 0 | ETA: ---")
        self.progress_summary_label = tk.Label(
            self.error_card,
            textvariable=self.progress_summary_var,
            font=("Arial", 10, "bold"),
            bg="#f7f7fa",
            fg="#2f3b52",
            anchor="w",
        )
        self.progress_summary_label.pack(fill="x", padx=14, pady=(0, 4))

        self.progress = ttk.Progressbar(self.error_card, orient="horizontal", length=560, mode="determinate")
        self.progress.pack(pady=(0, 8), padx=12, fill="x")

        self.live_log_frame = tk.Frame(self.main_frame, bg="#f3f4f7")
        self.live_log_frame.pack(fill="both", padx=12, pady=(6, 6), expand=False)
        self.live_log_table = ttk.Treeview(
            self.live_log_frame,
            columns=("time", "row", "s1", "s2", "msg"),
            show="headings",
            height=6,
        )
        self.live_log_table.heading("time", text="Time")
        self.live_log_table.heading("row", text="#")
        self.live_log_table.heading("s1", text="State")
        self.live_log_table.heading("s2", text="Result")
        self.live_log_table.heading("msg", text="Message")
        self.live_log_table.column("time", width=74, anchor="w")
        self.live_log_table.column("row", width=44, anchor="center")
        self.live_log_table.column("s1", width=72, anchor="center")
        self.live_log_table.column("s2", width=72, anchor="center")
        self.live_log_table.column("msg", width=360, anchor="w")
        self.live_log_table.tag_configure("start")
        self.live_log_table.tag_configure("ok", background="#d8f3dc", foreground="#1f3a2a")
        self.live_log_table.tag_configure("unavailable", background="#ffe6c7", foreground="#4a3820")
        self.live_log_table.tag_configure("fail", background="#ffd9d9", foreground="#4a1f1f")
        self.live_log_table.pack(side="left", fill="x", expand=True)
        self.live_log_scroll = ttk.Scrollbar(self.live_log_frame, orient="vertical", command=self.live_log_table.yview)
        self.live_log_table.configure(yscrollcommand=self.live_log_scroll.set)
        self.live_log_scroll.pack(side="right", fill="y")

        self.label_detail = tk.Label(
            self.main_frame, text="No activity yet...",
            font=("Arial", 9),
            bg="#f3f4f7", fg="#777777", wraplength=560, anchor="w", justify="left"
        )
        self.label_detail.pack(fill="x", padx=12, pady=(0, 8))

        self.exit_btn = tk.Button(self.main_frame, text="THOÁT", command=self.exit_app, width=12, bg="#f0f0f0")
        self.exit_btn.pack(pady=8)

        # Keep this list for menu history actions
        self._history_sheet_items = []
        self.live_error_details: dict[int, str] = {}

        self.sheet_url_var.trace_add("write", lambda *_: self.refresh_error_history_ui())
        self.sheet_name_var.trace_add("write", lambda *_: self.refresh_error_history_ui())
        self.refresh_error_history_ui()
        self.refresh_saved_sheets_list()

    def _new_block_vars(self, data: dict | None = None) -> dict:
        d = data or {}
        return {
            "name_var": tk.StringVar(value=str(d.get("name", "Post")).strip() or "Post"),
            "manual_link_var": tk.StringVar(value=str(d.get("manual_link", "")).strip()),
            "col_profile_var": tk.StringVar(value=str(d.get("col_profile", "")).strip().upper()),
            "col_content_var": tk.StringVar(value=str(d.get("col_content", "")).strip().upper()),
            "col_url_var": tk.StringVar(value=str(d.get("col_url", "")).strip().upper()),
            "col_drive_var": tk.StringVar(value=str(d.get("col_drive", "")).strip().upper()),
            "col_screenshot_var": tk.StringVar(value=str(d.get("col_screenshot", "")).strip().upper()),
            "col_air_date_var": tk.StringVar(value=str(d.get("col_air_date", "")).strip().upper()),
            "start_line_var": tk.StringVar(value=str(d.get("start_line", "4")).strip() or "4"),
        }

    def _ensure_default_mapping_blocks(self):
        defaults = [
            {"name": "Post 1", "col_profile": "C", "col_content": "D", "col_url": "E", "col_drive": "F", "col_screenshot": "G", "col_air_date": "", "start_line": "4"},
        ]
        self.mapping_blocks = [self._new_block_vars(x) for x in defaults]

    def _normalize_mode_name(self, mode_text: str) -> str:
        s = str(mode_text or "").strip().lower()
        if s in ("scan only text", "scan_only_text", "scan text", "text scan"):
            return "Scan"
        if s == "scan":
            return "Scan"
        if s == "booking":
            return "Booking"
        return "Seeding"

    def _default_mapping_configs_for_mode(self, mode_name: str) -> list[dict]:
        mode = self._normalize_mode_name(mode_name)
        if mode == "Scan":
            return [
                {
                    "name": "Scan 1",
                    "manual_link": "",
                    "col_profile": "",
                    "col_content": "E",
                    "col_url": "F",
                    "col_drive": "G",
                    "col_screenshot": "",
                    "col_air_date": "",
                    "start_line": "4",
                }
            ]
        return [
            {
                "name": "Post 1",
                "col_profile": "C",
                "col_content": "D",
                "col_url": "E",
                "col_drive": "F",
                "col_screenshot": "G",
                "col_air_date": "",
                "start_line": "4",
            }
        ]

    def _snapshot_current_mode_configs(self):
        mode = self._normalize_mode_name(getattr(self, "_active_mapping_mode", self.mapping_mode_var.get()))
        self.mapping_blocks_by_mode[mode] = self.get_mapping_configs()

    def _on_mode_changed(self, _event=None):
        new_mode = self._normalize_mode_name(self.mapping_mode_var.get())
        if self._is_loading_settings:
            self._active_mapping_mode = new_mode
            self._render_mapping_blocks()
            return
        self._snapshot_current_mode_configs()
        if not self.mapping_blocks_by_mode.get(new_mode):
            self.mapping_blocks_by_mode[new_mode] = self._default_mapping_configs_for_mode(new_mode)
        self._active_mapping_mode = new_mode
        self._load_mapping_blocks(self.mapping_blocks_by_mode.get(new_mode) or [], render=True)

    def _add_mapping_block(self, seed: dict | None = None):
        idx = len(self.mapping_blocks) + 1
        mode_name = self._normalize_mode_name(self.mapping_mode_var.get())
        if mode_name == "Scan":
            block_seed = {
                "name": f"Scan {idx}",
                "manual_link": "",
                "col_profile": "",
                "col_content": "",
                "col_url": "",
                "col_drive": "",
                "col_screenshot": "",
                "col_air_date": "",
                "start_line": "4",
            }
        else:
            block_seed = {"name": f"Post {idx}", "start_line": "4"}
        if seed:
            block_seed.update(seed)
        self.mapping_blocks.append(self._new_block_vars(block_seed))
        self._render_mapping_blocks()

    def _get_block_port(self, idx: int) -> int:
        return get_post_port(idx, 9223)

    def _get_block_profile(self, idx: int) -> str:
        if idx <= 0:
            return LOCAL_PROFILE_PATH
        return os.path.join(TEMP_DIR, f"chrome_profile_worker_{idx}")

    def launch_chrome_for_block(self, idx: int):
        port = self._get_block_port(idx)
        profile = self._get_block_profile(idx)
        try:
            os.makedirs(profile, exist_ok=True)
        except Exception:
            pass
        ok, info = launch_chrome_for_login(port, profile_path=profile)
        block_name = f"Post {idx + 1}"
        if 0 <= idx < len(self.mapping_blocks):
            block_name = (self.mapping_blocks[idx]["name_var"].get() or block_name).strip() or block_name
        if ok:
            messagebox.showinfo(
                "Chrome đã mở",
                f"{block_name} mở Chrome ở port {port}.\n\nBạn đăng nhập xong rồi bấm RUN."
            )
        else:
            messagebox.showerror("Lỗi", f"Không mở được Chrome cho {block_name}: {info}")

    def _remove_mapping_block(self, idx: int):
        if len(self.mapping_blocks) <= 1:
            return
        if 0 <= idx < len(self.mapping_blocks):
            self.mapping_blocks.pop(idx)
        self._render_mapping_blocks()

    def _pick_air_date_for_block(self, idx: int, anchor_widget=None):
        if tk is None:
            return
        if not (0 <= idx < len(self.mapping_blocks)):
            return
        block = self.mapping_blocks[idx]
        var = block.get("col_air_date_var")
        if var is None:
            return

        now = datetime.now()
        current = str(var.get() or "").strip()
        parsed = get_air_date_token(current)
        selected_token = ""
        if parsed and len(parsed) == 8 and parsed.isdigit():
            try:
                now = datetime(int(parsed[:4]), int(parsed[4:6]), int(parsed[6:8]))
                selected_token = parsed
            except Exception:
                pass
        if not selected_token:
            selected_token = now.strftime("%Y%m%d")

        win = tk.Toplevel(self.root)
        win.title("Chọn Air Date")
        win.resizable(False, False)
        win.transient(self.root)

        frm = tk.Frame(win, padx=10, pady=10, bg="#ffffff")
        frm.pack(fill="both", expand=True)

        cursor_year = now.year
        cursor_month = now.month
        month_text = tk.StringVar()

        header = tk.Frame(frm, bg="#ffffff")
        header.pack(fill="x")

        weekday_row = tk.Frame(frm, bg="#ffffff")
        weekday_row.pack(fill="x", pady=(6, 2))

        days_frame = tk.Frame(frm, bg="#ffffff")
        days_frame.pack(fill="both", expand=True)

        footer = tk.Frame(frm, bg="#ffffff")
        footer.pack(fill="x", pady=(8, 0))

        def _set_date(y: int, m: int, d: int):
            var.set(f"{y:04d}-{m:02d}-{d:02d}")
            win.destroy()

        def _clear():
            var.set("")
            win.destroy()

        def _today():
            dt = datetime.now()
            _set_date(dt.year, dt.month, dt.day)

        def _render_calendar():
            month_text.set(f"Tháng {cursor_month:02d}/{cursor_year}")

            for c in days_frame.winfo_children():
                c.destroy()

            first_weekday, days_in_month = calendar.monthrange(cursor_year, cursor_month)
            day = 1
            for r in range(6):
                for c in range(7):
                    cell = r * 7 + c
                    if cell < first_weekday or day > days_in_month:
                        tk.Label(days_frame, text=" ", width=4, bg="#ffffff").grid(row=r, column=c, padx=1, pady=1)
                    else:
                        token = f"{cursor_year:04d}{cursor_month:02d}{day:02d}"
                        bg = "#2f80ed" if token == selected_token else "#f5f5f5"
                        fg = "#ffffff" if token == selected_token else "#222222"
                        btn = tk.Button(
                            days_frame,
                            text=str(day),
                            width=4,
                            bg=bg,
                            fg=fg,
                            relief="flat",
                            command=lambda dd=day: _set_date(cursor_year, cursor_month, dd),
                        )
                        btn.grid(row=r, column=c, padx=1, pady=1)
                        day += 1

        def _prev_month():
            nonlocal cursor_year, cursor_month
            cursor_month -= 1
            if cursor_month < 1:
                cursor_month = 12
                cursor_year -= 1
            _render_calendar()

        def _next_month():
            nonlocal cursor_year, cursor_month
            cursor_month += 1
            if cursor_month > 12:
                cursor_month = 1
                cursor_year += 1
            _render_calendar()

        tk.Button(header, text="◀", width=4, command=_prev_month).pack(side="left")
        tk.Label(header, textvariable=month_text, bg="#ffffff", font=("Arial", 10, "bold")).pack(side="left", expand=True)
        tk.Button(header, text="▶", width=4, command=_next_month).pack(side="right")

        for i, wd in enumerate(["T2", "T3", "T4", "T5", "T6", "T7", "CN"]):
            tk.Label(weekday_row, text=wd, width=4, bg="#ffffff", fg="#555555").grid(row=0, column=i, padx=1)

        tk.Button(footer, text="Xóa", width=8, command=_clear).pack(side="left")
        tk.Button(footer, text="Hôm nay", width=8, command=_today).pack(side="left", padx=(6, 0))
        tk.Button(footer, text="Đóng", width=8, command=win.destroy).pack(side="right")

        _render_calendar()

        # Place popup next to clicked button for quicker date picking.
        try:
            win.update_idletasks()
            popup_w = max(260, win.winfo_width())
            popup_h = max(220, win.winfo_height())
            screen_w = win.winfo_screenwidth()
            screen_h = win.winfo_screenheight()

            if anchor_widget is not None and anchor_widget.winfo_exists():
                ax = anchor_widget.winfo_rootx()
                ay = anchor_widget.winfo_rooty()
                aw = anchor_widget.winfo_width()
                ah = anchor_widget.winfo_height()
                x = ax + aw + 6
                y = ay
                # If right edge overflows, show on the left side of button.
                if x + popup_w > screen_w - 8:
                    x = max(8, ax - popup_w - 6)
                # Keep popup inside vertical bounds.
                if y + popup_h > screen_h - 8:
                    y = max(8, ay + ah - popup_h)
            else:
                # Fallback to cursor-near placement.
                x = min(max(8, self.root.winfo_pointerx() + 8), max(8, screen_w - popup_w - 8))
                y = min(max(8, self.root.winfo_pointery() - 10), max(8, screen_h - popup_h - 8))

            win.geometry(f"+{x}+{y}")
        except Exception:
            pass

        win.grab_set()

    def _render_mapping_blocks(self):
        for child in self.mapping_header.winfo_children():
            child.destroy()
        for child in self.mapping_grid.winfo_children():
            child.destroy()

        # Reset persisted grid column layout from previous mode renders
        # (especially Scan mode, which uses weighted columns).
        for i in range(0, 24):
            try:
                self.mapping_header.grid_columnconfigure(i, minsize=0, weight=0, uniform="")
            except Exception:
                pass
            try:
                self.mapping_grid.grid_columnconfigure(i, minsize=0, weight=0, uniform="")
            except Exception:
                pass

        self.mapping_entries = []
        self.mapping_remove_buttons = []
        self.mapping_launch_buttons = []
        self.chk_capture5 = None
        self.btn_add_block = None
        mode_name = (self.mapping_mode_var.get() or "Seeding").strip().lower()
        is_scan_mode = mode_name == "scan"
        is_scan_like_mode = is_scan_mode
        render_blocks = [(i, b) for i, b in enumerate(self.mapping_blocks)]

        self.mapping_header.grid_columnconfigure(0, minsize=96)
        self.mapping_grid.grid_columnconfigure(0, minsize=96)
        for col_idx in range(len(render_blocks)):
            self.mapping_header.grid_columnconfigure(col_idx + 1, minsize=96)
            self.mapping_grid.grid_columnconfigure(col_idx + 1, minsize=96)

        if is_scan_like_mode:
            self.mapping_header.pack_forget()
            cards_per_row = 2
            for c in range(cards_per_row):
                self.mapping_grid.grid_columnconfigure(c, weight=1, uniform="scan_cards")
            for view_idx, (block_idx, block) in enumerate(render_blocks):
                old_result_col = (block["col_screenshot_var"].get() or "").strip().upper()
                new_result_col = (block["col_drive_var"].get() or "").strip().upper()
                if (not new_result_col) and old_result_col:
                    block["col_drive_var"].set(old_result_col)

                card = tk.LabelFrame(
                    self.mapping_grid,
                    text=f"Scan {view_idx + 1}",
                    bg="#f7f7fa",
                    fg="#4a4a4a",
                    padx=8,
                    pady=6,
                )
                grid_row = view_idx // cards_per_row
                grid_col = view_idx % cards_per_row
                card.grid(row=grid_row, column=grid_col, sticky="nsew", padx=(0, 4), pady=(0, 6))
                card.grid_columnconfigure(0, minsize=88)
                card.grid_columnconfigure(1, minsize=130)

                labels = [
                    ("Tên Post", "name_var"),
                    ("Text Column", "col_content_var"),
                    ("Image Column", "col_url_var"),
                    ("Result Column", "col_drive_var"),
                    ("Start Line", "start_line_var"),
                ]
                for row_idx, (label_text, key_name) in enumerate(labels):
                    tk.Label(card, text=label_text, bg="#f7f7fa", anchor="w", width=12).grid(
                        row=row_idx, column=0, sticky="w", padx=(2, 4), pady=2
                    )
                    ent = tk.Entry(card, textvariable=block[key_name], width=12)
                    ent.grid(row=row_idx, column=1, sticky="w", padx=(0, 4), pady=2)
                    self.mapping_entries.append(ent)

                if len(render_blocks) > 1:
                    rm_btn = tk.Button(
                        card,
                        text="−",
                        width=2,
                        bg="#f8d7da",
                        fg="#9d2026",
                        command=lambda x=block_idx: self._remove_mapping_block(x),
                    )
                    rm_btn.place(relx=1.0, rely=0.0, anchor="ne", x=-6, y=6)
                    self.mapping_remove_buttons.append(rm_btn)

            self.btn_add_block = tk.Button(
                self.mapping_grid,
                text="+ Thêm Block",
                width=14,
                bg="#d9edf7",
                command=self._add_mapping_block,
            )
            button_row = (len(render_blocks) + cards_per_row - 1) // cards_per_row
            self.btn_add_block.grid(row=button_row, column=0, sticky="w", pady=(0, 2))
            return
        else:
            self.mapping_header.pack_forget()
            target_grid = tk.Frame(
                self.mapping_grid,
                bg="#f7f7fa",
                relief="groove",
                bd=1,
                padx=8,
                pady=6,
            )
            target_grid.grid(row=0, column=0, sticky="w", pady=(0, 4))
            target_grid.grid_columnconfigure(0, minsize=96)
            for col_idx in range(len(render_blocks)):
                target_grid.grid_columnconfigure(col_idx + 1, minsize=96)
            tk.Label(target_grid, text="Tên Post", bg="#f7f7fa", anchor="w", width=12).grid(
                row=0, column=0, sticky="w", pady=2
            )
            for block_idx, block in render_blocks:
                col_frame = tk.Frame(target_grid, bg="#f7f7fa")
                col_frame.grid(row=0, column=block_idx + 1, sticky="w", padx=2, pady=2)
                name_entry = tk.Entry(col_frame, textvariable=block["name_var"], width=12, justify="center")
                name_entry.grid(row=0, column=1, padx=(0, 2))
                rm_btn = tk.Button(
                    col_frame,
                    text="−",
                    width=2,
                    bg="#f8d7da",
                    fg="#9d2026",
                    command=lambda x=block_idx: self._remove_mapping_block(x),
                )
                rm_btn.grid(row=0, column=2)
                self.mapping_entries.append(name_entry)
                self.mapping_remove_buttons.append(rm_btn)

            self.btn_add_block = tk.Button(
                self.mapping_grid,
                text="+ Thêm Block",
                width=14,
                bg="#d9edf7",
                command=self._add_mapping_block,
            )
            self.btn_add_block.grid(row=1, column=0, sticky="w", pady=(0, 2))
            if mode_name == "booking":
                self.mapping_grid.grid_columnconfigure(1, weight=1)
                self.chk_capture5 = tk.Checkbutton(
                    self.mapping_grid,
                    text="Chụp 5 tấm / 1 link",
                    variable=self.capture_five_per_link,
                    bg="#f7f7fa",
                    anchor="w",
                )
                self.chk_capture5.grid(row=1, column=1, sticky="e", padx=(6, 2), pady=(0, 2))
            if mode_name == "seeding":
                labels = [
                    ("Air Date", "col_air_date_var"),
                    ("Link URL", "col_url_var"),
                    ("Drive URL", "col_drive_var"),
                    ("Screenshot", "col_screenshot_var"),
                    ("Start Line", "start_line_var"),
                ]
            else:
                labels = [
                    ("Air Date", "col_air_date_var"),
                    ("Profile", "col_profile_var"),
                    ("Content", "col_content_var"),
                    ("Link URL", "col_url_var"),
                    ("Drive URL", "col_drive_var"),
                    ("Screenshot", "col_screenshot_var"),
                    ("Start Line", "start_line_var"),
                ]
            show_chrome_row = True
            row_offset = 1

        for row_idx, (label_text, key_name) in enumerate(labels):
            grid_row = row_idx + row_offset
            tk.Label(target_grid, text=label_text, bg="#f7f7fa", anchor="w", width=12).grid(row=grid_row, column=0, sticky="w", pady=2)
            for view_col, (block_idx, block) in enumerate(render_blocks, start=1):
                if key_name == "col_air_date_var":
                    cell = tk.Frame(target_grid, bg="#f7f7fa")
                    cell.grid(row=grid_row, column=view_col, sticky="w", padx=2, pady=2)
                    ent = tk.Entry(cell, textvariable=block[key_name], width=11)
                    ent.pack(side="left")
                    btn = tk.Button(cell, text="...", width=3)
                    btn.configure(command=lambda x=block_idx, w=btn: self._pick_air_date_for_block(x, w))
                    btn.pack(side="left", padx=(2, 0))
                    self.mapping_entries.append(ent)
                    self.mapping_entries.append(btn)
                else:
                    ent = tk.Entry(target_grid, textvariable=block[key_name], width=11)
                    ent.grid(row=grid_row, column=view_col, sticky="w", padx=2, pady=2)
                    self.mapping_entries.append(ent)

        if show_chrome_row:
            chrome_row_idx = len(labels) + row_offset
            tk.Label(target_grid, text="Chrome", bg="#f7f7fa", anchor="w", width=12).grid(
                row=chrome_row_idx, column=0, sticky="w", pady=2
            )
            for view_col, (block_idx, _block) in enumerate(render_blocks, start=1):
                launch_btn = tk.Button(
                    target_grid,
                    text=f"Chrome {self._get_block_port(block_idx)}",
                    width=12,
                    bg="#d9edf7",
                    command=lambda x=block_idx: self.launch_chrome_for_block(x),
                )
                launch_btn.grid(row=chrome_row_idx, column=view_col, sticky="w", padx=2, pady=2)
                self.mapping_launch_buttons.append(launch_btn)

    def _build_menu(self):
        menubar = tk.Menu(self.root)

        menu_file = tk.Menu(menubar, tearoff=0)
        menu_file.add_command(label="Lưu cấu hình", command=self.save_settings)
        menu_file.add_command(label="Tải lại app", command=self.reload_app)
        menu_file.add_separator()
        menu_file.add_command(label="Thoát", command=self.exit_app)
        menubar.add_cascade(label="Tệp", menu=menu_file)

        menu_run = tk.Menu(menubar, tearoff=0)
        menu_run.add_command(label="Bắt đầu", command=self.start_processing)
        menu_run.add_command(label="Chạy lại các dòng lỗi", command=self.start_processing_error_rows)
        menu_run.add_command(label="Tạm dừng / Tiếp tục", command=self.toggle_pause)
        menu_run.add_command(label="Launch Chrome", command=self.launch_chrome_for_login)
        menubar.add_cascade(label="Chạy", menu=menu_run)

        menu_error = tk.Menu(menubar, tearoff=0)
        menu_error.add_checkbutton(
            label="Chỉ chạy các dòng lỗi đã lưu",
            variable=self.only_run_error_rows,
            onvalue=True,
            offvalue=False,
        )
        menu_error.add_command(label="Lưu lịch sử dòng lỗi", command=self._save_error_history_current_sheet)
        menu_error.add_command(label="Xuất bảng log ra Excel", command=self.export_live_log_excel)
        self.menu_error_saved_sheets = tk.Menu(menu_error, tearoff=0)
        menu_error.add_cascade(label="Danh sách sheet lỗi đã lưu", menu=self.menu_error_saved_sheets)
        menu_error.add_separator()
        menu_error.add_command(label="Xóa lịch sử dòng lỗi", command=self._clear_error_history)
        menubar.add_cascade(label="Lỗi", menu=menu_error)

        menu_help = tk.Menu(menubar, tearoff=0)
        menu_help.add_command(label="Mở log.txt", command=lambda: self._open_path(LOG_PATH))
        menu_help.add_command(label="Mở app_settings.json", command=lambda: self._open_path(SETTINGS_PATH))
        menu_help.add_command(label="Mở error_history.json", command=lambda: self._open_path(ERROR_HISTORY_PATH))
        menubar.add_cascade(label="Hỗ trợ", menu=menu_help)

        self.root.config(menu=menubar)
        self.menubar = menubar
        self.refresh_saved_sheets_list()

    def _open_path(self, path: str):
        try:
            if not os.path.exists(path):
                with open(path, "a", encoding="utf-8"):
                    pass
            os.startfile(path)  # type: ignore[attr-defined]
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi mở file", str(e))
            write_log(f"[WARN] Open path failed: {path} -> {e}")

    def _clear_error_history(self):
        try:
            save_error_history({})
            self.refresh_error_history_ui()
            if messagebox:
                messagebox.showinfo("Đã xóa", "Đã xóa lịch sử các dòng lỗi.")
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi", str(e))

    def _clear_current_sheet_error_history(self):
        try:
            sheet_url = self.sheet_url_var.get().strip()
            sheet_name = self.sheet_name_var.get().strip()
            if not sheet_url:
                if messagebox:
                    messagebox.showwarning("Thiếu Sheet URL", "Bạn chưa nhập Sheet URL.")
                return
            set_error_rows_for_sheet(sheet_url, sheet_name=sheet_name, rows=set(), details={})
            self.live_error_details = {}
            self.refresh_error_history_ui()
            if messagebox:
                messagebox.showinfo("Đã xóa", "Đã xóa lịch sử lỗi của sheet hiện tại.")
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi", str(e))

    def _save_error_history_current_sheet(self):
        try:
            sheet_url = self.sheet_url_var.get().strip()
            sheet_name = self.sheet_name_var.get().strip()
            if not sheet_url:
                if messagebox:
                    messagebox.showwarning("Thiếu Sheet URL", "Bạn chưa nhập Sheet URL.")
                return
            rows = get_error_rows_for_sheet(sheet_url)
            details = get_error_details_for_sheet(sheet_url)
            if getattr(self, "live_error_details", None):
                details.update({int(k): str(v) for k, v in self.live_error_details.items()})
                rows = set(rows) | set(details.keys())
            # If app history is empty, scan current sheet for ERR markers.
            if not rows:
                rows = self._collect_error_rows_from_sheet()
                details = self._collect_error_details_from_sheet(rows)
            # Force write/update timestamp for current sheet key.
            set_error_rows_for_sheet(sheet_url, sheet_name=sheet_name, rows=rows, details=details)
            self.refresh_error_history_ui()
            if messagebox:
                messagebox.showinfo("Đã lưu", f"Đã lưu lịch sử lỗi cho link Sheet hiện tại ({len(rows)} dòng).")
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi", str(e))

    def _collect_error_rows_from_sheet(self) -> set[int]:
        rows: set[int] = set()
        try:
            sheet_url = self.sheet_url_var.get().strip()
            sheet_name = self.sheet_name_var.get().strip()
            if not sheet_url or not sheet_name:
                return rows

            mapping_blocks = self.get_mapping_configs()
            scan_cols = []
            for block in mapping_blocks:
                idx_content = col_letter_to_index((block.get("col_content") or "").strip().upper())
                idx_drive = col_letter_to_index((block.get("col_drive") or "").strip().upper())
                for c in [idx_content, idx_drive]:
                    if c:
                        scan_cols.append(c)
            scan_cols = sorted(set(scan_cols))
            if not scan_cols:
                return rows

            try:
                starts = []
                for block in mapping_blocks:
                    try:
                        starts.append(int(str(block.get("start_line", "4")).strip() or "4"))
                    except Exception:
                        continue
                start_line = min(starts) if starts else 4
            except Exception:
                start_line = 4

            if not os.path.exists(JSON_PATH):
                write_log(f"[WARN] _collect_error_rows_from_sheet: credentials not found at {JSON_PATH}")
                return rows

            creds = ServiceAccountCredentials.from_json_keyfile_name(
                JSON_PATH,
                [
                    "https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive",
                ],
            )
            client = gspread.authorize(creds)
            worksheet = client.open_by_url(sheet_url).worksheet(sheet_name)

            for col_idx in scan_cols:
                vals = worksheet.col_values(col_idx)
                for r in range(start_line, len(vals) + 1):
                    v = str(vals[r - 1]).strip().upper()
                    if v.startswith("ERR"):
                        rows.add(r)
        except Exception as e:
            write_log(f"[WARN] _collect_error_rows_from_sheet failed: {e}")
        return rows

    def _collect_error_details_from_sheet(self, rows: set[int]) -> dict[int, str]:
        details: dict[int, str] = {}
        try:
            if not rows:
                return details
            sheet_url = self.sheet_url_var.get().strip()
            sheet_name = self.sheet_name_var.get().strip()
            if not sheet_url or not sheet_name:
                return details

            mapping_blocks = self.get_mapping_configs()
            scan_cols = []
            for block in mapping_blocks:
                idx_content = col_letter_to_index((block.get("col_content") or "").strip().upper())
                idx_drive = col_letter_to_index((block.get("col_drive") or "").strip().upper())
                for c in [idx_content, idx_drive]:
                    if c:
                        scan_cols.append(c)
            scan_cols = sorted(set(scan_cols))
            if not scan_cols:
                return details

            if not os.path.exists(JSON_PATH):
                return details

            creds = ServiceAccountCredentials.from_json_keyfile_name(
                JSON_PATH,
                [
                    "https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive",
                ],
            )
            client = gspread.authorize(creds)
            worksheet = client.open_by_url(sheet_url).worksheet(sheet_name)

            cols_data = {}
            for col_idx in scan_cols:
                try:
                    cols_data[col_idx] = worksheet.col_values(col_idx)
                except Exception:
                    cols_data[col_idx] = []

            for r in sorted(rows):
                msg = ""
                for col_idx in scan_cols:
                    vals = cols_data.get(col_idx, [])
                    if r - 1 < len(vals):
                        v = str(vals[r - 1]).strip()
                        if v.upper().startswith("ERR"):
                            msg = v
                            break
                if msg:
                    details[r] = msg
        except Exception as e:
            write_log(f"[WARN] _collect_error_details_from_sheet failed: {e}")
        return details

    def refresh_error_history_ui(self):
        try:
            sheet_url = self.sheet_url_var.get().strip()
            rows = sorted(get_error_rows_for_sheet(sheet_url))
            details = get_error_details_for_sheet(sheet_url)
            # Initialize live view from saved history when switching sheet/reloading.
            self.live_error_details = {int(r): str(details.get(r, "")).strip() for r in rows}
            self._render_error_history_card(self.live_error_details)
            self.refresh_saved_sheets_list()
        except Exception as e:
            write_log(f"[WARN] refresh_error_history_ui failed: {e}")

    def _render_error_history_card(self, details_map: dict[int, str]):
        rows = sorted(details_map.keys())
        if not rows:
            header_text = "⚠ Lỗi theo link Sheet: chưa có"
            list_text = "• Chưa có dòng lỗi đã lưu."
        else:
            header_text = f"⚠ Lỗi theo link Sheet: {len(rows)} dòng"
            lines = []
            for r in rows:
                msg = (details_map.get(r) or "Dòng lỗi đã lưu").strip()
                if ":" in msg:
                    left, right = msg.split(":", 1)
                    left = left.strip()
                    right = right.strip()
                    if left and right:
                        lines.append(f"• #{r}  [{left}] {right}")
                    elif left:
                        lines.append(f"• #{r}  [{left}]")
                    else:
                        lines.append(f"• #{r}  {right or 'Dòng lỗi đã lưu'}")
                else:
                    lines.append(f"• #{r}  {msg}")
            list_text = "\n".join(lines)
        if hasattr(self, "error_header_var"):
            self.error_header_var.set(header_text)
        if hasattr(self, "error_rows_var"):
            self.error_rows_var.set(list_text)
        if hasattr(self, "error_rows_text"):
            try:
                self.error_rows_text.config(state="normal")
                self.error_rows_text.delete("1.0", "end")
                self.error_rows_text.insert("1.0", list_text)
                self.error_rows_text.config(state="disabled")
                self.error_rows_text.yview_moveto(0.0)
            except Exception:
                pass

    def update_error_row_live(self, row: int, message: str = "", is_fail: bool = False):
        try:
            r = int(row)
            if r <= 0:
                return
            if is_fail:
                msg = (message or "Có lỗi trong quá trình xử lý").strip()
                self.live_error_details[r] = msg[:220]
            else:
                self.live_error_details.pop(r, None)
            self._render_error_history_card(self.live_error_details)
        except Exception as e:
            write_log(f"[WARN] update_error_row_live failed: {e}")

    def refresh_saved_sheets_list(self):
        try:
            items = list_saved_error_sheets()
            self._history_sheet_items = items
            menu_obj = getattr(self, "menu_error_saved_sheets", None)
            if menu_obj is None:
                return

            menu_obj.delete(0, "end")
            if not items:
                menu_obj.add_command(label="(Chưa có sheet nào)", state="disabled")
                return

            current_url = self.sheet_url_var.get().strip()
            for it in items:
                sname = (it.get("sheet_name") or "").strip()
                if sname:
                    label = f"{sname} | {it['sheet_url']} | lỗi:{it['rows_count']}"
                else:
                    label = f"{it['sheet_url']} | lỗi:{it['rows_count']}"
                if it.get("updated_at"):
                    label += f" | {it['updated_at']}"
                if it["sheet_url"] == current_url:
                    label = "• " + label
                menu_obj.add_command(
                    label=label,
                    command=lambda u=it["sheet_url"], n=(it.get("sheet_name") or ""): self.load_sheet_from_history(u, n),
                )
        except Exception as e:
            write_log(f"[WARN] refresh_saved_sheets_list failed: {e}")

    def load_sheet_from_history(self, sheet_url: str, sheet_name: str = ""):
        try:
            target_url = (sheet_url or "").strip()
            if not target_url:
                return
            self.sheet_url_var.set(target_url)
            if (sheet_name or "").strip():
                self.sheet_name_var.set((sheet_name or "").strip())
            self.refresh_error_history_ui()
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi", str(e))

    def start_processing_error_rows(self):
        sheet_url = self.sheet_url_var.get().strip()
        rows = sorted(get_error_rows_for_sheet(sheet_url))
        if not rows:
            if messagebox:
                messagebox.showinfo("Không có lỗi", "Link Sheet hiện tại chưa có lịch sử dòng lỗi để chạy lại.")
            return
        self.only_run_error_rows.set(True)
        self.start_processing()

    def _get_initial_geometry(self) -> str:
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        preferred_w = 1380
        preferred_h = 940
        width = min(preferred_w, max(1120, screen_w - 40))
        height = min(preferred_h, max(780, screen_h - 80))
        x = max(20, (screen_w - width) // 2)
        y = max(20, (screen_h - height) // 2)
        return f"{width}x{height}+{x}+{y}"

    def _on_canvas_configure(self, event):
        target_h = max(event.height, self.main_frame.winfo_reqheight())
        self.main_canvas.itemconfigure(self.canvas_window, width=event.width, height=target_h)
        if hasattr(self, "progress"):
            self.progress.configure(length=max(320, event.width - 80))
        if hasattr(self, "label_detail"):
            self.label_detail.configure(wraplength=max(320, event.width - 40))
        if hasattr(self, "error_rows_text"):
            try:
                self.error_rows_text.configure(width=max(40, (event.width - 90) // 7))
            except Exception:
                pass

    def _on_mousewheel(self, event):
        self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_scroll_events(self):
        self.main_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _get_settings_payload(self) -> dict:
        self._snapshot_current_mode_configs()
        mode_key = self._normalize_mode_name(self.mapping_mode_var.get())
        return {
            "sheet_url": self.sheet_url_var.get().strip(),
            "sheet_name": self.sheet_name_var.get().strip(),
            "drive_id": self.drive_id_var.get().strip(),
            "credentials_path": self.credentials_path_var.get().strip(),
            "target_rows_input": self.target_rows_var.get().strip(),
            "mapping_mode": mode_key,
            "mapping_blocks": self.mapping_blocks_by_mode.get(mode_key, self.get_mapping_configs()),
            "mapping_blocks_by_mode": self.mapping_blocks_by_mode,
            "capture_five_per_link": bool(self.capture_five_per_link.get()),
            "force_run_all": bool(self.force_run_all.get()),
            "only_run_error_rows": bool(self.only_run_error_rows.get()),
            "auto_launch_chrome": bool(self.auto_launch_chrome.get()),
        }

    def get_mapping_configs(self) -> list[dict]:
        out = []
        mode_name = (self.mapping_mode_var.get() or "Seeding").strip().lower()
        for i, block in enumerate(self.mapping_blocks):
            item = {
                "name": (block["name_var"].get() or f"Post {i + 1}").strip() or f"Post {i + 1}",
                "manual_link": (block["manual_link_var"].get() or "").strip(),
                "col_profile": (block["col_profile_var"].get() or "").strip().upper(),
                "col_content": (block["col_content_var"].get() or "").strip().upper(),
                "col_url": (block["col_url_var"].get() or "").strip().upper(),
                "col_drive": (block["col_drive_var"].get() or "").strip().upper(),
                "col_screenshot": (block["col_screenshot_var"].get() or "").strip().upper(),
                "col_air_date": (block["col_air_date_var"].get() or "").strip().upper(),
                "start_line": (block["start_line_var"].get() or "4").strip() or "4",
            }
            if mode_name == "seeding":
                item["col_profile"] = ""
                item["col_content"] = ""
            elif mode_name == "scan":
                item["col_profile"] = ""
                item["col_screenshot"] = ""
                item["col_air_date"] = ""
            out.append(item)
        return out

    def _load_mapping_blocks(self, blocks_data: list[dict] | None, render: bool = True):
        self.mapping_blocks = []
        for raw in blocks_data or []:
            if not isinstance(raw, dict):
                continue
            self.mapping_blocks.append(self._new_block_vars(raw))
        if not self.mapping_blocks:
            self._ensure_default_mapping_blocks()
        if render and hasattr(self, "mapping_grid"):
            self._render_mapping_blocks()

    def load_settings(self):
        try:
            if not os.path.exists(SETTINGS_PATH):
                return
            self._is_loading_settings = True
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            self.sheet_url_var.set(str(data.get("sheet_url", self.sheet_url_var.get())).strip())
            self.sheet_name_var.set(str(data.get("sheet_name", self.sheet_name_var.get())).strip())
            self.drive_id_var.set(str(data.get("drive_id", self.drive_id_var.get())).strip())
            self.target_rows_var.set(str(data.get("target_rows_input", self.target_rows_var.get())).strip())
            saved_credentials_path = str(data.get("credentials_path", self.credentials_path_var.get())).strip()
            if saved_credentials_path:
                # Keep compatibility with older configs where folder text might be mangled,
                # but filename remains valid.
                norm_cred = os.path.normpath(saved_credentials_path)
                if not os.path.exists(norm_cred):
                    candidate = os.path.join(APP_DIR, os.path.basename(norm_cred))
                    if os.path.exists(candidate):
                        norm_cred = candidate
                self.credentials_path_var.set(norm_cred)
            mode_value = self._normalize_mode_name(str(data.get("mapping_mode", self.mapping_mode_var.get())).strip() or "Seeding")
            blocks_data = data.get("mapping_blocks")
            if not isinstance(blocks_data, list):
                # Backward compatibility with old single-block config.
                blocks_data = [
                    {
                        "name": "Post 1",
                        "start_line": str(data.get("start_line", "4")).strip() or "4",
                        "col_url": str(data.get("col_url", "K")).strip().upper(),
                        "col_profile": str(data.get("col_profile", "B")).strip().upper(),
                        "col_content": str(data.get("col_content", "I")).strip().upper(),
                        "col_screenshot": str(data.get("col_screenshot", "J")).strip().upper(),
                        "col_drive": str(data.get("col_drive", "L")).strip().upper(),
                        "col_air_date": str(data.get("col_air_date", "")).strip().upper(),
                    }
                ]
            mode_map_raw = data.get("mapping_blocks_by_mode")
            mode_map: dict[str, list[dict]] = {}
            if isinstance(mode_map_raw, dict):
                for k, v in mode_map_raw.items():
                    mk = self._normalize_mode_name(k)
                    if isinstance(v, list):
                        mode_map[mk] = [x for x in v if isinstance(x, dict)]
            if mode_value not in mode_map or not mode_map.get(mode_value):
                mode_map[mode_value] = [x for x in blocks_data if isinstance(x, dict)]
            for mk in ("Seeding", "Booking", "Scan"):
                if not mode_map.get(mk):
                    mode_map[mk] = self._default_mapping_configs_for_mode(mk)
            self.mapping_blocks_by_mode = mode_map
            self.mapping_mode_var.set(mode_value)
            self._active_mapping_mode = mode_value
            self._load_mapping_blocks(self.mapping_blocks_by_mode.get(mode_value) or [], render=False)
            if hasattr(self, "mapping_grid"):
                self._render_mapping_blocks()
            self.force_run_all.set(bool(data.get("force_run_all", self.force_run_all.get())))
            self.only_run_error_rows.set(bool(data.get("only_run_error_rows", self.only_run_error_rows.get())))
            self.auto_launch_chrome.set(bool(data.get("auto_launch_chrome", self.auto_launch_chrome.get())))
            self.capture_five_per_link.set(bool(data.get("capture_five_per_link", self.capture_five_per_link.get())))
            if hasattr(self, "share_email_var"):
                email = get_service_account_email(self.credentials_path_var.get().strip()) or "link-verification@hazel-tea-485816-u3.iam.gserviceaccount.com"
                self.share_email_var.set(email)
            write_log(f"[INFO] Loaded settings from {SETTINGS_PATH}")
        except Exception as e:
            write_log(f"[WARN] Load settings failed: {e}")
        finally:
            self._is_loading_settings = False

    def save_settings(self, silent: bool = False):
        try:
            payload = self._get_settings_payload()
            with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            write_log(f"[INFO] Saved settings to {SETTINGS_PATH}")
            if (not silent) and messagebox:
                messagebox.showinfo("Đã lưu", f"Đã lưu cấu hình vào:\n{SETTINGS_PATH}")
        except Exception as e:
            write_log(f"[ERROR] Save settings failed: {e}")
            if (not silent) and messagebox:
                messagebox.showerror("Lỗi", f"Không lưu được cấu hình:\n{e}")

    def set_inputs_enabled(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.entry_sheet_url.config(state=state)
        self.entry_sheet_name.config(state=state)
        self.entry_drive_id.config(state=state)
        self.entry_credentials_path.config(state=state)
        for ent in getattr(self, "mapping_entries", []):
            try:
                ent.config(state=state)
            except Exception:
                pass
        for btn in getattr(self, "mapping_remove_buttons", []):
            try:
                btn.config(state=state)
            except Exception:
                pass
        # Keep per-post Chrome launch buttons enabled even while running,
        # so user can reopen login windows on demand.
        for btn in getattr(self, "mapping_launch_buttons", []):
            try:
                btn.config(state="normal")
            except Exception:
                pass
        if getattr(self, "btn_add_block", None):
            try:
                self.btn_add_block.config(state=state)
            except Exception:
                pass
        self.start_btn.config(state=state)
        if hasattr(self, "error_card_save_btn"):
            self.error_card_save_btn.config(state=state)
        if hasattr(self, "error_card_clear_btn"):
            self.error_card_clear_btn.config(state=state)
        self.export_log_btn.config(state=state)
        self.save_btn.config(state=state)
        self.reload_btn.config(state=state)
        self.checkbox.config(state=state)
        self.checkbox_errors_only.config(state=state)
        if hasattr(self, "entry_target_rows"):
            self.entry_target_rows.config(state=state)
        if hasattr(self, "btn_clear_target_rows"):
            self.btn_clear_target_rows.config(state=state)
        if hasattr(self, "chk_capture5"):
            try:
                if self.chk_capture5:
                    self.chk_capture5.config(state=state)
            except Exception:
                pass
        if hasattr(self, "mapping_mode_combo"):
            self.mapping_mode_combo.config(state="readonly" if enabled else "disabled")
        if hasattr(self, "btn_launch_chrome"):
            self.btn_launch_chrome.config(state=state)
        self.btn_paste_sheet_url.config(state=state)
        self.btn_paste_sheet_name.config(state=state)
        self.btn_paste_drive_id.config(state=state)
        if not enabled:
            self.pause_btn.config(state="normal")
        else:
            self.pause_btn.config(state="disabled")

    def toggle_pause(self):
        self.is_paused = not self.is_paused
        if self.is_paused:
            self.pause_btn.config(text="▶", bg="#c6e3b5")
            self.label_status.config(text="TẠM DỪNG", fg="#ff6b6b")
            # While paused, allow utility controls.
            self.reload_btn.config(state="normal")
            self.save_btn.config(state="normal")
            self.export_log_btn.config(state="normal")
            self.checkbox.config(state="normal")
            self.checkbox_errors_only.config(state="normal")
            if hasattr(self, "entry_target_rows"):
                self.entry_target_rows.config(state="normal")
            if hasattr(self, "btn_clear_target_rows"):
                self.btn_clear_target_rows.config(state="normal")
            if hasattr(self, "chk_capture5"):
                try:
                    if self.chk_capture5:
                        self.chk_capture5.config(state="normal")
                except Exception:
                    pass
            if hasattr(self, "mapping_mode_combo"):
                self.mapping_mode_combo.config(state="readonly")
            if hasattr(self, "btn_launch_chrome"):
                self.btn_launch_chrome.config(state="normal")
            self.btn_paste_sheet_url.config(state="normal")
            self.btn_paste_sheet_name.config(state="normal")
            self.btn_paste_drive_id.config(state="normal")
            for btn in getattr(self, "mapping_launch_buttons", []):
                try:
                    btn.config(state="normal")
                except Exception:
                    pass
            if getattr(self, "btn_add_block", None):
                try:
                    self.btn_add_block.config(state="normal")
                except Exception:
                    pass
            for btn in getattr(self, "mapping_remove_buttons", []):
                try:
                    btn.config(state="normal")
                except Exception:
                    pass
        else:
            self.pause_btn.config(text="⏸", bg="#fff3cd")
            self.label_status.config(text="ĐANG CHẠY", fg="#1877F2")
            # Resume running lock-state.
            self.reload_btn.config(state="disabled")
            self.save_btn.config(state="disabled")
            self.export_log_btn.config(state="disabled")
            self.checkbox.config(state="disabled")
            self.checkbox_errors_only.config(state="disabled")
            if hasattr(self, "entry_target_rows"):
                self.entry_target_rows.config(state="disabled")
            if hasattr(self, "btn_clear_target_rows"):
                self.btn_clear_target_rows.config(state="disabled")
            if hasattr(self, "chk_capture5"):
                try:
                    if self.chk_capture5:
                        self.chk_capture5.config(state="disabled")
                except Exception:
                    pass
            if hasattr(self, "mapping_mode_combo"):
                self.mapping_mode_combo.config(state="disabled")
            if hasattr(self, "btn_launch_chrome"):
                self.btn_launch_chrome.config(state="disabled")
            self.btn_paste_sheet_url.config(state="disabled")
            self.btn_paste_sheet_name.config(state="disabled")
            self.btn_paste_drive_id.config(state="disabled")
            if getattr(self, "btn_add_block", None):
                try:
                    self.btn_add_block.config(state="disabled")
                except Exception:
                    pass
            for btn in getattr(self, "mapping_remove_buttons", []):
                try:
                    btn.config(state="disabled")
                except Exception:
                    pass

    def paste_to(self, target_var):
        try:
            text = self.root.clipboard_get()
        except Exception:
            messagebox.showerror("Clipboard trống", "Không đọc được dữ liệu từ clipboard.")
            return
        target_var.set(text.strip())

    def copy_share_email(self):
        email = self.share_email_var.get().strip()
        if not email or email.startswith("Không đọc"):
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(email)
            self.root.update_idletasks()
            messagebox.showinfo("Đã copy", "Đã copy email vào clipboard. Chia sẻ Sheet và folder Drive với email này, chọn quyền Editor.")
        except Exception:
            messagebox.showerror("Lỗi", "Không copy được vào clipboard.")

    def export_live_log_excel(self):
        if filedialog is None:
            if messagebox:
                messagebox.showerror("Lỗi", "Môi trường hiện tại không hỗ trợ hộp thoại lưu file.")
            return
        if not hasattr(self, "live_log_table"):
            return

        rows = []
        for iid in self.live_log_table.get_children():
            vals = self.live_log_table.item(iid, "values")
            tags = self.live_log_table.item(iid, "tags") or ()
            if vals:
                rows.append((list(vals), list(tags)))
        if not rows:
            if messagebox:
                messagebox.showinfo("Không có dữ liệu", "Bảng log hiện đang trống.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Lưu bảng log",
            defaultextension=".xlsx",
            initialfile=f"evidence_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not out_path:
            return

        headers = ["Time", "#", "State", "Result", "Message"]
        try:
            if out_path.lower().endswith(".csv"):
                with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    w.writerow(headers)
                    for r, _tags in rows:
                        w.writerow(r)
            else:
                exported_with_color = False
                # 1) Prefer openpyxl
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Log"
                    ws.append(headers)
                    fill_ok = PatternFill(fill_type="solid", fgColor="FFD8F3DC")
                    fill_fail = PatternFill(fill_type="solid", fgColor="FFFFD9D9")
                    fill_unavailable = PatternFill(fill_type="solid", fgColor="FFFFE6C7")

                    for r, tags in rows:
                        ws.append(r)
                        row_idx = ws.max_row
                        tag_set = set(tags or [])
                        state = str(r[3]).strip().upper() if len(r) > 3 else ""
                        msg = str(r[4]).lower() if len(r) > 4 else ""
                        row_fill = None
                        if "fail" in tag_set or state == "FAIL":
                            row_fill = fill_fail
                        elif "unavailable" in tag_set or "nội dung không khả dụng" in msg:
                            row_fill = fill_unavailable
                        elif "ok" in tag_set or state == "OK":
                            row_fill = fill_ok
                        if row_fill:
                            for col in range(1, 6):
                                ws.cell(row=row_idx, column=col).fill = row_fill
                    wb.save(out_path)
                    exported_with_color = True
                except Exception:
                    exported_with_color = False

                # 2) Fallback to xlsxwriter (also keeps colors)
                if not exported_with_color:
                    try:
                        import xlsxwriter

                        wb = xlsxwriter.Workbook(out_path)
                        ws = wb.add_worksheet("Log")
                        fmt_ok = wb.add_format({"bg_color": "#D8F3DC"})
                        fmt_fail = wb.add_format({"bg_color": "#FFD9D9"})
                        fmt_unavailable = wb.add_format({"bg_color": "#FFE6C7"})
                        for c, h in enumerate(headers):
                            ws.write(0, c, h)
                        for i, (r, tags) in enumerate(rows, start=1):
                            tag_set = set(tags or [])
                            state = str(r[3]).strip().upper() if len(r) > 3 else ""
                            msg = str(r[4]).lower() if len(r) > 4 else ""
                            fmt = None
                            if "fail" in tag_set or state == "FAIL":
                                fmt = fmt_fail
                            elif "unavailable" in tag_set or "nội dung không khả dụng" in msg:
                                fmt = fmt_unavailable
                            elif "ok" in tag_set or state == "OK":
                                fmt = fmt_ok
                            for c, v in enumerate(r):
                                ws.write(i, c, v, fmt)
                        wb.close()
                        exported_with_color = True
                    except Exception:
                        exported_with_color = False

                # 3) Last fallback: HTML table saved as .xls (opens in Excel with colors)
                if not exported_with_color:
                    try:
                        native_xlsx = out_path if out_path.lower().endswith(".xlsx") else out_path + ".xlsx"
                        write_colored_xlsx_builtin(native_xlsx, headers, rows)
                        out_path = native_xlsx
                        exported_with_color = True
                    except Exception:
                        exported_with_color = False

                # 4) Last-resort fallback: HTML .xls with colors
                if not exported_with_color:
                    fallback = out_path
                    if not fallback.lower().endswith(".xls"):
                        fallback = fallback + ".xls"
                    css = """
                    <style>
                    table { border-collapse: collapse; font-family: Arial, sans-serif; font-size: 11pt; }
                    th, td { border: 1px solid #d0d0d0; padding: 4px 6px; }
                    .ok { background: #D8F3DC; }
                    .fail { background: #FFD9D9; }
                    .unavailable { background: #FFE6C7; }
                    </style>
                    """
                    lines = [
                        "<html><head><meta charset='utf-8'>",
                        css,
                        "</head><body><table>",
                        "<tr>" + "".join(f"<th>{html_lib.escape(h)}</th>" for h in headers) + "</tr>",
                    ]
                    for r, tags in rows:
                        tag_set = set(tags or [])
                        state = str(r[3]).strip().upper() if len(r) > 3 else ""
                        msg = str(r[4]).lower() if len(r) > 4 else ""
                        cls = ""
                        if "fail" in tag_set or state == "FAIL":
                            cls = "fail"
                        elif "unavailable" in tag_set or "nội dung không khả dụng" in msg:
                            cls = "unavailable"
                        elif "ok" in tag_set or state == "OK":
                            cls = "ok"
                        row_cells = "".join(f"<td>{html_lib.escape(str(v))}</td>" for v in r)
                        lines.append(f"<tr class='{cls}'>{row_cells}</tr>")
                    lines.append("</table></body></html>")
                    with open(fallback, "w", encoding="utf-8") as f:
                        f.write("\n".join(lines))
                    out_path = fallback
                    if messagebox:
                        messagebox.showwarning(
                            "Thiếu thư viện Excel",
                            "Máy này thiếu thư viện Excel, đã dùng fallback .xls (HTML) có màu.",
                        )
            if messagebox:
                messagebox.showinfo("Đã xuất", f"Đã xuất bảng log:\n{out_path}")
        except Exception as e:
            if messagebox:
                messagebox.showerror("Lỗi xuất file", str(e))

    def reset_live_log(self):
        try:
            if hasattr(self, "live_log_table"):
                for iid in self.live_log_table.get_children():
                    self.live_log_table.delete(iid)
        except Exception:
            pass
        self.update_progress_summary(0, 0, 0, 0, "---")

    def update_progress_summary(
        self,
        done: int,
        total: int,
        ok_count: int,
        fail_count: int,
        eta_text: str = "---",
        unavailable_count: int = 0,
    ):
        try:
            unavailable_text = f" | Unavailable: {unavailable_count}" if int(unavailable_count or 0) > 0 else ""
            self.progress_summary_var.set(
                f"✔ Progress: {done}/{total} | Success: {ok_count} | Failed: {fail_count}{unavailable_text} | ETA: {eta_text}"
            )
        except Exception:
            pass

    def add_live_log(self, row: int, state_left: str, state_right: str, message: str, tag: str = ""):
        try:
            ts = datetime.now().strftime("%H:%M:%S")
            if not hasattr(self, "live_log_table"):
                return
            self.live_log_table.insert(
                "",
                0,
                values=(ts, f"#{row}", state_left, state_right, message[:240]),
                tags=(tag,) if tag else (),
            )
        except Exception:
            pass

    def show_completion_popup(self, title: str, summary_text: str, severity: str = "info"):
        if tk is None:
            return
        win = tk.Toplevel(self.root)
        win.title(title)
        win.configure(bg="#f7f9ff")
        win.resizable(False, False)
        win.attributes("-topmost", True)
        win.transient(self.root)
        win.grab_set()

        border_color = "#2f80ed"
        icon = "ℹ"
        icon_color = "#2f80ed"
        if severity == "warn":
            border_color = "#e09f00"
            icon = "⚠"
            icon_color = "#e09f00"
        elif severity == "error":
            border_color = "#c0392b"
            icon = "✖"
            icon_color = "#c0392b"

        outer = tk.Frame(win, bg=border_color, padx=2, pady=2)
        outer.pack(fill="both", expand=True)
        body = tk.Frame(outer, bg="#ffffff", padx=16, pady=14)
        body.pack(fill="both", expand=True)

        top_row = tk.Frame(body, bg="#ffffff")
        top_row.pack(fill="x", pady=(0, 8))
        tk.Label(top_row, text=icon, font=("Arial", 20, "bold"), fg=icon_color, bg="#ffffff").pack(side="left")
        tk.Label(top_row, text=title, font=("Arial", 12, "bold"), fg="#1d2a44", bg="#ffffff").pack(side="left", padx=(8, 0))

        tk.Label(
            body,
            text=summary_text,
            justify="left",
            anchor="w",
            font=("Arial", 11, "bold"),
            fg="#1f2d4d",
            bg="#ffffff",
        ).pack(fill="x", pady=(0, 12))

        btn = tk.Button(body, text="OK", width=10, command=win.destroy, bg="#2f80ed", fg="#ffffff")
        btn.pack(anchor="e")
        btn.focus_set()

        try:
            win.update_idletasks()
            sw = win.winfo_screenwidth()
            sh = win.winfo_screenheight()
            ww = win.winfo_width()
            wh = win.winfo_height()
            x = max(0, (sw - ww) // 2)
            y = max(0, (sh - wh) // 2)
            win.geometry(f"+{x}+{y}")
        except Exception:
            pass

    def launch_chrome_for_login(self):
        browser_port = 9223
        
        ok, info = launch_chrome_for_login(browser_port)
        if ok:
            messagebox.showinfo(
                "Chrome đã mở",
                f"Chrome mở trên port {browser_port}.\n\nBây giờ bạn có thể:\n"
                f"1. Đăng nhập Facebook, TikTok, Instagram, YouTube\n"
                f"2. Sau đó bấm BẮT ĐẦU để chạy xử lý\n"
                f"3. Chrome sẽ nhớ tất cả đăng nhập"
            )
        else:
            messagebox.showerror("Lỗi", f"Không mở được Chrome: {info}")

    def start_processing(self):
        global JSON_PATH
        run_mode = (self.mapping_mode_var.get() or "Seeding").strip().lower()
        is_scan_image_mode = run_mode == "scan"
        is_scan_like_mode = is_scan_image_mode
        sheet_url = normalize_sheet_input(self.sheet_url_var.get().strip())
        sheet_name = self.sheet_name_var.get().strip()
        drive_id = normalize_drive_folder_input(self.drive_id_var.get().strip())
        cred_input = self.credentials_path_var.get().strip()
        if sheet_url:
            self.sheet_url_var.set(sheet_url)
        if drive_id:
            self.drive_id_var.set(drive_id)

        if not sheet_url or not sheet_name or ((not is_scan_like_mode) and not drive_id):
            required_text = (
                "Vui lòng nhập đầy đủ Sheet URL, Sheet Name."
                if is_scan_like_mode
                else "Vui lòng nhập đầy đủ Sheet URL, Sheet Name và Drive Folder ID."
            )
            messagebox.showerror(
                "Thiếu thông tin",
                required_text
            )
            return
        if not cred_input:
            messagebox.showerror(
                "Thiếu credentials",
                "Hãy nhập đường dẫn credentials.json hoặc dán nội dung JSON vào ô Credentials JSON."
            )
            return

        if os.path.exists(cred_input):
            JSON_PATH = cred_input
        else:
            # Allow pasting raw JSON directly in Credentials input.
            try:
                data = json.loads(cred_input)
            except Exception:
                messagebox.showerror(
                    "Credentials không hợp lệ",
                    "Ô Credentials JSON không phải đường dẫn file và cũng không phải JSON hợp lệ.\n"
                    "Hãy dán trực tiếp nội dung JSON vào ô Credentials JSON."
                )
                return

            required = ["type", "client_email", "private_key"]
            missing = [k for k in required if not str(data.get(k, "")).strip()]
            if missing:
                messagebox.showerror("Thiếu trường", f"JSON thiếu trường bắt buộc: {', '.join(missing)}")
                return

            out_path = os.path.join(BASE_DIR, "credentials.inline.json")
            try:
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception as e:
                messagebox.showerror("Lỗi lưu file", str(e))
                return
            JSON_PATH = out_path
            self.credentials_path_var.set(out_path)
            self.share_email_var.set(str(data.get("client_email", "")).strip() or "link-verification@hazel-tea-485816-u3.iam.gserviceaccount.com")

        block_configs = self.get_mapping_configs()
        target_rows_text = (self.target_rows_var.get() or "").strip()
        target_rows: list[int] = []
        if target_rows_text:
            target_rows, target_rows_err = parse_target_rows_text(target_rows_text)
            if target_rows_err:
                messagebox.showerror("Rows to rerun không hợp lệ", target_rows_err)
                return
            if not target_rows:
                messagebox.showerror("Rows to rerun trống", "Bạn đã nhập lọc hàng nhưng không parse được hàng hợp lệ.")
                return
            if self.only_run_error_rows.get():
                self.only_run_error_rows.set(False)
                write_log("[INFO] Manual row filter is set -> tắt 'Retry Failed Only' để chỉ chạy đúng danh sách hàng đã nhập.")

        mappings = []
        for i, block in enumerate(block_configs):
            col_url = (block.get("col_url") or "").strip().upper()
            if not col_url:
                continue
            if col_url and (not col_letter_to_index(col_url)):
                messagebox.showerror("Lỗi dữ liệu", f"Block {i+1}: Cột Link URL không hợp lệ ({col_url}).")
                return
            col_profile = (block.get("col_profile") or "").strip().upper()
            col_content = (block.get("col_content") or "").strip().upper()
            col_screenshot = (block.get("col_screenshot") or "").strip().upper()
            col_drive = (block.get("col_drive") or "").strip().upper()
            col_air_date_raw = (block.get("col_air_date") or "").strip()
            col_air_date = col_air_date_raw.upper()
            optional_cols = [
                ("Profile", col_profile),
                ("Content", col_content),
                ("Screenshot", col_screenshot),
                ("Drive URL", col_drive),
            ]
            for label, col_ref in optional_cols:
                if col_ref and not col_letter_to_index(col_ref):
                    messagebox.showerror("Lỗi dữ liệu", f"Block {i+1}: Cột {label} không hợp lệ ({col_ref}).")
                    return
            if is_scan_like_mode:
                if not col_content:
                    messagebox.showerror("Lỗi dữ liệu", f"Block {i+1}: Text Column không được để trống.")
                    return
                if not col_drive:
                    messagebox.showerror("Lỗi dữ liệu", f"Block {i+1}: Result Column không được để trống.")
                    return
            fixed_air_date = ""
            if col_air_date_raw:
                if col_letter_to_index(col_air_date):
                    pass
                else:
                    fixed_air_date = get_air_date_token(col_air_date_raw)
                    if not fixed_air_date:
                        messagebox.showerror(
                            "Lỗi dữ liệu",
                            f"Block {i+1}: Air Date phải là ký tự cột (vd: H) hoặc ngày hợp lệ (vd: 2026-03-10).",
                        )
                        return
                    col_air_date = ""
            try:
                start_line = int(str(block.get("start_line", "4")).strip() or "4")
            except ValueError:
                messagebox.showerror("Lỗi dữ liệu", f"Block {i+1}: Start Line phải là số.")
                return
            mappings.append(
                {
                    "name": (
                        (
                            (block.get("name") or "").strip()
                            or (f"Scan {i+1}")
                        )
                        if is_scan_like_mode
                        else (block.get("name") or f"Post {i+1}").strip() or f"Post {i+1}"
                    ),
                    "start_line": start_line,
                    "col_url": col_url,
                    "col_profile": col_profile,
                    "col_content": col_content,
                    "col_screenshot": col_screenshot,
                    "col_drive": col_drive,
                    "col_air_date": col_air_date,
                    "fixed_air_date": fixed_air_date,
                    "mode": run_mode,
                }
            )

        if not mappings:
            messagebox.showerror("Thiếu cấu hình", "Cần ít nhất 1 block hợp lệ để chạy.")
            return
        if is_scan_image_mode:
            ocr_ok, ocr_msg = check_ocr_dependencies()
            if not ocr_ok:
                messagebox.showerror(
                    "Thiếu OCR",
                    f"Scan cần OCR để đọc chữ trong ảnh.\n\n{ocr_msg}",
                )
                return
        browser_port = 9223
        if self.auto_launch_chrome.get() and run_mode != "scan":
            # Auto-prepare Chrome for all mapped posts, not only Post 1.
            for i in range(len(mappings)):
                p = self._get_block_port(i)
                prof = self._get_block_profile(i)
                ok, info = launch_chrome_for_login(p, profile_path=prof)
                if not ok:
                    write_log(f"[WARN] Auto launch Chrome failed (Post {i+1}, port {p}): {info}")

        self.is_running = True
        self.is_paused = False
        self.save_settings(silent=True)
        self.set_inputs_enabled(False)
        self.label_status.config(text="ĐANG CHẠY", fg="#1877F2")
        self.label_detail.config(text="Bắt đầu xử lý...")
        self.reset_live_log()

        threading.Thread(
            target=lambda: main_logic(
                self,
                drive_id,
                sheet_url,
                sheet_name,
                mappings=mappings,
                browser_port=browser_port,
                target_rows=target_rows,
            ),
            daemon=True
        ).start()

    def reload_app(self):
        self.is_running = False
        self.is_paused = False
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.driver = None
        self.load_settings()
        self.progress["value"] = 0
        self.pause_btn.config(text="⏸", bg="#fff3cd")
        self.label_status.config(text="Sẵn sàng", fg="#1877F2")
        self.label_detail.config(text="Đã load lại app.")
        self.reset_live_log()
        self.set_inputs_enabled(True)

    def exit_app(self):
        self.is_running = False
        self.main_canvas.unbind_all("<MouseWheel>")
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.destroy()
        os._exit(0)

# ================= CORE =================
def main_logic(app: ProgressApp, drive_id: str, sheet_url: str, sheet_name: str, start_line: int = 4, col_url_letter: str = "K", col_profile_letter: str = "B", col_content_letter: str = "I", col_screenshot_letter: str = "J", col_drive_letter: str = "L", browser_port: int = 9223, mappings: list[dict] | None = None, primary_profile_path: str | None = None, target_rows: list[int] | set[int] | tuple[int, ...] | None = None, target_block_name: str | None = None):
    def ui_call(fn, *args, **kwargs):
        """
        Run UI actions on Tk main thread to avoid random crashes on some machines.
        """
        try:
            root = getattr(app, "root", None)
            if root is not None and hasattr(root, "after") and threading.current_thread() is not threading.main_thread():
                root.after(0, lambda: fn(*args, **kwargs))
            else:
                fn(*args, **kwargs)
        except Exception as e:
            write_log(f"[WARN] UI call failed: {e}")

    def can_show_native_dialog() -> bool:
        """
        Only allow Tk native dialogs when running real desktop GUI mode.
        Web/headless workers must never call messagebox to avoid Tcl/Tk crashes.
        """
        if messagebox is None:
            return False
        root = getattr(app, "root", None)
        return bool(root is not None and hasattr(root, "tk"))

    def ui_set_progress(value: int):
        app.progress["value"] = value

    def ui_set_detail(text: str):
        app.label_detail.config(text=text)

    def ui_set_status(text: str, fg: str = ""):
        app.label_status.config(text=text, fg=fg)

    def ui_set_done():
        app.label_status.config(text="HOÀN TẤT", fg="#34C759")

    def ui_update_summary(done: int, total: int, ok_count: int, fail_count: int, eta_text: str, unavailable_count: int = 0):
        if hasattr(app, "update_progress_summary"):
            app.update_progress_summary(done, total, ok_count, fail_count, eta_text, unavailable_count)

    def ui_add_log(row: int, state_left: str, state_right: str, message: str, tag: str):
        if hasattr(app, "add_live_log"):
            app.add_live_log(row, state_left, state_right, message, tag)

    tracked_error_rows: set[int] = set()
    tracked_error_details: dict[int, str] = {}
    history_ready = False
    requested_data_start_row = max(1, int(start_line or 4))
    try:
        for item in list(mappings or []):
            if isinstance(item, dict):
                requested_data_start_row = min(
                    requested_data_start_row,
                    max(1, int(str(item.get("start_line", requested_data_start_row)).strip() or requested_data_start_row)),
                )
    except Exception:
        pass
    try:
        if target_rows:
            requested_data_start_row = min(
                requested_data_start_row,
                max(1, int(min(int(r) for r in list(target_rows) if int(r) > 0))),
            )
    except Exception:
        pass
    try:
        highlight_sheet_errors_enabled = bool(getattr(app, "highlight_sheet_errors", None).get())
    except Exception:
        highlight_sheet_errors_enabled = False

    try:
        write_log("=== START FINAL TOOL v2.2 ===")

        if not os.path.exists(TEMP_DIR):
            os.makedirs(TEMP_DIR)

        if not os.path.exists(JSON_PATH):
            raise FileNotFoundError(
                "Khong tim thay credentials.json.\n"
                f"Duong dan dang tim: {JSON_PATH}\n"
                "Hay dat credentials.json canh file .exe hoac set env GOOGLE_CREDENTIALS_PATH."
            )

        creds = ServiceAccountCredentials.from_json_keyfile_name(
            JSON_PATH,
            [
                "https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"
            ]
        )
        client = gspread.authorize(creds)
        drive_service = build("drive", "v3", credentials=creds)
        sheets_service = build("sheets", "v4", credentials=creds)

        spreadsheet = client.open_by_url(sheet_url)
        spreadsheet_id = spreadsheet.id
        worksheet = spreadsheet.worksheet(sheet_name)
        sheet_id = worksheet.id
        data_start_row = requested_data_start_row
        sheet_total_rows = max(0, int(getattr(worksheet, "row_count", 0) or 0) - data_start_row + 1)
        _column_cache: dict[tuple[str, tuple[int, ...], int], dict[int, list[str]]] = {}

        def _pad_column_values(values: list[str], total_rows: int) -> list[str]:
            out = list(values or [])
            if total_rows > 0 and len(out) < total_rows:
                out.extend([""] * (total_rows - len(out)))
            return out

        def _flatten_range_rows(rows: list[Any], total_rows: int) -> list[str]:
            out: list[str] = []
            for r in list(rows or []):
                if isinstance(r, list) and r:
                    out.append(str(r[0]).strip())
                else:
                    out.append("")
            return _pad_column_values(out, total_rows)

        def _batch_fetch_columns(
            col_indices: list[int] | tuple[int, ...] | set[int],
            *,
            value_render_option: str = "UNFORMATTED_VALUE",
            start_row: int = data_start_row,
            total_rows: int = sheet_total_rows,
        ) -> dict[int, list[str]]:
            unique_cols = tuple(sorted({int(col) for col in list(col_indices or []) if int(col) > 0}))
            if not unique_cols:
                return {}
            cache_key = (str(value_render_option or "UNFORMATTED_VALUE").upper(), unique_cols, int(start_row))
            cached = _column_cache.get(cache_key)
            if cached is not None:
                return cached
            ranges_by_col: dict[int, str] = {}
            end_row = max(int(start_row), int(start_row + max(total_rows, 0) - 1))
            for col_idx in unique_cols:
                col_letter = col_index_to_letter(col_idx)
                if not col_letter:
                    continue
                ranges_by_col[col_idx] = f"{col_letter}{start_row}:{col_letter}{end_row}"
            if not ranges_by_col:
                _column_cache[cache_key] = {}
                return {}
            fetched: dict[int, list[str]] = {}

            def _get_aligned_rows(rng: str) -> list[Any]:
                try:
                    # Keep exact row alignment even when there are blanks/merged cells.
                    return worksheet.get(
                        rng,
                        value_render_option=value_render_option,
                        maintain_size=True,
                        pad_values=True,
                    ) or []
                except TypeError:
                    try:
                        return worksheet.get(
                            rng,
                            value_render_option=value_render_option,
                            pad_values=True,
                        ) or []
                    except Exception:
                        return worksheet.get(rng, value_render_option=value_render_option) or []

            for col_idx, rng in ranges_by_col.items():
                try:
                    rows = _get_aligned_rows(rng)
                except Exception as single_exc:
                    write_log(f"[WARN] get range failed {rng} ({value_render_option}): {single_exc}")
                    rows = []
                fetched[col_idx] = _flatten_range_rows(rows, total_rows)
            for col_idx in unique_cols:
                fetched.setdefault(col_idx, [""] * max(total_rows, 0))
            _column_cache[cache_key] = fetched
            return fetched

        def wait_page_ready(driver, timeout: int = PAGE_READY_TIMEOUT):
            try:
                WebDriverWait(driver, timeout).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
            except Exception:
                time.sleep(PAGE_READY_FALLBACK_SLEEP)

        def load_existing_drive_files():
            files_by_name = {}
            page_token = None
            while True:
                try:
                    resp = drive_service.files().list(
                        q=f"'{drive_id}' in parents and trashed = false",
                        fields="nextPageToken, files(id,name)",
                        supportsAllDrives=True,
                        includeItemsFromAllDrives=True,
                        corpora="allDrives",
                        pageSize=1000,
                        pageToken=page_token,
                    ).execute()
                except Exception as drive_exc:
                    msg = str(drive_exc)
                    msg_lower = msg.lower()
                    if ("file not found" in msg_lower) or ("404" in msg_lower) or ("notfound" in msg_lower):
                        raise RuntimeError(
                            "Drive Folder ID không tồn tại hoặc service account chưa có quyền truy cập.\n"
                            f"Drive ID hiện tại: {drive_id}"
                        ) from drive_exc
                    raise
                for f in resp.get("files", []):
                    n = f.get("name")
                    fid = f.get("id")
                    if n and fid:
                        files_by_name.setdefault(n, []).append(fid)
                page_token = resp.get("nextPageToken")
                if not page_token:
                    break
            return files_by_name

        def build_chrome_options(user_data_dir: str, headless: bool, debug_port: int) -> Options:
            options = Options()
            normalized_profile = _resolve_writable_profile_dir(
                user_data_dir or LOCAL_PROFILE_PATH,
                browser_port=debug_port,
                log_prefix="WebDriver: ",
            )
            if normalized_profile:
                options.add_argument(f"--user-data-dir={normalized_profile}")
            options.add_argument(f"--window-size={CAPTURE_WINDOW_SIZE}")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(f"--remote-debugging-port={debug_port}")
            options.add_argument("--force-device-scale-factor=1")
            options.add_argument("--high-dpi-support=1")
            options.add_argument("--remote-allow-origins=*")
            options.add_argument("--no-first-run")
            options.add_argument("--no-default-browser-check")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-background-networking")
            options.add_argument("--disable-sync")
            options.add_argument("--disable-features=TranslateUI")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.page_load_strategy = "eager"
            if headless:
                headless_mode = os.environ.get("EVIDENCE_CHROME_HEADLESS_MODE", "old").strip().lower()
                options.add_argument("--headless=new" if headless_mode == "new" else "--headless")
            return options

        scan_only_request = bool(mappings) and all(
            str((m or {}).get("mode", "seeding")).strip().lower() == "scan"
            for m in mappings
        )

        service = None
        def seed_profile_if_needed(target_profile: str):
            if not target_profile:
                return
            try:
                normalized_target = _resolve_writable_profile_dir(
                    target_profile,
                    browser_port=browser_port,
                    log_prefix="Seed: ",
                )
                os.makedirs(normalized_target, exist_ok=True)
                if os.path.isdir(os.path.join(normalized_target, "Default")):
                    return
                seed_profile = ""
                normalized_local_profile = _normalize_profile_dir(LOCAL_PROFILE_PATH)
                if os.path.abspath(normalized_target) != os.path.abspath(normalized_local_profile) and os.path.isdir(normalized_local_profile):
                    seed_profile = normalized_local_profile
                if not seed_profile and os.path.isdir(FB_PROFILE_PATH):
                    seed_profile = FB_PROFILE_PATH
                if not seed_profile and os.path.isdir(FB_PROFILE_PATH_ALT):
                    seed_profile = FB_PROFILE_PATH_ALT
                if not seed_profile or os.path.abspath(seed_profile) == os.path.abspath(normalized_target):
                    return
                shutil.copytree(
                    seed_profile,
                    normalized_target,
                    dirs_exist_ok=True,
                    ignore=shutil.ignore_patterns(
                        "Cache",
                        "Code Cache",
                        "GPUCache",
                        "GrShaderCache",
                        "ShaderCache",
                        "Crashpad",
                        "Singleton*",
                        "lockfile",
                        "*.tmp",
                    ),
                )
                write_log(f"[INFO] Seeded profile '{normalized_target}' from '{seed_profile}'")
            except Exception as e:
                write_log(f"[WARN] Profile seed failed ({target_profile}): {e}")

        profile_candidates = []
        for cand in [
            primary_profile_path,
            LOCAL_PROFILE_PATH,
            FB_PROFILE_PATH,
            FB_PROFILE_PATH_ALT,
            os.path.join(TEMP_DIR, "chrome_profile_temp"),
        ]:
            cand = str(cand or "").strip()
            if cand and cand not in profile_candidates:
                profile_candidates.append(cand)
        browser_resolve_issues: list[str] = []
        def find_chrome_binary() -> str | None:
            nonlocal browser_resolve_issues
            chrome_path, issues = find_compatible_browser_binary()
            browser_resolve_issues = list(issues or [])
            for item in browser_resolve_issues:
                write_log(f"[WARN] Skip browser binary: {item}")
            return chrome_path

        if not scan_only_request:
            attach_only_existing_browser = bool(getattr(app, "attach_only_existing_browser", False))
            try:
                service = resolve_chromedriver_service()
            except Exception as e:
                service = None
                write_log(f"[WARN] Resolve chromedriver service failed, will fallback to Selenium Manager: {e}")
            started = False
            last_err = None

            # 1) Prefer attaching to an already launched Chrome (from "Launch Chrome").
            try:
                attach_opts = Options()
                attach_opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{browser_port}")
                app.driver = create_chrome_driver(attach_opts, service=service)
                write_log(f"[INFO] Attached to existing Chrome debug session on port {browser_port}")
                started = True
            except Exception as e:
                last_err = e
                write_log(f"[INFO] No attachable Chrome on port {browser_port}: {e}")
                if attach_only_existing_browser:
                    raise Exception(
                        f"ATTACH_ONLY_MODE: Không attach được Chrome đang mở ở port {browser_port}. "
                        "Hãy mở Chrome với --remote-debugging-port và đăng nhập trước."
                    )

            # 2) If not attachable, start new Chrome with profile candidates.
            if not started:
                for profile in profile_candidates:
                    seed_profile_if_needed(profile)
                    for headless in (False, True):
                        try:
                            app.driver = create_chrome_driver(
                                build_chrome_options(user_data_dir=profile, headless=headless, debug_port=browser_port),
                                service=service,
                            )
                            write_log(f"[INFO] Chrome started with profile='{profile}', headless={headless}")
                            started = True
                            break
                        except Exception as e:
                            last_err = e
                            write_log(f"[WARN] Chrome start failed (profile='{profile}', headless={headless}): {e}")
                            try:
                                if app.driver:
                                    app.driver.quit()
                            except:
                                pass
                    if started:
                        break
            if not started:
                chrome_path = find_chrome_binary()
                if chrome_path:
                    debug_profile = primary_profile_path or LOCAL_PROFILE_PATH
                    seed_profile_if_needed(debug_profile)
                    for port in [browser_port]:
                        try:
                            args = [
                                chrome_path,
                                f"--remote-debugging-port={port}",
                                f"--user-data-dir={debug_profile}",
                                f"--window-size={CAPTURE_WINDOW_SIZE}",
                                "--force-device-scale-factor=1",
                                "--high-dpi-support=1",
                                "--no-first-run",
                                "--no-default-browser-check",
                                "--disable-extensions",
                                "--disable-background-networking",
                                "--disable-sync",
                                "--disable-gpu",
                            ]
                            subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                            time.sleep(2)
                            attach_opts = Options()
                            attach_opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
                            app.driver = webdriver.Chrome(service=service, options=attach_opts)
                            write_log(f"[INFO] Attached to debug Chrome on port {port}")
                            started = True
                            break
                        except Exception as e:
                            last_err = e
                            write_log(f"[WARN] Attach debug Chrome failed (port={port}): {e}")
                if not started:
                    if browser_resolve_issues:
                        write_log(
                            "[WARN] Browser compatibility issues detected while resolving binary: "
                            + " | ".join(browser_resolve_issues[:4])
                        )
                        if not chrome_path:
                            raise Exception(
                                "CHROME_START_FAILED: Không tìm được browser tương thích với macOS hiện tại. "
                                f"Chi tiết: {browser_resolve_issues[0]}"
                            )
                    raise Exception(f"CHROME_START_FAILED: {last_err}")
        else:
            write_log("[INFO] Scan mode: skip Selenium/Chrome startup.")

        requested_rows: set[int] = set()
        for item in target_rows or []:
            try:
                row_value = int(item)
            except Exception:
                continue
            if row_value >= 1:
                requested_rows.add(row_value)
        requested_block_name = str(target_block_name or "").strip().lower()

        mapping_list = mappings or [
            {
                "name": "Post 1",
                "start_line": start_line,
                "col_url": col_url_letter,
                "col_profile": col_profile_letter,
                "col_content": col_content_letter,
                "col_screenshot": col_screenshot_letter,
                "col_drive": col_drive_letter,
                "col_air_date": "",
                "mode": "seeding",
            }
        ]
        normalized_mappings = []
        for i, m in enumerate(mapping_list):
            m_name = str((m or {}).get("name", f"Post {i+1}")).strip() or f"Post {i+1}"
            m_mode = str((m or {}).get("mode", "seeding")).strip().lower() or "seeding"
            if m_mode == "scan only text":
                write_log(f"[INFO] '{m_name}': mode 'Scan Only Text' đã bị gỡ, tự chuyển sang 'Scan'.")
                m_mode = "scan"
            m_col_url = str((m or {}).get("col_url", "")).strip().upper()
            if not col_letter_to_index(m_col_url):
                write_log(f"[WARN] Skip {m_name}: invalid Link URL column '{m_col_url}'")
                continue
            m_col_profile = str((m or {}).get("col_profile", "")).strip().upper()
            m_col_content = str((m or {}).get("col_content", "")).strip().upper()
            m_col_screenshot = str((m or {}).get("col_screenshot", "")).strip().upper()
            m_col_drive = str((m or {}).get("col_drive", "")).strip().upper()
            m_col_air_date_raw = str((m or {}).get("col_air_date", "")).strip()
            m_col_air_date = m_col_air_date_raw.upper()
            m_fixed_air_date = str((m or {}).get("fixed_air_date", "")).strip()
            optional_pairs = [
                ("Profile", m_col_profile),
                ("Content", m_col_content),
                ("Screenshot", m_col_screenshot),
                ("Drive", m_col_drive),
            ]
            sanitized_optional = {}
            for label, col_ref in optional_pairs:
                if col_ref and not col_letter_to_index(col_ref):
                    write_log(f"[WARN] {m_name}: invalid {label} column '{col_ref}', set empty.")
                    sanitized_optional[label] = ""
                else:
                    sanitized_optional[label] = col_ref
            if m_mode == "seeding":
                sanitized_optional["Profile"] = ""
                sanitized_optional["Content"] = ""
            elif m_mode == "scan":
                sanitized_optional["Profile"] = ""
                sanitized_optional["Screenshot"] = ""
                m_col_air_date = ""
                m_fixed_air_date = ""
            if m_col_air_date and not col_letter_to_index(m_col_air_date):
                parsed_fixed = get_air_date_token(m_col_air_date_raw)
                if parsed_fixed:
                    m_fixed_air_date = parsed_fixed
                    m_col_air_date = ""
                else:
                    write_log(f"[WARN] {m_name}: invalid AirDate '{m_col_air_date_raw}', ignore.")
                    m_col_air_date = ""
            try:
                m_start = int(str((m or {}).get("start_line", "4")).strip() or "4")
            except Exception:
                m_start = 4
            normalized_mappings.append(
                {
                    "block_index": i,
                    "name": m_name,
                    "start_line": m_start,
                    "col_url": m_col_url,
                    "col_profile": sanitized_optional["Profile"],
                    "col_content": sanitized_optional["Content"],
                    "col_screenshot": sanitized_optional["Screenshot"],
                    "col_drive": sanitized_optional["Drive"],
                    "col_air_date": m_col_air_date,
                    "fixed_air_date": m_fixed_air_date,
                    "mode": m_mode,
                }
            )
        if requested_block_name:
            filtered_mappings = [m for m in normalized_mappings if str(m.get("name", "")).strip().lower() == requested_block_name]
            if filtered_mappings:
                normalized_mappings = filtered_mappings
        if not normalized_mappings:
            raise Exception("KHONG_CO_BLOCK_HOP_LE")
        if any(str(m.get("mode", "")).strip().lower() == "scan" for m in normalized_mappings):
            ocr_ok, ocr_msg = check_ocr_dependencies()
            if not ocr_ok:
                ui_call(ui_set_status, "THIẾU OCR", "#ef4444")
                ui_call(ui_set_detail, ocr_msg)
                raise Exception(f"OCR_UNAVAILABLE: {ocr_msg}")

        prepared_blocks = []
        need_upload = False
        max_rows = 0
        scan_url_cols: set[int] = set()
        url_formula_cols: set[int] = set()
        unformatted_cols: set[int] = set()
        for m in normalized_mappings:
            idx_url = col_letter_to_index(m["col_url"])
            idx_profile = col_letter_to_index(m["col_profile"]) if m["col_profile"] else None
            idx_content = col_letter_to_index(m["col_content"]) if m["col_content"] else None
            idx_drive = col_letter_to_index(m["col_drive"]) if m["col_drive"] else None
            idx_screenshot = col_letter_to_index(m["col_screenshot"]) if m["col_screenshot"] else None
            idx_air_date = col_letter_to_index(m["col_air_date"]) if m["col_air_date"] else None
            mode_name = str(m.get("mode", "seeding")).strip().lower()
            if idx_url:
                unformatted_cols.add(idx_url)
            if idx_profile:
                unformatted_cols.add(idx_profile)
            if idx_content:
                unformatted_cols.add(idx_content)
            if idx_drive:
                unformatted_cols.add(idx_drive)
            if idx_screenshot:
                unformatted_cols.add(idx_screenshot)
            if idx_air_date:
                unformatted_cols.add(idx_air_date)
            if mode_name == "scan" and idx_url:
                scan_url_cols.add(idx_url)
            if idx_url:
                url_formula_cols.add(idx_url)

        unformatted_column_cache = _batch_fetch_columns(unformatted_cols, value_render_option="UNFORMATTED_VALUE")
        formula_column_cache = _batch_fetch_columns(scan_url_cols, value_render_option="FORMULA") if scan_url_cols else {}
        formula_url_column_cache = _batch_fetch_columns(url_formula_cols, value_render_option="FORMULA") if url_formula_cols else {}

        def _extract_http_url(raw_value: str) -> str:
            raw = str(raw_value or "").strip()
            if not raw:
                return ""
            # Fast path
            if raw.startswith("http://") or raw.startswith("https://"):
                return raw
            # Formula form: =HYPERLINK("https://...", "label")
            m = re.search(r'HYPERLINK\(\s*"([^"]+)"', raw, flags=re.IGNORECASE)
            if m:
                cand = str(m.group(1) or "").strip()
                if cand.startswith("http://") or cand.startswith("https://"):
                    return cand
            # Generic fallback: first URL-like token
            m2 = re.search(r'https?://[^\s")]+', raw, flags=re.IGNORECASE)
            if m2:
                return str(m2.group(0) or "").strip()
            return ""
        for m in normalized_mappings:
            idx_url = col_letter_to_index(m["col_url"])
            idx_profile = col_letter_to_index(m["col_profile"]) if m["col_profile"] else None
            idx_content = col_letter_to_index(m["col_content"]) if m["col_content"] else None
            idx_drive = col_letter_to_index(m["col_drive"]) if m["col_drive"] else None
            idx_screenshot = col_letter_to_index(m["col_screenshot"]) if m["col_screenshot"] else None
            idx_air_date = col_letter_to_index(m["col_air_date"]) if m["col_air_date"] else None
            mode_name = str(m.get("mode", "seeding")).strip().lower()
            try:
                block_start_line = max(1, int(str(m.get("start_line", 4)).strip() or 4))
            except Exception:
                block_start_line = 4
            start_offset = max(0, block_start_line - data_start_row)
            if mode_name == "scan":
                scan_expected_texts = (
                    list(unformatted_column_cache.get(idx_content, []))[start_offset:] if idx_content else []
                )
                scan_result_values = (
                    list(unformatted_column_cache.get(idx_drive, []))[start_offset:] if idx_drive else []
                )
                total_scan_rows = max(
                    len(scan_expected_texts),
                    len(scan_result_values),
                    (len(unformatted_column_cache.get(idx_url, [])) - start_offset) if idx_url else 0,
                    1,
                )
                links = resolve_links_for_scan_values(
                    list(unformatted_column_cache.get(idx_url, []))[start_offset : start_offset + total_scan_rows] if idx_url else [],
                    list(formula_column_cache.get(idx_url, []))[start_offset : start_offset + total_scan_rows] if idx_url else [],
                )
                links = filldown_scan_links_for_merged_rows(
                    links,
                    expected_texts=scan_expected_texts,
                    result_values=scan_result_values,
                )
                total_scan_rows = max(len(links), len(scan_expected_texts), len(scan_result_values))
                if len(scan_expected_texts) < total_scan_rows:
                    scan_expected_texts.extend([""] * (total_scan_rows - len(scan_expected_texts)))
                if len(scan_result_values) < total_scan_rows:
                    scan_result_values.extend([""] * (total_scan_rows - len(scan_result_values)))
                row_numbers = list(range(block_start_line, block_start_line + total_scan_rows))
            else:
                shown_links = list(unformatted_column_cache.get(idx_url, []))[start_offset:] if idx_url else []
                formula_links = list(formula_url_column_cache.get(idx_url, []))[start_offset:] if idx_url else []
                total_rows = max(len(shown_links), len(formula_links))
                links = []
                for i in range(total_rows):
                    shown = shown_links[i] if i < len(shown_links) else ""
                    formula = formula_links[i] if i < len(formula_links) else ""
                    resolved = _extract_http_url(str(shown))
                    if not resolved:
                        resolved = _extract_http_url(str(formula))
                    links.append(resolved or str(shown or ""))
                row_numbers = list(range(block_start_line, block_start_line + len(links)))
                scan_expected_texts = []
            if mode_name == "scan":
                results = scan_result_values
            else:
                results = list(unformatted_column_cache.get(idx_drive, []))[start_offset:] if idx_drive else []
            captions_existing = list(unformatted_column_cache.get(idx_content, []))[start_offset:] if idx_content else []
            screenshots_existing = list(unformatted_column_cache.get(idx_screenshot, []))[start_offset:] if idx_screenshot else []
            air_dates = list(unformatted_column_cache.get(idx_air_date, []))[start_offset:] if idx_air_date else []
            prepared_blocks.append(
                {
                    "block_index": int(m.get("block_index", 0)),
                    "name": m["name"],
                    "start_line": m["start_line"],
                    "col_url": m["col_url"],
                    "col_profile": m["col_profile"],
                    "col_content": m["col_content"],
                    "col_screenshot": m["col_screenshot"],
                    "col_drive": m["col_drive"],
                    "col_air_date": m["col_air_date"],
                    "idx_profile": idx_profile,
                    "idx_content": idx_content,
                    "idx_drive": idx_drive,
                    "idx_screenshot": idx_screenshot,
                    "idx_air_date": idx_air_date,
                    "fixed_air_date": m.get("fixed_air_date", ""),
                    "mode": mode_name or "seeding",
                    "links": links,
                    "row_numbers": row_numbers,
                    "scan_expected_texts": scan_expected_texts,
                    "results": results,
                    "captions_existing": captions_existing,
                    "screenshots_existing": screenshots_existing,
                    "air_dates": air_dates,
                }
            )
            if (str(m.get("mode", "seeding")).strip().lower() not in ("scan",)) and (idx_drive or idx_screenshot):
                need_upload = True
            max_rows = max(max_rows, len(links))

        write_log(f"[DEBUG] Using Drive Folder ID: {drive_id}")
        existing_files_by_name = load_existing_drive_files() if need_upload else {}
        if need_upload:
            write_log(f"[INFO] Drive folder preload complete: {len(existing_files_by_name)} distinct names")
        else:
            write_log("[INFO] Skip upload step (Drive/Screenshot column not configured).")

        try:
            only_error_mode = bool(getattr(app, "only_run_error_rows", None).get())
        except Exception:
            only_error_mode = False

        stored_error_rows = get_error_rows_for_sheet(sheet_url)
        stored_error_details = get_error_details_for_sheet(sheet_url)
        if only_error_mode:
            tracked_error_rows = set(stored_error_rows)
            tracked_error_details = dict(stored_error_details)
        else:
            # In normal runs, only show failures from the current run.
            # This avoids mixing historical pending rows into new job UI.
            tracked_error_rows = set()
            tracked_error_details = {}
        history_ready = True
        if only_error_mode:
            if tracked_error_rows:
                write_log(f"[INFO] Error-only mode: loaded {len(tracked_error_rows)} rows from history")
            else:
                write_log("[INFO] Error-only mode: no stored error rows for this sheet")
        elif stored_error_rows:
            write_log(
                f"[INFO] Normal mode: ignore {len(stored_error_rows)} historical error row(s); "
                "will show only current-run failures."
            )

        def _is_target_row(start_at: int, row_num: int, link_val: str, mode_key: str = "") -> bool:
            if row_num < start_at:
                return False
            if requested_rows and row_num not in requested_rows:
                return False
            if only_error_mode and row_num not in tracked_error_rows:
                return False
            mk = str(mode_key or "").strip().lower()
            if mk == "scan":
                return bool(normalize_scan_source_url(link_val))
            return str(link_val).strip().startswith("http")

        def _discover_candidate_url_columns(scan_start: int = 4) -> list[str]:
            letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            max_scan_cols = min(len(letters), int(getattr(worksheet, "col_count", 26) or 26))
            found: list[str] = []
            for col_idx in range(1, max_scan_cols + 1):
                try:
                    values = worksheet.col_values(col_idx)
                except Exception:
                    continue
                hit_count = 0
                for cell in values[max(0, scan_start - 1):]:
                    raw = str(cell or "").strip()
                    if raw.startswith("http"):
                        hit_count += 1
                if hit_count > 0:
                    found.append(f"{letters[col_idx - 1]}({hit_count})")
            return found

        target_total = 0
        eligibility_debug: list[dict] = []
        for block in prepared_blocks:
            block_mode_key = str(block.get("mode", "seeding")).strip().lower()
            row_nums = block.get("row_numbers") or []
            stats = {
                "name": str(block.get("name", "") or ""),
                "col_url": str(block.get("col_url", "") or "").strip().upper(),
                "start_line": int(block.get("start_line", 4) or 4),
                "total_rows": 0,
                "below_start": 0,
                "filtered_requested_rows": 0,
                "filtered_error_only": 0,
                "empty_link": 0,
                "invalid_link": 0,
                "eligible": 0,
                "first_nonempty_row": 0,
                "first_nonempty_value": "",
                "first_eligible_row": 0,
            }
            for i, lnk in enumerate(block["links"]):
                r = row_nums[i] if i < len(row_nums) else (i + 4)
                raw_link = str(lnk or "").strip()
                stats["total_rows"] += 1
                if r < int(block.get("start_line", 4) or 4):
                    stats["below_start"] += 1
                    continue
                if requested_rows and r not in requested_rows:
                    stats["filtered_requested_rows"] += 1
                    continue
                if only_error_mode and r not in tracked_error_rows:
                    stats["filtered_error_only"] += 1
                    continue
                if not raw_link:
                    stats["empty_link"] += 1
                    continue
                if not stats["first_nonempty_row"]:
                    stats["first_nonempty_row"] = int(r)
                    stats["first_nonempty_value"] = raw_link[:140]
                if block_mode_key == "scan":
                    if not normalize_scan_source_url(raw_link):
                        stats["invalid_link"] += 1
                        continue
                else:
                    if not raw_link.startswith("http"):
                        stats["invalid_link"] += 1
                        continue
                if _is_target_row(block["start_line"], r, raw_link, mode_key=block_mode_key):
                    target_total += 1
                    stats["eligible"] += 1
                    if not stats["first_eligible_row"]:
                        stats["first_eligible_row"] = int(r)
            eligibility_debug.append(stats)
        for s in eligibility_debug:
            try:
                start_row = int(s.get("start_line") or 0)
                first_valid_row = int(s.get("first_eligible_row") or 0)
                first_nonempty_row = int(s.get("first_nonempty_row") or 0)
            except Exception:
                continue
            if (
                first_valid_row > 0
                and start_row > 0
                and first_valid_row > start_row
                and not requested_rows
                and not only_error_mode
            ):
                col = str(s.get("col_url") or "").strip().upper() or "?"
                block_name = str(s.get("name") or "").strip() or "-"
                sample = str(s.get("first_nonempty_value") or "").strip()
                if first_nonempty_row > 0 and first_nonempty_row == start_row and sample:
                    write_log(
                        f"[INFO] {block_name}: Start Line={start_row} but {col}{start_row} is not a valid URL ('{sample}'). "
                        f"First eligible row is {first_valid_row}."
                    )
                else:
                    write_log(
                        f"[INFO] {block_name}: Start Line={start_row}. First eligible row is {first_valid_row}."
                    )
        if target_total == 0:
            configured_url_cols = ", ".join(
                sorted({str(block.get("col_url", "")).strip().upper() for block in prepared_blocks if str(block.get("col_url", "")).strip()})
            ) or "-"
            scan_start = min([int(block.get("start_line", 4) or 4) for block in prepared_blocks] or [4])
            detected_url_cols = _discover_candidate_url_columns(scan_start=scan_start)
            empty_msg = (
                f"Không có dòng hợp lệ để xử lý. Link URL hiện đang trỏ tới cột: {configured_url_cols}."
            )
            if detected_url_cols:
                empty_msg += f" Sheet này đang có URL ở: {', '.join(detected_url_cols[:8])}."
            if eligibility_debug:
                parts = []
                for s in eligibility_debug[:4]:
                    parts.append(
                        f"{s['name'] or '-'}[{s['col_url'] or '-'}]: "
                        f"rows={s['total_rows']}, ok={s['eligible']}, "
                        f"belowStart={s['below_start']}, reqFilter={s['filtered_requested_rows']}, "
                        f"retryFilter={s['filtered_error_only']}, empty={s['empty_link']}, badLink={s['invalid_link']}"
                    )
                empty_msg += " Chi tiết lọc: " + " | ".join(parts) + "."
            empty_msg += " Kiểm tra lại Link URL, Start Line hoặc chế độ retry."
            write_log("[WARN] target_total=0: no eligible links found to process.")
            ui_call(ui_set_detail, empty_msg)
            ui_call(ui_set_status, "KHÔNG CÓ DỮ LIỆU", "#b45309")
            ui_call(ui_add_log, 0, "WARN", "EMPTY", empty_msg, "warning")

        run_started_at = time.time()
        started_count = 0
        processed_count = 0
        success_count = 0
        fail_count = 0
        unavailable_count = 0
        ui_call(ui_update_summary, 0, target_total, 0, 0, "---", 0)

        counter_lock = threading.Lock()
        error_lock = threading.Lock()
        drive_cache_lock = threading.Lock()
        sheet_write_lock = threading.Lock()
        last_sheet_write_ts = [0.0]
        # Keep per-row failing blocks to avoid clearing errors when another block on same row succeeds.
        row_issue_blocks: dict[int, set[str]] = {}

        def _is_quota_error(exc: Exception) -> bool:
            s = str(exc).lower()
            return ("429" in s) or ("quota exceeded" in s) or ("rate limit" in s)

        def safe_sheet_write(write_fn, op_desc: str = "sheet_write", max_retry: int = 8):
            """
            Serialize + throttle writes to avoid Google Sheets write quota bursts
            when multiple blocks/workers run in parallel.
            """
            base_wait = 0.55 if scan_only_request else 0.85
            last_err = None
            for attempt in range(max_retry):
                try:
                    with sheet_write_lock:
                        now = time.time()
                        wait_more = base_wait - (now - last_sheet_write_ts[0])
                        if wait_more > 0:
                            time.sleep(wait_more)
                        out = write_fn()
                        last_sheet_write_ts[0] = time.time()
                        return out
                except Exception as e:
                    last_err = e
                    if not _is_quota_error(e):
                        raise
                    sleep_s = min(20.0, (1.3 ** attempt) * 1.2)
                    write_log(
                        f"[WARN] {op_desc} quota hit (attempt {attempt + 1}/{max_retry}), sleep {sleep_s:.1f}s: {e}"
                    )
                    time.sleep(sleep_s)
            if last_err:
                raise last_err

        def _calc_eta(done_count: int) -> str:
            elapsed = max(0.0, time.time() - run_started_at)
            if done_count > 0 and done_count < target_total:
                avg = elapsed / done_count
                remain = int(avg * (target_total - done_count))
                return f"{remain}s"
            return "---"

        def _start_row(block_name: str, row: int, url: str):
            nonlocal started_count
            with counter_lock:
                started_count += 1
                done = processed_count
                okv = success_count
                failv = fail_count
                unavailv = unavailable_count
                percent = int((done / max(1, target_total)) * 100)
                eta = _calc_eta(done)
            ui_call(ui_set_progress, percent)
            ui_call(ui_set_detail, f"{block_name} - hàng {row}")
            ui_call(ui_update_summary, done, target_total, okv, failv, eta, unavailv)
            ui_call(ui_add_log, row, "START", "START", f"{block_name}: Link {url[:110]}", "start")
            return eta

        def _finish_row_ok(
            block_name: str,
            row: int,
            url: str,
            eta: str,
            msg: str | None = None,
            log_tag: str = "ok",
            issue_columns: list[str] | None = None,
        ):
            nonlocal processed_count, success_count, fail_count, unavailable_count
            clear_row_error = False
            with error_lock:
                row_blocks = set(row_issue_blocks.get(row) or set())
                if block_name in row_blocks:
                    row_blocks.discard(block_name)
                    if row_blocks:
                        row_issue_blocks[row] = row_blocks
                    else:
                        row_issue_blocks.pop(row, None)
                if not row_blocks:
                    tracked_error_rows.discard(row)
                    tracked_error_details.pop(row, None)
                    clear_row_error = True
            with counter_lock:
                processed_count += 1
                success_count += 1
                done = processed_count
                okv = success_count
                failv = fail_count
                unavailv = unavailable_count
                percent = int((done / max(1, target_total)) * 100)
                eta = _calc_eta(done)
            text = msg if msg else f"{block_name}: {url[:110]}"
            ui_call(ui_add_log, row, "OK", "OK", text, log_tag)
            if clear_row_error and hasattr(app, "update_error_row_live"):
                ui_call(app.update_error_row_live, row, "", False)
            if hasattr(app, "update_issue_cells_live"):
                ui_call(app.update_issue_cells_live, row, block_name, issue_columns or [], "", "", True)
            ui_call(ui_set_progress, percent)
            ui_call(ui_update_summary, done, target_total, okv, failv, eta, unavailv)

        def _finish_row_fail(block_name: str, row: int, err: str, eta: str, issue_columns: list[str] | None = None):
            nonlocal processed_count, fail_count, unavailable_count
            with error_lock:
                row_issue_blocks.setdefault(row, set()).add(block_name)
                tracked_error_rows.add(row)
                err_text = str(err).strip()
                if not err_text:
                    err_text = "Lỗi xử lý"
                if err_text.lower().startswith(str(block_name).strip().lower() + ":"):
                    msg_store = err_text
                else:
                    msg_store = f"{block_name}: {err_text}"
                tracked_error_details[row] = msg_store[:220]
            with counter_lock:
                processed_count += 1
                fail_count += 1
                done = processed_count
                okv = success_count
                failv = fail_count
                unavailv = unavailable_count
                percent = int((done / max(1, target_total)) * 100)
                eta = _calc_eta(done)
            ui_call(ui_add_log, row, "FAIL", "FAIL", f"{block_name}: {err}", "fail")
            if hasattr(app, "update_error_row_live"):
                ui_call(app.update_error_row_live, row, msg_store, True)
            if hasattr(app, "update_issue_cells_live"):
                ui_call(app.update_issue_cells_live, row, block_name, issue_columns or [], msg_store, "failed", False)
            ui_call(ui_set_progress, percent)
            ui_call(ui_update_summary, done, target_total, okv, failv, eta, unavailv)

        def _finish_row_unavailable(block_name: str, row: int, message: str, eta: str, issue_columns: list[str] | None = None):
            nonlocal processed_count, fail_count, unavailable_count
            text = str(message or "").strip() or "Nội dung không khả dụng"
            with error_lock:
                row_issue_blocks.setdefault(row, set()).add(block_name)
                tracked_error_rows.add(row)
                msg_store = f"{block_name}: {text}"
                tracked_error_details[row] = msg_store[:220]
            with counter_lock:
                processed_count += 1
                unavailable_count += 1
                done = processed_count
                okv = success_count
                failv = fail_count
                unavailv = unavailable_count
                percent = int((done / max(1, target_total)) * 100)
                eta = _calc_eta(done)
            ui_call(ui_add_log, row, "WARN", "UNAVAILABLE", f"{block_name}: {text}", "unavailable")
            if hasattr(app, "update_error_row_live"):
                ui_call(app.update_error_row_live, row, msg_store, True)
            if hasattr(app, "update_issue_cells_live"):
                ui_call(app.update_issue_cells_live, row, block_name, issue_columns or [], msg_store, "unavailable", False)
            ui_call(ui_set_progress, percent)
            ui_call(ui_update_summary, done, target_total, okv, failv, eta, unavailv)

        def _webdriver_session_alive(driver_obj) -> bool:
            if driver_obj is None:
                return False
            try:
                _ = str(driver_obj.current_url or "")
                return True
            except Exception:
                return False

        def _is_recoverable_driver_error(exc: Exception) -> bool:
            msg = str(exc or "").strip().lower()
            if not msg:
                return False
            recoverable_markers = (
                "disconnected",
                "unable to connect to renderer",
                "failed to check if window was closed",
                "target window already closed",
                "invalid session id",
                "session deleted",
                "chrome not reachable",
                "web view not found",
            )
            return any(marker in msg for marker in recoverable_markers)

        def _start_worker_driver(block_index: int, block_mode: str, fast_recovery: bool = False):
            if scan_only_request:
                return None
            if block_index == 0 and app.driver:
                if _webdriver_session_alive(app.driver):
                    return app.driver
                try:
                    app.driver.quit()
                except Exception:
                    pass
                app.driver = None
            worker_port = get_post_port(block_index, browser_port)
            worker_profile = get_block_profile(block_index, block_mode, browser_port=worker_port)
            os.makedirs(worker_profile, exist_ok=True)
            # If the user already opened this block's Chrome and logged in manually,
            # attach to that exact debugger session instead of starting a fresh browser.
            try:
                attach_opts = Options()
                attach_opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{worker_port}")
                attached = create_chrome_driver(attach_opts, service=service)
                write_log(f"[INFO] Attached worker block {block_index} to existing Chrome debug session on port {worker_port}")
                return attached
            except Exception as e:
                write_log(f"[INFO] No attachable worker Chrome on port {worker_port}: {e}")
            # Copy login session from main profile so parallel workers don't ask login again.
            # In fast recovery mode we skip this expensive copy step to restart sooner.
            if not fast_recovery:
                try:
                    has_local_session = os.path.isdir(os.path.join(worker_profile, "Default"))
                    seed_profile = LOCAL_PROFILE_PATH if os.path.isdir(LOCAL_PROFILE_PATH) else ""
                    if not seed_profile and os.path.isdir(FB_PROFILE_PATH):
                        seed_profile = FB_PROFILE_PATH
                    if not seed_profile and os.path.isdir(FB_PROFILE_PATH_ALT):
                        seed_profile = FB_PROFILE_PATH_ALT
                    if seed_profile and not has_local_session:
                        shutil.copytree(
                            seed_profile,
                            worker_profile,
                            dirs_exist_ok=True,
                            ignore=shutil.ignore_patterns(
                                "Cache",
                                "Code Cache",
                                "GPUCache",
                                "GrShaderCache",
                                "ShaderCache",
                                "Crashpad",
                                "Singleton*",
                                "lockfile",
                                "*.tmp",
                            ),
                        )
                except Exception as e:
                    write_log(f"[WARN] Worker profile seed failed ({block_index}): {e}")
            last = None
            # Local UX: prefer visible Chrome first so user can re-login when needed.
            headless_plan = (False,) if fast_recovery else (False, True)
            for headless in headless_plan:
                try:
                    return create_chrome_driver(
                        build_chrome_options(user_data_dir=worker_profile, headless=headless, debug_port=worker_port),
                        service=service,
                    )
                except Exception as e:
                    last = e
            raise Exception(f"WORKER_CHROME_START_FAILED[{block_index}]: {last}")

        def _run_block(block: dict):
            block_index = int(block.get("block_index", 0))
            worker_driver = _start_worker_driver(block_index, str(block.get("mode", "seeding")))
            local_is_main_driver = bool(worker_driver) and (worker_driver is app.driver)
            try:
                # Each worker gets its own Google API clients to avoid SSL/socket corruption
                # when sharing httplib2 transport across threads.
                local_creds = ServiceAccountCredentials.from_json_keyfile_name(
                    JSON_PATH,
                    [
                        "https://spreadsheets.google.com/feeds",
                        "https://www.googleapis.com/auth/drive",
                    ],
                )
                local_client = gspread.authorize(local_creds)
                local_spreadsheet = local_client.open_by_url(sheet_url)
                local_worksheet = local_spreadsheet.worksheet(sheet_name)
                local_drive_service = build("drive", "v3", credentials=local_creds)
                block_name = block["name"]
                idx_profile = block["idx_profile"]
                idx_content = block["idx_content"]
                idx_drive = block["idx_drive"]
                idx_screenshot = block["idx_screenshot"]
                idx_air_date = block["idx_air_date"]
                fixed_air_date = str(block.get("fixed_air_date", "")).strip()
                block_mode = str(block.get("mode", "seeding")).strip().lower()
                is_scan_mode = block_mode == "scan"
                try:
                    scan_negative_filter_enabled = is_scan_mode and bool(getattr(app, "scan_negative_filter", None).get())
                    scan_keyword_filter_enabled = is_scan_mode and bool(getattr(app, "scan_keyword_filter", None).get())
                except Exception:
                    scan_negative_filter_enabled = False
                    scan_keyword_filter_enabled = False
                col_profile_letter = str(block.get("col_profile", "")).strip().upper()
                col_content_letter = str(block.get("col_content", "")).strip().upper()
                col_screenshot_letter = str(block.get("col_screenshot", "")).strip().upper()
                col_drive_letter = str(block.get("col_drive", "")).strip().upper()
                highlight_columns: list[int] = []
                issue_columns: list[str] = []
                if is_scan_mode:
                    if idx_drive:
                        highlight_columns.append(idx_drive)
                        if col_drive_letter:
                            issue_columns.append(col_drive_letter)
                else:
                    for col_idx, col_letter in (
                        (idx_profile, col_profile_letter),
                        (idx_content, col_content_letter),
                        (idx_drive, col_drive_letter),
                        (idx_screenshot, col_screenshot_letter),
                    ):
                        if col_idx and col_idx not in highlight_columns:
                            highlight_columns.append(col_idx)
                        if col_idx and col_letter:
                            normalized_letter = str(col_letter).strip().upper()
                            if normalized_letter and normalized_letter not in issue_columns:
                                issue_columns.append(normalized_letter)
                log_block_name = block_name
                links = block["links"]
                row_numbers = block.get("row_numbers") or []
                scan_expected_texts = block.get("scan_expected_texts", [])
                results = block["results"]
                captions_existing = block["captions_existing"]
                screenshots_existing = block["screenshots_existing"]
                air_dates = block["air_dates"]
                start_at = block["start_line"]

                def _highlight_row_cells(row: int, fill_rgb: dict[str, float], color_name: str):
                    if not highlight_sheet_errors_enabled or not highlight_columns or row < 1:
                        return
                    requests = []
                    for col_idx in highlight_columns:
                        requests.append(
                            {
                                "repeatCell": {
                                    "range": {
                                        "sheetId": int(local_worksheet.id),
                                        "startRowIndex": int(row - 1),
                                        "endRowIndex": int(row),
                                        "startColumnIndex": int(col_idx - 1),
                                        "endColumnIndex": int(col_idx),
                                    },
                                    "cell": {
                                        "userEnteredFormat": {
                                            "backgroundColor": dict(fill_rgb),
                                        }
                                    },
                                    "fields": "userEnteredFormat.backgroundColor",
                                }
                            }
                        )
                    if not requests:
                        return
                    safe_sheet_write(
                        lambda: local_spreadsheet.batch_update({"requests": requests}),
                        op_desc=f"highlight_{color_name}_row_{row}",
                    )

                def _mark_row_unavailable(row: int):
                    _highlight_row_cells(
                        row,
                        {"red": 1.0, "green": 0.95, "blue": 0.68},
                        "unavailable",
                    )

                def _mark_row_failed(row: int):
                    _highlight_row_cells(
                        row,
                        {"red": 0.98, "green": 0.82, "blue": 0.82},
                        "failed",
                    )

                def _mark_row_success(row: int):
                    _highlight_row_cells(
                        row,
                        {"red": 1.0, "green": 1.0, "blue": 1.0},
                        "success",
                    )

                try:
                    multi_capture_enabled = bool(getattr(app, "capture_five_per_link", None).get())
                except Exception:
                    multi_capture_enabled = False
                captures_per_link = 5 if (multi_capture_enabled and block_mode == "booking") else 1

                retry_once_by_row: set[int] = set()
                pending_indexes = list(range(len(links)))
                while pending_indexes:
                    idx = pending_indexes.pop(0)
                    url = links[idx] if idx < len(links) else ""
                    if not app.is_running:
                        break
                    while getattr(app, "is_paused", False):
                        time.sleep(0.5)

                    row = row_numbers[idx] if idx < len(row_numbers) else (idx + 4)
                    if row < start_at:
                        continue
                    if requested_rows and row not in requested_rows:
                        continue
                    if only_error_mode and row not in tracked_error_rows:
                        continue

                    url = str(url).strip()
                    if is_scan_mode:
                        url = normalize_scan_source_url(url)
                        if not url:
                            continue
                    else:
                        if not url.startswith("http"):
                            continue

                    expected_scan_text = ""
                    if is_scan_mode and idx < len(scan_expected_texts):
                        expected_scan_text = str(scan_expected_texts[idx]).strip()

                    eta_text = _start_row(log_block_name, row, url)

                    if (not is_scan_mode) and (not app.force_run_all.get()):
                        drive_ready = False
                        content_ready = False
                        screenshot_ready = False
                        if idx_drive and idx < len(results):
                            drive_val = str(results[idx] or "").strip().lower()
                            drive_ready = ("drive.google.com" in drive_val) or drive_val.startswith("http")
                        if idx_content and idx < len(captions_existing):
                            content_ready = bool(str(captions_existing[idx] or "").strip())
                        if idx_screenshot and idx < len(screenshots_existing):
                            shot_val = str(screenshots_existing[idx] or "").strip().lower()
                            screenshot_ready = bool(shot_val) and (
                                "drive.google.com" in shot_val
                                or shot_val.startswith("http")
                                or "image(" in shot_val
                            )

                        required_outputs_ready = True
                        if idx_drive:
                            required_outputs_ready = required_outputs_ready and drive_ready
                        if idx_content:
                            required_outputs_ready = required_outputs_ready and content_ready
                        if idx_screenshot:
                            required_outputs_ready = required_outputs_ready and screenshot_ready

                        if required_outputs_ready:
                            ui_call(
                                ui_add_log,
                                row,
                                "INFO",
                                "SKIP",
                                f"{log_block_name}: Bỏ qua vì dòng đã có đủ output (Drive/Caption/Screenshot)",
                                "start",
                            )
                            _finish_row_ok(log_block_name, row, url, eta_text)
                            continue

                    try:
                        unavailable = False
                        is_tiktok = False
                        tiktok_shop_app_only = False
                        profile_name = ""
                        caption = ""
                        _post_time = ""
                        tiktok_oembed_payload: dict = {}
                        tiktok_oembed_png: bytes = b""
                        ocr_text = ""
                        text_scan_source = ""
                        if is_scan_mode:
                            try:
                                image_bytes = download_image_bytes_for_scan(url, drive_service=local_drive_service)
                                if image_bytes:
                                    ocr_text = ocr_text_from_image_bytes(image_bytes, expected_text=expected_scan_text)
                            except Exception as ocr_e:
                                write_log(f"[WARN] OCR failed row {row}: {ocr_e}")
                                ocr_text = ""
                        else:
                            worker_driver.get(url)
                            wait_page_ready(worker_driver, timeout=PAGE_READY_TIMEOUT)

                            url_lower = url.lower()
                            is_tiktok = "tiktok.com" in url_lower or "vt.tiktok.com" in url_lower
                            if is_tiktok:
                                # Redirect chains (especially vt.tiktok.com) may need a short settle window.
                                wait_tiktok_redirect_ready(
                                    worker_driver,
                                    requested_url=url,
                                    timeout_sec=TIKTOK_REDIRECT_WAIT_SEC,
                                )
                                try:
                                    worker_driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    time.sleep(TIKTOK_SCROLL_WAIT_1)
                                    worker_driver.execute_script("window.scrollTo(0, window.innerHeight / 2);")
                                    time.sleep(TIKTOK_SCROLL_WAIT_2)
                                except Exception:
                                    pass
                                if is_tiktok_slider_challenge_present(worker_driver):
                                    focused_now = focus_chrome_for_tiktok_challenge(worker_driver)
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "CAPTCHA",
                                        (
                                            f"{block_name}: TikTok yêu cầu kéo xác minh. "
                                            f"Đã đẩy cửa sổ lên trước ({'OK' if focused_now else 'limited'}), chờ bạn xử lý..."
                                        ),
                                        "start",
                                    )
                                    solved, waited_sec = wait_for_tiktok_slider_clear(worker_driver)
                                    if not solved:
                                        raise Exception(f"TIKTOK_CAPTCHA_TIMEOUT_{int(waited_sec)}s")
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "OK",
                                        "CAPTCHA",
                                        f"{block_name}: Xác minh TikTok thành công, tiếp tục chụp.",
                                        "ok",
                                    )
                                    time.sleep(TIKTOK_CAPTCHA_POST_CLEAR_WAIT_SEC)

                                # TikTok may randomly show "Access Denied" on first load but succeed after reload.
                                max_denied_retry = max(1, int(TIKTOK_ACCESS_DENIED_RETRY_MAX or 4))
                                denied_cleared = False
                                for denied_attempt in range(max_denied_retry + 1):
                                    if not is_tiktok_access_denied_page(worker_driver, url):
                                        denied_cleared = True
                                        break
                                    if denied_attempt >= max_denied_retry:
                                        break
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "RETRY",
                                        (
                                            f"{block_name}: TikTok tạm chặn Access Denied, "
                                            f"đang thử tải lại ({denied_attempt + 1}/{max_denied_retry})..."
                                        ),
                                        "start",
                                    )
                                    try:
                                        worker_driver.get(url)
                                        wait_page_ready(worker_driver, timeout=PAGE_READY_TIMEOUT)
                                        wait_tiktok_redirect_ready(
                                            worker_driver,
                                            requested_url=url,
                                            timeout_sec=TIKTOK_REDIRECT_WAIT_SEC,
                                        )
                                        time.sleep(max(0.6, float(TIKTOK_ACCESS_DENIED_RETRY_SLEEP_SEC or 1.6)))
                                    except Exception:
                                        pass
                                if not denied_cleared and is_tiktok_access_denied_page(worker_driver, url):
                                    unavailable = True
                                    profile_name, caption = "", "Nội dung không khả dụng"
                                    _post_time = ""
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "UNAVAILABLE",
                                        (
                                            f"{block_name}: TikTok bị chặn Access Denied sau "
                                            f"{max_denied_retry} lần thử, đánh dấu không khả dụng"
                                        ),
                                        "unavailable",
                                    )
                                if not unavailable:
                                    try:
                                        current_tiktok_url = wait_tiktok_redirect_ready(
                                            worker_driver,
                                            requested_url=url,
                                            timeout_sec=TIKTOK_REDIRECT_WAIT_SEC,
                                        )
                                    except Exception:
                                        current_tiktok_url = ""
                                    if current_tiktok_url and (not _is_expected_tiktok_page(url, current_tiktok_url)):
                                        raise Exception(
                                            f"TIKTOK_URL_MISMATCH expected={url[:120]} current={current_tiktok_url[:120]}"
                                        )
                            time.sleep(PER_LINK_BASE_WAIT)

                            unavailable = unavailable or is_unavailable_content_page(worker_driver, url)
                            if is_tiktok and unavailable:
                                tiktok_shop_app_only = is_tiktok_shop_app_only_notice(worker_driver, url)
                                if tiktok_shop_app_only:
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "UNAVAILABLE",
                                        f"{block_name}: TikTok Shop yêu cầu xem trong app, bỏ qua fallback ảnh",
                                        "unavailable",
                                    )
                                else:
                                    try:
                                        tiktok_oembed_payload = fetch_tiktok_oembed_data(url)
                                    except Exception:
                                        tiktok_oembed_payload = {}
                                    oembed_author = str(tiktok_oembed_payload.get("author_name") or "").strip()
                                    oembed_title = str(tiktok_oembed_payload.get("title") or "").strip()
                                    oembed_thumb = str(tiktok_oembed_payload.get("thumbnail_url") or "").strip()
                                    if oembed_thumb:
                                        try:
                                            tiktok_oembed_png = download_image_bytes_for_scan(
                                                oembed_thumb,
                                                timeout=20,
                                                drive_service=None,
                                            )
                                        except Exception:
                                            tiktok_oembed_png = b""
                                    if oembed_author:
                                        profile_name = oembed_author
                                    if oembed_title:
                                        caption = oembed_title
                                    if tiktok_oembed_png:
                                        unavailable = False
                                        ui_call(
                                            ui_add_log,
                                            row,
                                            "INFO",
                                            "FALLBACK",
                                            f"{block_name}: TikTok bị chặn, dùng oEmbed thumbnail để tiếp tục",
                                            "ok",
                                        )
                            if unavailable:
                                profile_name, caption = "", "Nội dung không khả dụng"
                                _post_time = ""
                            elif is_tiktok and tiktok_oembed_payload and (tiktok_oembed_png or profile_name or caption):
                                # Keep oEmbed metadata when page access is blocked but fallback data exists.
                                pass
                            else:
                                profile_name, caption = get_fb_profile_and_caption(worker_driver, url)
                                _post_time = get_fb_post_datetime(worker_driver)
                            try:
                                worker_driver.execute_cdp_cmd(
                                    "Emulation.setPageScaleFactor",
                                    {"pageScaleFactor": CAPTURE_ZOOM_PERCENT / 100.0},
                                )
                            except Exception:
                                pass
                            try:
                                worker_driver.execute_script(f"document.body.style.zoom='{CAPTURE_ZOOM_PERCENT}%'")
                                time.sleep(ZOOM_SETTLE_SLEEP)
                            except Exception:
                                pass

                        link_drive = ""
                        direct_url = ""
                        if (not is_scan_mode) and (idx_drive or idx_screenshot):
                            effective_captures = 1 if (unavailable or bool(tiktok_oembed_png)) else captures_per_link
                            sheet_air_raw = str(air_dates[idx]).strip() if (idx_air_date and idx < len(air_dates)) else ""
                            air_date = get_air_date_token(sheet_air_raw) or fixed_air_date or get_air_date_token(_post_time)
                            platform_token = sanitize_filename_token(detect_platform_label(url), fallback="Other", max_len=24)
                            kol_token = sanitize_filename_token(profile_name, fallback="UnknownKOL", max_len=60)
                            date_token = sanitize_filename_token(air_date, fallback="NoDate", max_len=16)
                            base_name = f"Post_{platform_token}_{kol_token}_{date_token}_Row_{row}"
                            captured_pngs: list[bytes] = []

                            def _upload_png_as(file_name: str, png_data: bytes):
                                media = MediaIoBaseUpload(io.BytesIO(png_data), mimetype="image/png", resumable=False)
                                with drive_cache_lock:
                                    existing_files = list(existing_files_by_name.get(file_name, []))

                                if existing_files:
                                    is_new_file_local = False
                                    file_id_local = existing_files[0]
                                    if len(existing_files) > 1:
                                        for dup in existing_files[1:]:
                                            try:
                                                local_drive_service.files().delete(fileId=dup, supportsAllDrives=True).execute()
                                            except Exception:
                                                pass
                                        with drive_cache_lock:
                                            existing_files_by_name[file_name] = [file_id_local]
                                    local_drive_service.files().update(
                                        fileId=file_id_local,
                                        media_body=media,
                                        supportsAllDrives=True,
                                    ).execute()
                                else:
                                    is_new_file_local = True
                                    file_meta_local = {"name": file_name, "parents": [drive_id]}
                                    uploaded_local = local_drive_service.files().create(
                                        body=file_meta_local,
                                        media_body=media,
                                        fields="id",
                                        supportsAllDrives=True,
                                    ).execute()
                                    file_id_local = uploaded_local.get("id")
                                    if not file_id_local:
                                        raise Exception("UPLOAD_FAIL")
                                    with drive_cache_lock:
                                        existing_files_by_name[file_name] = [file_id_local]

                                if is_new_file_local:
                                    local_drive_service.permissions().create(
                                        fileId=file_id_local,
                                        body={"type": "anyone", "role": "reader"},
                                        supportsAllDrives=True,
                                    ).execute()
                                file_info_local = local_drive_service.files().get(
                                    fileId=file_id_local,
                                    fields="webViewLink",
                                    supportsAllDrives=True,
                                ).execute()
                                web_link_local = file_info_local.get("webViewLink")
                                direct_local = f"https://drive.google.com/uc?export=view&id={file_id_local}&ts={int(time.time())}"
                                return file_id_local, web_link_local, direct_local

                            for shot_idx in range(1, effective_captures + 1):
                                using_oembed_capture = bool(shot_idx == 1 and tiktok_oembed_png)
                                png_bytes = tiktok_oembed_png if using_oembed_capture else b""
                                if shot_idx == 1 and (not using_oembed_capture):
                                    if has_please_wait_overlay(worker_driver):
                                        ui_call(
                                            ui_add_log,
                                            row,
                                            "INFO",
                                            "WAIT",
                                            f"{block_name}: Phát hiện 'Please wait', đợi thêm {int(PLEASE_WAIT_EXTRA_CAPTURE_DELAY_SEC)}s trước khi chụp",
                                            "start",
                                        )
                                        time.sleep(PLEASE_WAIT_EXTRA_CAPTURE_DELAY_SEC)
                                        cleared_wait, waited_wait = wait_for_please_wait_clear(worker_driver)
                                        if not cleared_wait:
                                            unavailable = True
                                            profile_name, caption = "", "Nội dung không khả dụng"
                                            _post_time = ""
                                            ui_call(
                                                ui_add_log,
                                                row,
                                                "WARN",
                                                "UNAVAILABLE",
                                                (
                                                    f"{block_name}: Trang kẹt 'Please wait' hơn "
                                                    f"{int(waited_wait)}s, đánh dấu không khả dụng"
                                                ),
                                                "unavailable",
                                            )
                                            break
                                    first_capture_delay = SCREENSHOT_CAPTURE_DELAY + (
                                        TIKTOK_FIRST_CAPTURE_EXTRA_SEC if is_tiktok else 0.0
                                    )
                                    time.sleep(first_capture_delay)
                                else:
                                    time.sleep(MULTI_CAPTURE_INTERVAL_SEC)
                                if not png_bytes:
                                    png_bytes = worker_driver.get_screenshot_as_png()
                                blank_retry = max(0, int(BLANK_SCREEN_MAX_RETRIES or 0))
                                blank_attempt = 0
                                while (not using_oembed_capture) and blank_attempt < blank_retry and is_blank_like_screenshot_png(png_bytes):
                                    blank_attempt += 1
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "RETRY",
                                        (
                                            f"{block_name}: Ảnh chụp có vẻ trắng/rỗng, "
                                            f"đợi {int(BLANK_SCREEN_RETRY_DELAY_SEC)}s rồi chụp lại "
                                            f"({blank_attempt}/{blank_retry})"
                                        ),
                                        "start",
                                    )
                                    time.sleep(max(0.2, float(BLANK_SCREEN_RETRY_DELAY_SEC or 2.0)))
                                    png_bytes = worker_driver.get_screenshot_as_png()
                                if (not using_oembed_capture) and is_blank_like_screenshot_png(png_bytes):
                                    unavailable = True
                                    profile_name, caption = "", "Nội dung không khả dụng"
                                    _post_time = ""
                                    link_drive = ""
                                    direct_url = ""
                                    captured_pngs = []
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "WARN",
                                        "UNAVAILABLE",
                                        (
                                            f"{block_name}: Chụp lại sau {int(BLANK_SCREEN_RETRY_DELAY_SEC)}s "
                                            "vẫn trắng/rỗng, đánh dấu không khả dụng"
                                        ),
                                        "unavailable",
                                    )
                                    break
                                captured_pngs.append(png_bytes)
                                if effective_captures > 1:
                                    file_name = f"{base_name}_S{shot_idx}.png"
                                else:
                                    file_name = f"{base_name}.png"
                                file_id, web_link, direct_link = _upload_png_as(file_name, png_bytes)
                                if shot_idx == 1:
                                    link_drive = web_link
                                    direct_url = direct_link
                                ui_call(
                                    ui_add_log,
                                    row,
                                    "INFO",
                                    "SHOT",
                                    f"{block_name}: Đã chụp {shot_idx}/{effective_captures}",
                                    "start",
                                )
                            if effective_captures > 1 and idx_screenshot:
                                collage_png = build_collage_png(captured_pngs)
                                if collage_png:
                                    collage_name = f"{base_name}_ALL.png"
                                    _fid_all, web_all, direct_all = _upload_png_as(collage_name, collage_png)
                                    link_drive = web_all or link_drive
                                    direct_url = direct_all or direct_url
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "INFO",
                                        "SHOT",
                                        f"{block_name}: Đã gộp {captures_per_link} ảnh vào 1 ô",
                                        "start",
                                    )

                        is_youtube = ("youtube.com" in url) or ("youtu.be" in url)
                        is_facebook = ("facebook.com" in url) or ("fb.watch" in url) or ("m.facebook.com" in url)
                        if unavailable:
                            col_i = "Nội dung không khả dụng"
                        elif is_youtube:
                            profile_name = get_youtube_channel(worker_driver) or profile_name
                            col_i = (get_youtube_title(worker_driver) or "").strip()
                        else:
                            if is_facebook:
                                profile_name = clean_fb_profile_name(profile_name)
                            col_i = caption.strip() if caption else ""
                        profile_name = normalize_account_name(profile_name, url)
                        if (not unavailable) and is_facebook and (not profile_name) and (not col_i.strip()):
                            unavailable = True
                            col_i = "Nội dung không khả dụng"

                        updates = []
                        scan_ok = False
                        negative_term_hit = ""
                        keyword_term_hit = ""
                        if is_scan_mode:
                            if scan_negative_filter_enabled:
                                negative_term_hit = detect_scan_negative_term(expected_scan_text, ocr_text)
                            if scan_keyword_filter_enabled:
                                keyword_term_hit = detect_scan_keyword_term(expected_scan_text, ocr_text)
                            scan_ok = is_scan_match(expected_scan_text, ocr_text)
                            e_norm = normalize_match_text(expected_scan_text)
                            o_norm = normalize_match_text(ocr_text)
                            ratio_dbg = difflib.SequenceMatcher(None, e_norm, o_norm).ratio() if e_norm and o_norm else 0.0
                            write_log(
                                f"[SCAN] row={row} match={int(scan_ok)} ratio={ratio_dbg:.2f} "
                                f"exp='{e_norm[:90]}' ocr='{o_norm[:90]}'"
                            )
                            if idx_drive:
                                updates.append(
                                    {
                                        "range": f"{col_drive_letter}{row}",
                                        "values": [["1" if (scan_ok and not negative_term_hit) else "0"]],
                                    }
                                )
                        else:
                            if idx_profile and profile_name:
                                updates.append({"range": f"{col_profile_letter}{row}", "values": [[profile_name]]})
                            if idx_drive:
                                updates.append({"range": f"{col_drive_letter}{row}", "values": [[link_drive]]})
                            if idx_screenshot and direct_url:
                                updates.append({"range": f"{col_screenshot_letter}{row}", "values": [[f'=IMAGE(\"{direct_url}\")']]})
                            if idx_content and col_i:
                                updates.append({"range": f"{col_content_letter}{row}", "values": [[col_i]]})

                        if updates:
                            safe_sheet_write(
                                lambda: local_worksheet.batch_update(updates, value_input_option="USER_ENTERED"),
                                op_desc=f"batch_update_row_{row}",
                            )
                        if is_scan_mode:
                            if negative_term_hit:
                                _mark_row_failed(row)
                                _finish_row_fail(
                                    log_block_name,
                                    row,
                                    f"Từ tiêu cực: {negative_term_hit}",
                                    eta_text,
                                    issue_columns=issue_columns,
                                )
                            elif keyword_term_hit:
                                _mark_row_failed(row)
                                _finish_row_fail(
                                    log_block_name,
                                    row,
                                    f"Từ khóa: {keyword_term_hit}",
                                    eta_text,
                                    issue_columns=issue_columns,
                                )
                            elif scan_ok:
                                _mark_row_success(row)
                                _finish_row_ok(
                                    log_block_name,
                                    row,
                                    url,
                                    eta_text,
                                    msg=f"{log_block_name}: MATCH",
                                    log_tag="ok",
                                    issue_columns=issue_columns,
                                )
                            else:
                                _mark_row_failed(row)
                                _finish_row_fail(log_block_name, row, "NO_MATCH", eta_text, issue_columns=issue_columns)
                        elif unavailable:
                            _mark_row_unavailable(row)
                            _finish_row_unavailable(log_block_name, row, "Nội dung không khả dụng", eta_text, issue_columns=issue_columns)
                        else:
                            _mark_row_success(row)
                            _finish_row_ok(log_block_name, row, url, eta_text, issue_columns=issue_columns)
                    except Exception as e:
                        if (not is_scan_mode) and _is_recoverable_driver_error(e):
                            ui_call(
                                ui_add_log,
                                row,
                                "WARN",
                                "RETRY",
                                f"{log_block_name}: Chrome bị ngắt kết nối, đang khởi động lại phiên trình duyệt...",
                                "start",
                            )
                            write_log(f"[WARN] Recoverable browser error at row {row}: {e}")
                            try:
                                if (not local_is_main_driver) and worker_driver:
                                    try:
                                        worker_driver.quit()
                                    except Exception:
                                        pass
                                worker_driver = _start_worker_driver(
                                    block_index,
                                    str(block.get("mode", "seeding")),
                                    fast_recovery=True,
                                )
                                local_is_main_driver = bool(worker_driver) and (worker_driver is app.driver)
                            except Exception as restart_exc:
                                write_log(f"[WARN] Browser recovery failed at row {row}: {restart_exc}")
                            else:
                                if row not in retry_once_by_row:
                                    retry_once_by_row.add(row)
                                    ui_call(
                                        ui_add_log,
                                        row,
                                        "OK",
                                        "RETRY",
                                        f"{log_block_name}: Đã khôi phục phiên Chrome, thử lại dòng này 1 lần...",
                                        "ok",
                                    )
                                    pending_indexes.insert(0, idx)
                                    continue
                                write_log(f"[WARN] Row {row} already retried once after browser recovery; mark failed.")
                        write_log(f"{log_block_name} row {row} ERROR: {e}")
                        _mark_row_failed(row)
                        _finish_row_fail(log_block_name, row, str(e), eta_text, issue_columns=issue_columns)
                        if is_scan_mode and idx_drive:
                            try:
                                safe_sheet_write(
                                    lambda: local_worksheet.update_acell(f"{col_drive_letter}{row}", "0"),
                                    op_desc=f"update_result_0_row_{row}",
                                )
                            except Exception:
                                pass
                        if (not is_scan_mode) and idx_drive:
                            try:
                                safe_sheet_write(
                                    lambda: local_worksheet.update_acell(f"{col_drive_letter}{row}", f"ERR: {str(e)[:80]}"),
                                    op_desc=f"update_drive_err_row_{row}",
                                )
                            except Exception:
                                pass
                        if (not is_scan_mode) and idx_content:
                            try:
                                safe_sheet_write(
                                    lambda: local_worksheet.update_acell(f"{col_content_letter}{row}", f"ERR_CAPTION: {str(e)[:80]}"),
                                    op_desc=f"update_caption_err_row_{row}",
                                )
                            except Exception:
                                pass
            finally:
                if (not local_is_main_driver) and worker_driver:
                    try:
                        worker_driver.quit()
                    except Exception:
                        pass

        # Run all configured posts in parallel (no fixed upper limit).
        worker_total = max(1, len(prepared_blocks))
        if len(prepared_blocks) > 1:
            with ThreadPoolExecutor(max_workers=worker_total) as ex:
                futures = [ex.submit(_run_block, b) for b in prepared_blocks]
                for fu in as_completed(futures):
                    fu.result()
        elif prepared_blocks:
            _run_block(prepared_blocks[0])

        if history_ready:
            set_error_rows_for_sheet(
                sheet_url,
                sheet_name=sheet_name,
                rows=tracked_error_rows,
                details=tracked_error_details,
            )
            write_log(f"[INFO] Error history saved: {len(tracked_error_rows)} row(s) pending")
            if hasattr(app, "refresh_error_history_ui"):
                ui_call(app.refresh_error_history_ui)

        # Force-sync UI error panel from the same runtime source used for final summary,
        # so "Lỗi theo link Sheet" and "Failed" never diverge after run completion.
        try:
            if hasattr(app, "live_error_details"):
                app.live_error_details = {int(k): str(v) for k, v in tracked_error_details.items()}
            if hasattr(app, "_render_error_history_card"):
                ui_call(app._render_error_history_card, dict(tracked_error_details))
        except Exception:
            pass

        final_fail_count = fail_count
        fail_count = final_fail_count
        ui_call(ui_update_summary, processed_count, target_total, success_count, final_fail_count, "---", unavailable_count)
        ui_call(ui_set_done)
        if can_show_native_dialog():
            try:
                stopped_early = (not getattr(app, "is_running", True)) and (processed_count < target_total)
                summary_text = (
                    f"Đã xử lý: {processed_count}/{target_total}\n"
                    f"Success: {success_count}\n"
                    f"Failed: {final_fail_count}\n"
                    f"Unavailable: {unavailable_count}"
                )
                if stopped_early:
                    if hasattr(app, "show_completion_popup"):
                        ui_call(app.show_completion_popup, "Đã dừng", f"Tiến trình đã dừng giữa chừng.\n\n{summary_text}", "warn")
                    else:
                        ui_call(messagebox.showwarning, "Đã dừng", f"Tiến trình đã dừng giữa chừng.\n\n{summary_text}")
                elif fail_count > 0 or unavailable_count > 0:
                    if hasattr(app, "show_completion_popup"):
                        ui_call(app.show_completion_popup, "Hoàn tất (có lỗi)", summary_text, "warn")
                    else:
                        ui_call(messagebox.showwarning, "Hoàn tất (có lỗi)", summary_text)
                else:
                    if hasattr(app, "show_completion_popup"):
                        ui_call(app.show_completion_popup, "Hoàn tất", summary_text, "info")
                    else:
                        ui_call(messagebox.showinfo, "Hoàn tất", summary_text)
            except Exception:
                pass
        ui_call(app.set_inputs_enabled, True)

    except Exception as e:
        error_text = str(e).strip() or "Unknown error"
        write_log(f"FATAL: {error_text}")
        if history_ready:
            set_error_rows_for_sheet(
                sheet_url,
                sheet_name=sheet_name,
                rows=tracked_error_rows,
                details=tracked_error_details,
            )
            if hasattr(app, "refresh_error_history_ui"):
                ui_call(app.refresh_error_history_ui)
        try:
            ui_call(ui_set_status, "LỖI HỆ THỐNG", "#ef4444")
            ui_call(ui_set_detail, error_text)
        except Exception:
            pass
        if hasattr(app, "_job_store") and isinstance(getattr(app, "_job_store", None), dict):
            try:
                app._job_store["error"] = error_text
            except Exception:
                pass
        ui_call(app.set_inputs_enabled, True)
        if can_show_native_dialog():
            ui_call(messagebox.showerror, "Lỗi hệ thống", error_text)
        else:
            print(f"FATAL: {error_text}")
            raise
    finally:
        if app.driver:
            try:
                app.driver.quit()
            except:
                pass

# ================= RUN =================
def run_headless(drive_id: str, sheet_url: str, sheet_name: str, start_line: int = 4, force_run_all: bool = False, browser_port: int = 9223):
    class _Flag:
        def __init__(self, v=False):
            self._v = bool(v)

        def get(self):
            return self._v

    class _LabelStub:
        def config(self, **kwargs):
            # no-op for headless
            return

    class _RootStub:
        def update(self):
            return

    class _AppStub:
        def __init__(self):
            self.is_running = True
            self.driver = None
            self.start_line = start_line
            self.force_run_all = _Flag(force_run_all)
            self.only_run_error_rows = _Flag(False)
            self.progress = {"value": 0}
            self.label_detail = _LabelStub()
            self.label_status = _LabelStub()
            self.root = _RootStub()

        def set_inputs_enabled(self, enabled: bool):
            return

    app = _AppStub()

    # Run in a separate thread to avoid blocking callers
    def _target():
        try:
            main_logic(app, drive_id, sheet_url, sheet_name, start_line=start_line, browser_port=browser_port)
        except Exception as e:
            write_log(f"run_headless ERROR: {e}")

    t = threading.Thread(target=_target, daemon=True)
    t.start()
    return t

def launch_chrome_for_login(browser_port: int = 9223, profile_path: str | None = None) -> tuple[bool, str]:
    """
    Launch Chrome on specified port for user to login to sites.
    Returns (success: bool, profile_info: str)
    """
    try:
        replaced_headless = False
        has_desktop_session = (
            os.name == "nt"
            or sys.platform == "darwin"
            or bool(os.environ.get("DISPLAY") or os.environ.get("WAYLAND_DISPLAY"))
        )

        def is_port_open(port: int, timeout_sec: float = 0.8) -> bool:
            try:
                with socket.create_connection(("127.0.0.1", port), timeout=timeout_sec):
                    return True
            except Exception:
                return False

        def get_debugger_version(port: int) -> dict:
            endpoint = f"http://127.0.0.1:{port}/json/version"
            try:
                if requests is not None:
                    resp = requests.get(endpoint, timeout=2)
                    if resp.ok:
                        return resp.json() or {}
                    return {}
                import urllib.request
                with urllib.request.urlopen(endpoint, timeout=2) as resp:
                    return json.loads(resp.read().decode("utf-8", errors="ignore") or "{}")
            except Exception as e:
                write_log(f"[WARN] Failed to read debugger version on port {port}: {e}")
                return {}

        def debugger_is_headless(port: int) -> bool:
            meta = get_debugger_version(port)
            browser_text = str(meta.get("Browser") or "")
            user_agent = str(meta.get("User-Agent") or "")
            return "HeadlessChrome" in browser_text or "HeadlessChrome" in user_agent

        def find_listener_pid(port: int) -> int | None:
            try:
                if os.name == "nt":
                    result = subprocess.run(
                        ["netstat", "-ano", "-p", "tcp"],
                        capture_output=True,
                        text=True,
                        timeout=6,
                        check=False,
                    )
                    for raw in (result.stdout or "").splitlines():
                        line = raw.strip()
                        if not line.upper().startswith("TCP"):
                            continue
                        parts = line.split()
                        if len(parts) < 5:
                            continue
                        local_addr = parts[1]
                        state = parts[3].upper()
                        pid_text = parts[4]
                        if state != "LISTENING":
                            continue
                        local_port = local_addr.rsplit(":", 1)[-1].strip("[]")
                        if local_port == str(port) and pid_text.isdigit():
                            return int(pid_text)
                    return None
                result = subprocess.run(
                    ["lsof", "-nPi", f"TCP:{port}", "-sTCP:LISTEN"],
                    capture_output=True,
                    text=True,
                    timeout=6,
                    check=False,
                )
                for raw in (result.stdout or "").splitlines()[1:]:
                    parts = raw.split()
                    if len(parts) > 1 and parts[1].isdigit():
                        return int(parts[1])
            except Exception as e:
                write_log(f"[WARN] Failed to resolve pid for port {port}: {e}")
            return None

        def terminate_process_tree(pid: int) -> bool:
            if pid <= 0:
                return False
            try:
                if os.name == "nt":
                    result = subprocess.run(
                        ["taskkill", "/PID", str(pid), "/F", "/T"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        timeout=10,
                        check=False,
                    )
                    return result.returncode == 0
                os.kill(pid, signal.SIGTERM)
                return True
            except Exception as e:
                write_log(f"[WARN] Failed to terminate pid {pid}: {e}")
                return False

        def wait_for_port_closed(port: int, timeout_sec: float = 8.0) -> bool:
            deadline = time.time() + timeout_sec
            while time.time() < deadline:
                if not is_port_open(port, timeout_sec=0.4):
                    return True
                time.sleep(0.25)
            return not is_port_open(port, timeout_sec=0.4)

        browser_resolve_issues: list[str] = []
        def find_chrome_binary() -> str | None:
            nonlocal browser_resolve_issues
            chrome_path, issues = find_compatible_browser_binary()
            browser_resolve_issues = list(issues or [])
            for item in browser_resolve_issues:
                write_log(f"[WARN] Skip browser binary: {item}")
            return chrome_path

        def open_visible_window(chrome_path: str, profile: str):
            args_visible = [
                f"--remote-debugging-port={browser_port}",
                "--remote-debugging-address=127.0.0.1",
                f"--user-data-dir={profile}",
                "--new-window",
                "--window-size=1200,900",
                "about:blank",
            ]
            if sys.platform == "darwin":
                marker = "/Contents/MacOS/"
                app_bundle = None
                if marker in chrome_path:
                    app_bundle = chrome_path.split(marker, 1)[0]
                open_cmd = ["open", "-na", app_bundle or "Google Chrome", "--args", *args_visible]
                try:
                    subprocess.Popen(open_cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    return
                except Exception as e:
                    write_log(f"[WARN] open -na failed on macOS, fallback to binary: {e}")
            subprocess.Popen([chrome_path, *args_visible], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        def open_tab_in_existing_debugger(target_url: str = "about:blank") -> bool:
            try:
                endpoint = f"http://127.0.0.1:{browser_port}/json/new?{quote(target_url, safe=':/?=&,%#')}"
                if requests is not None:
                    resp = requests.put(endpoint, timeout=2)
                    return bool(resp.ok)
                import urllib.request
                req = urllib.request.Request(endpoint, method="PUT")
                with urllib.request.urlopen(req, timeout=2):
                    return True
            except Exception as e:
                write_log(f"[WARN] Failed to open tab via debugger port {browser_port}: {e}")
                return False

        def focus_existing_browser_window(title_hint: str | None = None) -> bool:
            if os.name != "nt":
                return False
            title_hint = str(title_hint or "").strip()
            title_expr = title_hint.replace("'", "''")
            script = """
$ws = New-Object -ComObject WScript.Shell
$targets = @()
if ('__TITLE__'.Length -gt 0) {
  $targets += '__TITLE__'
}
$targets += @('Google Chrome', 'Chrome', 'Microsoft Edge', 'Edge')
foreach ($t in $targets) {
  try {
    if ($ws.AppActivate($t)) { exit 0 }
  } catch {}
}
exit 1
"""
            script = script.replace("__TITLE__", title_expr)
            try:
                result = subprocess.run(
                    ["powershell", "-NoProfile", "-Command", script],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=4,
                    check=False,
                )
                return result.returncode == 0
            except Exception as e:
                write_log(f"[WARN] Failed to focus existing browser window: {e}")
                return False

        def open_focus_marker_tab() -> tuple[bool, str]:
            marker_title = f"Tool Evidence Chrome {browser_port}"
            marker_html = (
                "<html><head>"
                f"<title>{marker_title}</title>"
                "</head><body style='font-family:Segoe UI,Arial,sans-serif;"
                "background:#0f172a;color:#dbeafe;display:grid;place-items:center;"
                "height:100vh;margin:0'>"
                f"<div>Chrome debugger port {browser_port}</div>"
                "</body></html>"
            )
            target_url = "data:text/html," + marker_html
            opened = open_tab_in_existing_debugger(target_url)
            if opened:
                time.sleep(0.35)
            focused = focus_existing_browser_window(marker_title)
            return opened, marker_title if focused else marker_title

        def focus_browser_window_macos() -> bool:
            if sys.platform != "darwin":
                return False
            script = (
                'tell application "Google Chrome" to activate\n'
                'tell application "System Events" to tell process "Google Chrome" to set frontmost to true'
            )
            try:
                result = subprocess.run(
                    ["osascript", "-e", script],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=4,
                    check=False,
                )
                return result.returncode == 0
            except Exception as e:
                write_log(f"[WARN] Failed to activate Chrome on macOS: {e}")
                return False

        chrome_path = find_chrome_binary()
        if not chrome_path:
            if browser_resolve_issues:
                return (
                    False,
                    "Không tìm được browser tương thích với macOS hiện tại. "
                    f"Chi tiết: {browser_resolve_issues[0]}",
                )
            return False, "Chrome not found (missing chrome.exe)"

        if not has_desktop_session:
            return (
                False,
                "Môi trường web deploy không có giao diện desktop để mở Chrome. Hãy dùng local web/app nếu cần Chrome 9223 trên máy của bạn.",
            )

        profile = _resolve_writable_profile_dir(
            profile_path or LOCAL_PROFILE_PATH,
            browser_port=browser_port,
            log_prefix="Launch: ",
        )
        local_profile = _normalize_profile_dir(LOCAL_PROFILE_PATH)
        os.makedirs(profile, exist_ok=True)
        if os.path.abspath(profile) != os.path.abspath(local_profile):
            try:
                if not os.path.isdir(os.path.join(profile, "Default")) and os.path.isdir(local_profile):
                    shutil.copytree(
                        local_profile,
                        profile,
                        dirs_exist_ok=True,
                        ignore=shutil.ignore_patterns(
                            "Cache",
                            "Code Cache",
                            "GPUCache",
                            "GrShaderCache",
                            "ShaderCache",
                            "Crashpad",
                            "Singleton*",
                            "lockfile",
                            "*.tmp",
                        ),
                    )
                    write_log(f"[INFO] Seeded login profile '{profile}' from LOCAL profile")
            except Exception as e:
                write_log(f"[WARN] Failed to seed login profile '{profile}': {e}")
        # Cleanup stale lock files that can block opening profile window.
        for fn in ["SingletonLock", "SingletonCookie", "SingletonSocket", "lockfile"]:
            try:
                p = os.path.join(profile, fn)
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass

        # If debug port is already alive, still force opening a visible Chrome window.
        if is_port_open(browser_port):
            if debugger_is_headless(browser_port):
                pid = find_listener_pid(browser_port)
                if pid is None:
                    return False, f"Port {browser_port} đang bị HeadlessChrome chiếm nhưng không tìm được PID để mở lại window"
                if not terminate_process_tree(pid):
                    return False, f"Port {browser_port} đang bị HeadlessChrome chiếm và không thể dừng process {pid}"
                if not wait_for_port_closed(browser_port):
                    return False, f"Port {browser_port} vẫn còn bị chiếm sau khi dừng HeadlessChrome (PID {pid})"
                replaced_headless = True
                write_log(f"[INFO] Replaced headless Chrome on port {browser_port} (pid={pid}) with visible window request")
            else:
                opened_tab, marker_title = open_focus_marker_tab()
                focused = focus_existing_browser_window(marker_title)
                try:
                    open_visible_window(chrome_path, profile)
                except Exception as e:
                    write_log(f"[WARN] Port {browser_port} is open but failed to open visible window: {e}")
                    if os.name == "nt":
                        try:
                            fallback_cmd = (
                                f'start "" "{chrome_path}" --remote-debugging-port={browser_port} '
                                f'--remote-debugging-address=127.0.0.1 --user-data-dir="{profile}" --new-window about:blank'
                            )
                            subprocess.Popen(["cmd", "/c", fallback_cmd], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        except Exception:
                            pass
                if not focused:
                    focused = focus_existing_browser_window(marker_title)
                write_log(
                    f"[INFO] Chrome debug port {browser_port} already active. "
                    f"opened_tab={opened_tab}, focused={focused}, marker_title={marker_title}"
                )
                focus_browser_window_macos()
                if opened_tab or focused:
                    return True, f"Port {browser_port} already active; opened existing Chrome session"
                return True, f"Port {browser_port} already active; check the existing Chrome window"

        args = [
            chrome_path,
            f"--remote-debugging-port={browser_port}",
            f"--user-data-dir={profile}",
            "--new-window",
            f"--window-size={CAPTURE_WINDOW_SIZE}",
            "--force-device-scale-factor=1",
            "--high-dpi-support=1",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-background-networking",
            "--disable-sync",
            "about:blank",
        ]

        write_log(f"[INFO] Launch Chrome cmd: {args[0]}")
        subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        for _ in range(12):
            if is_port_open(browser_port):
                write_log(f"[INFO] Chrome launched on port {browser_port} for login. Profile: {profile}")
                focus_browser_window_macos()
                if replaced_headless:
                    return True, f"Port {browser_port} đã được mở lại thành window thật. Profile: {os.path.basename(profile)}"
                return True, f"Port {browser_port}, Profile: {os.path.basename(profile)}"
            time.sleep(0.5)

        # Fallback launch via shell alias (some machines only resolve chrome via App Paths/PATH).
        if os.name == "nt":
            fallback_cmd = (
                f'start "" chrome --remote-debugging-port={browser_port} '
                f'--user-data-dir="{profile}" --new-window about:blank'
            )
            subprocess.Popen(["cmd", "/c", fallback_cmd], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            for _ in range(8):
                if is_port_open(browser_port):
                    write_log(f"[INFO] Chrome launched via fallback shell command on port {browser_port}.")
                    focus_browser_window_macos()
                    if replaced_headless:
                        return True, f"Port {browser_port} đã được mở lại thành window thật. Profile: {os.path.basename(profile)}"
                    return True, f"Port {browser_port} (fallback), Profile: {os.path.basename(profile)}"
                time.sleep(0.5)

        write_log(f"[ERROR] Chrome started but port {browser_port} is not reachable after direct + fallback launch.")
        return False, f"Chrome did not expose debug port {browser_port}"
    except Exception as e:
        write_log(f"[ERROR] Failed to launch Chrome for login: {e}")
        return False, str(e)

if __name__ == "__main__":
    if tk is None:
        raise RuntimeError("Tkinter is unavailable in this environment. Use web_ui.py for web mode.")
    root = tk.Tk()
    app = ProgressApp(root)
    root.mainloop()


